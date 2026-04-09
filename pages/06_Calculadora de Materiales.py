import streamlit as st
import pandas as pd
import numpy as np
import requests
from bs4 import BeautifulSoup
import re, math, io, datetime, json, pathlib
import plotly.graph_objects as go
from io import BytesIO

import json
import os

STATE_FILE = "konte_state.json"

def save_state():
    try:
        with open(STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(st.session_state.kc_rows, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def load_state():
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r', encoding='utf-8') as f:
                rows = json.load(f)
                if isinstance(rows, list):
                    st.session_state.kc_rows = rows
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Konte — Calculadora de Materiales", layout="wide", page_icon="")

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL CONFIGURATION (LANGUAGE & UNITS)
# ─────────────────────────────────────────────────────────────────────────────
if "idioma" not in st.session_state: st.session_state["idioma"] = "Español"
if "unidades" not in st.session_state: st.session_state["unidades"] = "Métrico"
_lang = st.session_state["idioma"]
_unit = st.session_state["unidades"]

def _t(es, en): return en if _lang == "English" else es
def _u(metric, imperial): return imperial if _unit == "Imperial" else metric

# ─────────────────────────────────────────────────────────────────────────────
# REGIONAL TERMINOLOGY per norm
# ─────────────────────────────────────────────────────────────────────────────
REGIONAL_RAW = {
    "NSR-10 (Colombia)": {
        "pais": "Colombia", "moneda": "COP $", "moneda_code": "COP", "imperial": False,
        "cemento": ("Cemento gris (Argos/Holcim)", "Grey Cement (Argos/Holcim)"), "arena": ("Arena de peña (m³)", "Sand (m³)"),
        "grava": ("Triturado / Gravilla (m³)", "Gravel / Crushed stone (m³)"), "agua": ("Agua potable (lt)", "Potable water (lt)"),
        "varilla": ("Varilla corrugada", "Deformed rebar"), "bloque": ("Bloque de concreto", "Concrete block"),
        "ladrillo": ("Ladrillo tolete", "Clay brick"), "mortero": ("Mortero de pega", "Masonry mortar"),
        "pañete": ("Pañete / Revoque", "Stucco / Plaster"), "placa": ("Placa / Losa", "Concrete slab"),
        "columna": ("Columna", "Column"), "viga": ("Viga de amarre", "Tie beam"),
        "zapata": ("Zapata aislada", "Isolated footing"), "cimiento": ("Cimiento corrido", "Strip footing"),
        "pintura": ("Pintura de vinilo", "Latex paint"), "ceramica": ("Cerámica", "Ceramic tile"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Drywall / Panel yeso", "Drywall / Gypsum board"),
        "cubierta": ("Cubierta en lámina / teja", "Roofing (metal/tile)"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Concreto", "Concrete"), "hierro": ("Acero de refuerzo", "Steel rebar"),
        "peso_bolsa": 50, "rendimiento_pintura": 25,
    },
    "ACI 318-25 (EE.UU.)": {
        "pais": "USA", "moneda": "USD $", "moneda_code": "USD", "imperial": True,
        "cemento": ("Cemento Portland (Tipo I/II)", "Portland Cement (Type I/II)"), "arena": ("Agregado fino", "Fine aggregate"),
        "grava": ("Agregado grueso", "Coarse aggregate"), "agua": ("Agua", "Water"),
        "varilla": ("Barra corrugada", "Rebar / Deformed bar"), "bloque": ("Bloque de concreto", "Concrete block (CMU)"),
        "ladrillo": ("Ladrillo", "Brick"), "mortero": ("Mezcla de mortero", "Mortar mix"),
        "pañete": ("Estuco", "Stucco"), "placa": ("Losa de concreto", "Concrete slab"),
        "columna": ("Columna", "Column"), "viga": ("Viga", "Beam"),
        "zapata": ("Zapata aislada", "Isolated footing"), "cimiento": ("Cimiento corrido", "Strip footing"),
        "pintura": ("Pintura interior/exterior", "Interior/Exterior paint"), "ceramica": ("Baldosa cerámica", "Ceramic tile"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Drywall / Panel de yeso", "Drywall / Gypsum board"),
        "cubierta": ("Cubierta (lámina/teja)", "Roofing (metal/tile)"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Concreto", "Concrete"), "hierro": ("Acero de refuerzo", "Steel rebar"),
        "peso_bolsa": 42.6, "rendimiento_pintura": 300,
    },
    "ACI 318-19 (EE.UU.)": {
        "pais": "USA", "moneda": "USD $", "moneda_code": "USD", "imperial": True,
        "cemento": ("Cemento Portland (Tipo I/II)", "Portland Cement (Type I/II)"), "arena": ("Agregado fino", "Fine aggregate"),
        "grava": ("Agregado grueso", "Coarse aggregate"), "agua": ("Agua", "Water"),
        "varilla": ("Barra corrugada", "Rebar / Deformed bar"), "bloque": ("Bloque CMU", "CMU Block"),
        "ladrillo": ("Ladrillo", "Brick"), "mortero": ("Mortero", "Mortar"),
        "pañete": ("Estuco", "Stucco"), "placa": ("Losa", "Slab"),
        "columna": ("Columna", "Column"), "viga": ("Viga", "Beam"),
        "zapata": ("Zapata", "Footing"), "cimiento": ("Cimiento corrido", "Strip footing"),
        "pintura": ("Pintura", "Paint"), "ceramica": ("Baldosa", "Tile"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Drywall", "Drywall"),
        "cubierta": ("Cubierta", "Roofing"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Concreto", "Concrete"), "hierro": ("Acero de refuerzo", "Steel rebar"),
        "peso_bolsa": 42.6, "rendimiento_pintura": 300,
    },
    "ACI 318-14 (EE.UU.)": {
        "pais": "USA", "moneda": "USD $", "moneda_code": "USD", "imperial": True,
        "cemento": ("Cemento Portland", "Portland Cement"), "arena": ("Arena (agregado fino)", "Sand (fine agg.)"),
        "grava": ("Grava (agregado grueso)", "Gravel (coarse agg.)"), "agua": ("Agua", "Water"),
        "varilla": ("Varilla", "Rebar"), "bloque": ("Bloque CMU", "CMU Block"),
        "ladrillo": ("Ladrillo", "Brick"), "mortero": ("Mortero", "Mortar"),
        "pañete": ("Estuco", "Stucco"), "placa": ("Losa", "Slab"),
        "columna": ("Columna", "Column"), "viga": ("Viga", "Beam"),
        "zapata": ("Zapata", "Footing"), "cimiento": ("Cimiento corrido", "Strip footing"),
        "pintura": ("Pintura", "Paint"), "ceramica": ("Baldosa", "Tile"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Drywall", "Drywall"),
        "cubierta": ("Cubierta", "Roofing"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Concreto", "Concrete"), "hierro": ("Acero de refuerzo", "Steel rebar"),
        "peso_bolsa": 42.6, "rendimiento_pintura": 300,
    },
    "NEC-SE-HM (Ecuador)": {
        "pais": "Ecuador", "moneda": "USD $", "moneda_code": "USD", "imperial": False,
        "cemento": ("Cemento Holcim / Chimborazo", "Holcim / Chimborazo Cement"), "arena": ("Arena lavada (m³)", "Washed sand (m³)"),
        "grava": ("Ripio / Piedra triturada (m³)", "Crushed stone / Gravel (m³)"), "agua": ("Agua (lt)", "Water (lt)"),
        "varilla": ("Varilla corrugada (Adelca)", "Deformed rebar (Adelca)"), "bloque": ("Bloque de hormigón", "Concrete block"),
        "ladrillo": ("Ladrillo artesanal", "Artisanal brick"), "mortero": ("Mortero de pega", "Masonry mortar"),
        "pañete": ("Enlucido / Revoque", "Plaster / Stucco"), "placa": ("Losa de hormigón", "Concrete slab"),
        "columna": ("Columna", "Column"), "viga": ("Cadena / Viga", "Tie beam / Beam"),
        "zapata": ("Zapata aislada", "Isolated footing"), "cimiento": ("Cimiento corrido", "Strip footing"),
        "pintura": ("Pintura látex", "Latex paint"), "ceramica": ("Cerámica", "Ceramic tile"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Gypsum / Gyplac", "Gypsum board"),
        "cubierta": ("Cubierta (zinc / teja)", "Roofing (zinc/tile)"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Hormigón", "Concrete"), "hierro": ("Acero de refuerzo", "Steel rebar"),
        "peso_bolsa": 50, "rendimiento_pintura": 25,
    },
    "E.060 (Perú)": {
        "pais": "Perú", "moneda": "PEN S/", "moneda_code": "PEN", "imperial": False,
        "cemento": ("Cemento Sol / Pacasmayo", "Sol / Pacasmayo Cement"), "arena": ("Arena gruesa (m³)", "Coarse sand (m³)"),
        "grava": ("Piedra chancada (m³)", "Crushed stone (m³)"), "agua": ("Agua (lt)", "Water (lt)"),
        "varilla": ("Fierro corrugado", "Deformed rebar"), "bloque": ("Ladrillo King Kong", "King Kong Brick"),
        "ladrillo": ("Ladrillo de arcilla", "Clay brick"), "mortero": ("Mezcla de pega", "Masonry mortar"),
        "pañete": ("Tarrajeo / Revoque", "Plaster / Stucco"), "placa": ("Losa aligerada / maciza", "Ribbed / Solid slab"),
        "columna": ("Columna", "Column"), "viga": ("Viga", "Beam"),
        "zapata": ("Zapata aislada", "Isolated footing"), "cimiento": ("Cimiento corrido", "Strip footing"),
        "pintura": ("Pintura látex", "Latex paint"), "ceramica": ("Cerámico", "Ceramic tile"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Drywall / Placa de yeso", "Drywall / Gypsum board"),
        "cubierta": ("Cobertura / Calamina", "Roofing / Corrugated metal"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Concreto", "Concrete"), "hierro": ("Acero de refuerzo", "Steel rebar"),
        "peso_bolsa": 42.5, "rendimiento_pintura": 25,
    },
    "NTC-EM (México)": {
        "pais": "México", "moneda": "MXN $", "moneda_code": "MXN", "imperial": False,
        "cemento": ("Cemento Cemex / Cruz Azul", "Cemex / Cruz Azul Cement"), "arena": ("Arena (m³)", "Sand (m³)"),
        "grava": ("Grava (m³)", "Gravel (m³)"), "agua": ("Agua (lt)", "Water (lt)"),
        "varilla": ("Varilla corrugada", "Deformed rebar"), "bloque": ("Block de concreto", "Concrete block"),
        "ladrillo": ("Tabique de barro", "Clay brick"), "mortero": ("Mezcla de pegado", "Masonry mortar"),
        "pañete": ("Aplanado / Firme", "Plaster"), "placa": ("Losa de concreto", "Concrete slab"),
        "columna": ("Columna", "Column"), "viga": ("Trabe / Viga", "Beam"),
        "zapata": ("Zapata aislada", "Isolated footing"), "cimiento": ("Cimentación corrida", "Strip footing"),
        "pintura": ("Pintura vinílica", "Vinyl paint"), "ceramica": ("Azulejo / Cerámica", "Ceramic tile"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Panel de yeso / Tablaroca", "Gypsum board / Tablaroca"),
        "cubierta": ("Lámina galvanizada / Teja", "Galvanized sheet / Tile"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Concreto", "Concrete"), "hierro": ("Acero de refuerzo", "Steel rebar"),
        "peso_bolsa": 50, "rendimiento_pintura": 25,
    },
    "COVENIN 1753-2006 (Venezuela)": {
        "pais": "Venezuela", "moneda": "USD $", "moneda_code": "USD", "imperial": False,
        "cemento": ("Cemento Portland", "Portland Cement"), "arena": ("Arena (m³)", "Sand (m³)"),
        "grava": ("Piedra picada / Grava (m³)", "Crushed stone / Gravel (m³)"), "agua": ("Agua (lt)", "Water (lt)"),
        "varilla": ("Cabilla corrugada", "Deformed rebar (Cabilla)"), "bloque": ("Bloque de arcilla / hormigón", "Clay / Concrete block"),
        "ladrillo": ("Ladrillo de arcilla", "Clay brick"), "mortero": ("Mortero", "Mortar"),
        "pañete": ("Friso / Revoque", "Plaster"), "placa": ("Losa de concreto", "Concrete slab"),
        "columna": ("Columna", "Column"), "viga": ("Viga", "Beam"),
        "zapata": ("Zapata aislada", "Isolated footing"), "cimiento": ("Cimiento corrido", "Strip footing"),
        "pintura": ("Pintura de caucho", "Rubber-based paint"), "ceramica": ("Cerámica", "Ceramic tile"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Panel de yeso", "Gypsum board"),
        "cubierta": ("Lámina acanalada / Teja", "Corrugated sheet / Tile"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Concreto", "Concrete"), "hierro": ("Acero de refuerzo (cabilla)", "Steel rebar"),
        "peso_bolsa": 42.5, "rendimiento_pintura": 25,
    },
    "NB 1225001-2020 (Bolivia)": {
        "pais": "Bolivia", "moneda": "BOB Bs.", "moneda_code": "BOB", "imperial": False,
        "cemento": ("Cemento Fancesa / Viacha", "Fancesa / Viacha Cement"), "arena": ("Arena (m³)", "Sand (m³)"),
        "grava": ("Piedra chancada / Grava (m³)", "Gravel / Crushed rock (m³)"), "agua": ("Agua (lt)", "Water (lt)"),
        "varilla": ("Fierro corrugado", "Deformed rebar"), "bloque": ("Block de hormigón", "Concrete block"),
        "ladrillo": ("Ladrillo de arcilla", "Clay brick"), "mortero": ("Mortero", "Mortar"),
        "pañete": ("Revoque / Enlucido", "Plaster / Stucco"), "placa": ("Losa de hormigón", "Concrete slab"),
        "columna": ("Columna", "Column"), "viga": ("Viga", "Beam"),
        "zapata": ("Zapata aislada", "Isolated footing"), "cimiento": ("Cimiento corrido", "Strip footing"),
        "pintura": ("Pintura látex", "Latex paint"), "ceramica": ("Cerámica", "Ceramic"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Panel de yeso", "Gypsum board"),
        "cubierta": ("Cubierta", "Roofing"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Hormigón", "Concrete"), "hierro": ("Acero de refuerzo", "Steel rebar"),
        "peso_bolsa": 50, "rendimiento_pintura": 25,
    },
    "CIRSOC 201-2025 (Argentina)": {
        "pais": "Argentina", "moneda": "ARS $", "moneda_code": "ARS", "imperial": False,
        "cemento": ("Cemento Loma Negra / Holcim", "Loma Negra / Holcim Cement"), "arena": ("Arena gruesa (m³)", "Coarse sand (m³)"),
        "grava": ("Piedra partida / Canto rodado (m³)", "Crushed stone / Gravel (m³)"), "agua": ("Agua (lt)", "Water (lt)"),
        "varilla": ("Hierro de construcción", "Deformed rebar"), "bloque": ("Bloque de hormigón", "Concrete block"),
        "ladrillo": ("Ladrillo cerámico hueco", "Hollow clay brick"), "mortero": ("Mortero de pegado", "Masonry mortar"),
        "pañete": ("Revoque / Jaharro", "Plaster"), "placa": ("Losa de hormigón", "Concrete slab"),
        "columna": ("Columna", "Column"), "viga": ("Viga / Encadenado", "Beam / Tie beam"),
        "zapata": ("Zapata aislada", "Isolated footing"), "cimiento": ("Platea / Cimiento corrido", "Strip footing"),
        "pintura": ("Pintura látex interior/exterior", "Latex paint"), "ceramica": ("Cerámica", "Ceramic tile"),
        "porcelanato": ("Porcelanato", "Porcelain tile"), "drywall": ("Durlock / Placa de yeso", "Gypsum board"),
        "cubierta": ("Chapa metálica / Teja", "Sheet metal / Tile roofing"), "suelo_cemento": ("Suelo-cemento", "Soil-cement"),
        "concreto": ("Hormigón armado", "Concrete"), "hierro": ("Acero de refuerzo", "Steel rebar"),
        "peso_bolsa": 50, "rendimiento_pintura": 25,
    }
}

# Auto-traducción del diccionario según el idioma seleccionado
REGIONAL = {}
for norm, dic in REGIONAL_RAW.items():
    REGIONAL[norm] = {}
    for k, v in dic.items():
        if isinstance(v, tuple):
            REGIONAL[norm][k] = _t(v[0], v[1])
        else:
            REGIONAL[norm][k] = v

# ─────────────────────────────────────────────────────────────────────────────
# CONCRETE MIX DESIGNS DATABASE (from CM-V3.0)
# Key: dosification ratio, Values: per 1 m³
# ─────────────────────────────────────────────────────────────────────────────
MIX_DESIGNS = [
    {"dos": "1:2:2",   "fc_kgcm2": 280, "fc_mpa": 27.5, "cem_kg": 420, "arena_m3": 0.67, "grava_m3": 0.67, "agua_lt": 190, "agua_bolsa": 9.45},
    {"dos": "1:2:2.5", "fc_kgcm2": 240, "fc_mpa": 23.5, "cem_kg": 380, "arena_m3": 0.60, "grava_m3": 0.76, "agua_lt": 180, "agua_bolsa": 9.45},
    {"dos": "1:2:3",   "fc_kgcm2": 226, "fc_mpa": 22.2, "cem_kg": 350, "arena_m3": 0.55, "grava_m3": 0.84, "agua_lt": 170, "agua_bolsa": 9.45},
    {"dos": "1:2:3.5", "fc_kgcm2": 210, "fc_mpa": 20.6, "cem_kg": 320, "arena_m3": 0.52, "grava_m3": 0.90, "agua_lt": 170, "agua_bolsa": 9.45},
    {"dos": "1:2:4",   "fc_kgcm2": 200, "fc_mpa": 19.6, "cem_kg": 300, "arena_m3": 0.48, "grava_m3": 0.95, "agua_lt": 158, "agua_bolsa": 9.45},
    {"dos": "1:2.5:4", "fc_kgcm2": 189, "fc_mpa": 18.5, "cem_kg": 280, "arena_m3": 0.55, "grava_m3": 0.89, "agua_lt": 158, "agua_bolsa": 9.45},
    {"dos": "1:3:3",   "fc_kgcm2": 168, "fc_mpa": 16.5, "cem_kg": 300, "arena_m3": 0.72, "grava_m3": 0.72, "agua_lt": 158, "agua_bolsa": 9.45},
    {"dos": "1:3:4",   "fc_kgcm2": 159, "fc_mpa": 15.6, "cem_kg": 260, "arena_m3": 0.63, "grava_m3": 0.83, "agua_lt": 163, "agua_bolsa": 9.45},
    {"dos": "1:3:5",   "fc_kgcm2": 140, "fc_mpa": 13.7, "cem_kg": 230, "arena_m3": 0.55, "grava_m3": 0.92, "agua_lt": 148, "agua_bolsa": 9.45},
    {"dos": "1:3:6",   "fc_kgcm2": 119, "fc_mpa": 11.7, "cem_kg": 210, "arena_m3": 0.50, "grava_m3": 1.00, "agua_lt": 143, "agua_bolsa": 9.45},
    {"dos": "1:4:7",   "fc_kgcm2": 109, "fc_mpa": 10.7, "cem_kg": 175, "arena_m3": 0.55, "grava_m3": 0.98, "agua_lt": 133, "agua_bolsa": 9.45},
    {"dos": "1:4:8",   "fc_kgcm2": 99,  "fc_mpa": 9.7,  "cem_kg": 160, "arena_m3": 0.55, "grava_m3": 1.03, "agua_lt": 125, "agua_bolsa": 9.45},
]

# ─────────────────────────────────────────────────────────────────────────────
# REBAR DATABASE (from CM-V3.0) - Weight per 6m bar
# ─────────────────────────────────────────────────────────────────────────────
VARILLAS = [
    {"nombre": "N2 - 1/4\"",  "diam_mm": 6.35,  "diam_pulg": "1/4\"",   "kg_6m": 1.49,  "kg_m": 0.248},
    {"nombre": "N3 - 3/8\"",  "diam_mm": 9.52,  "diam_pulg": "3/8\"",   "kg_6m": 3.35,  "kg_m": 0.558},
    {"nombre": "N4 - 1/2\"",  "diam_mm": 12.70, "diam_pulg": "1/2\"",   "kg_6m": 5.96,  "kg_m": 0.994},
    {"nombre": "N5 - 5/8\"",  "diam_mm": 15.87, "diam_pulg": "5/8\"",   "kg_6m": 9.25,  "kg_m": 1.542},
    {"nombre": "N5.5 - 11/16\"", "diam_mm": 17.46, "diam_pulg": "11/16\"", "kg_6m": 9.79, "kg_m": 1.632},
    {"nombre": "N6 - 3/4\"",  "diam_mm": 19.05, "diam_pulg": "3/4\"",   "kg_6m": 13.33, "kg_m": 2.222},
    {"nombre": "N7 - 7/8\"",  "diam_mm": 22.22, "diam_pulg": "7/8\"",   "kg_6m": 18.13, "kg_m": 3.022},
    {"nombre": "N8 - 1\"",    "diam_mm": 25.40, "diam_pulg": "1\"",     "kg_6m": 23.70, "kg_m": 3.950},
    {"nombre": "N10 - 1¼\"",  "diam_mm": 31.75, "diam_pulg": "1 1/4\"", "kg_6m": 37.04, "kg_m": 6.173},
    {"nombre": "N12 - 1½\"",  "diam_mm": 38.10, "diam_pulg": "1 1/2\"", "kg_6m": 53.35, "kg_m": 8.892},
    {"nombre": "N16 - 2\"",   "diam_mm": 50.80, "diam_pulg": "2\"",     "kg_6m": 94.82, "kg_m": 15.803},
]

# ─────────────────────────────────────────────────────────────────────────────
# BLOCKS DATABASE
# ─────────────────────────────────────────────────────────────────────────────
BLOQUES = {
    "Bloque 10x20x40 cm": {"ancho": 0.10, "alto": 0.20, "largo": 0.40, "und_m2": 12.5, "mortero_m3_m2": 0.015},
    "Bloque 12x20x40 cm": {"ancho": 0.12, "alto": 0.20, "largo": 0.40, "und_m2": 12.5, "mortero_m3_m2": 0.016},
    "Bloque 15x20x40 cm": {"ancho": 0.15, "alto": 0.20, "largo": 0.40, "und_m2": 12.5, "mortero_m3_m2": 0.018},
    "Bloque 20x20x40 cm": {"ancho": 0.20, "alto": 0.20, "largo": 0.40, "und_m2": 12.5, "mortero_m3_m2": 0.020},
    "Ladrillo tolete (23x11.5x6.5 cm)": {"ancho": 0.115, "alto": 0.065, "largo": 0.230, "und_m2": 58.0, "mortero_m3_m2": 0.025},
    "Ladrillo perforado (25x12x6 cm)": {"ancho": 0.12, "alto": 0.060, "largo": 0.250, "und_m2": 60.0, "mortero_m3_m2": 0.024},
}

# ─────────────────────────────────────────────────────────────────────────────
# PRICE SCRAPING URLS (like APU_Mercado)
# ─────────────────────────────────────────────────────────────────────────────
PRICE_URLS = {
    "Colombia": {
        "cemento": [
            ("Homecenter", "https://www.homecenter.com.co/homecenter-co/product/379124/cemento-uso-general-50-kg/379124/", "homecenter"),
            ("Ultracem", "https://b2c.ultracem.co/Cemento-Gris-Uso-General/cemento-gris-50-kg-uso-general-r218", "ultracem"),
        ],
        "acero": [
            ("Homecenter", "https://www.homecenter.com.co/homecenter-co/product/115431/varilla-corrugada-12-x-6-m-pdr-60/115431/", "homecenter"),
        ],
        "teja_barro": [
            ("Homecenter", "https://www.homecenter.com.co/homecenter-co/category/cat2440036/tejas-de-barro-y-coloniales/", "homecenter"),
            ("Almacen Canaima", "https://www.almacencanaima.com/productos/49/tejas-de-barro", "almacencanaima"),
            ("GYJ Nacional", "https://gyj.com.co/bogota_65/productos.html", "gyj"),
        ],
        "ceramica": [
            ("Degres Colombia", "https://www.degrescolombia.com/collections/pisos", "degrescolombia"),
            ("Homecenter", "https://www.homecenter.com.co/homecenter-co/category/cat10007/pisos", "homecenter"),
        ],
        "precios_ref": {"cemento": 34000, "arena": 80000, "grava": 90000, "acero_kg": 4800, "teja_barro": 1800, "ceramica": 38000, "bloque": 3700, "ladrillo": 1800, "pintura": 55000},
    },
    "Perú": {
        "cemento": [
            ("Promart", "https://www.promart.pe/cemento-sol-portland-tipo-1-42.5-kg-12821/p", "promart"),
            ("Sodimac", "https://www.sodimac.com.pe/sodimac-pe/product/20658/cemento-portland-tipo-1-sol-42.5-kg/20658/", "sodimac"),
        ],
        "acero": [
            ("Promart", "https://www.promart.pe/fierro-corrugado-1-2-x-9-m-12995/p", "promart"),
        ],
        "precios_ref": {"cemento": 28, "arena": 65, "grava": 70, "acero_kg": 4.2},
    },
    "México": {
        "cemento": [
            ("Sodimac MX", "https://www.sodimac.com.mx/sodimac-mx/product/432098/cemento-gris-holcim-apasco-50-kg/432098/", "sodimac"),
        ],
        "acero": [
            ("Home Depot MX", "https://www.homedepot.com.mx/materiales-de-construccion/acero-de-refuerzo/varillas-y-alambres/varilla-corrugada-12-x-12-m-136932", "homedepot_mx"),
        ],
        "precios_ref": {"cemento": 220, "arena": 400, "grava": 450, "acero_kg": 28},
    },
    "Ecuador": {
        "cemento": [
            ("Disensa", "https://www.disensa.com.ec/cemento-holcim-fuerte-ecoplanet-tipo-gu-50-kg/p", "vtex"),
        ],
        "acero": [
            ("Disensa", "https://www.disensa.com.ec/varilla-corrugada-de-12mm-x-12m-as42-adelca/p", "vtex"),
        ],
        "precios_ref": {"cemento": 8.5, "arena": 18, "grava": 22, "acero_kg": 1.4},
    },
    "USA": {
        "precios_ref": {
            "cemento": 12.5,
            "arena": 45,
            "grava": 55,
            "acero_kg": 1.2,
            "teja_barro": 1.5,
            "ceramica": 15,
            "bloque": 1.8,
            "ladrillo": 0.8,
            "pintura": 35
        }
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# SCRAPING HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _clean_price(text):
    nums = re.findall(r"\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?", text)
    if not nums: return None
    for n in nums:
        val = float(n.replace(".", "").replace(",", ".")) if "," in n and n.index(",") > len(n)-4 else float(n.replace(",","").replace(".",""))
        if val > 0.5: return val
    return None

@st.cache_data(ttl=3600*12, show_spinner=False)
def _scrape_price(url, platform="sodimac"):
    try:
        h = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        r = requests.get(url, headers=h, timeout=8)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        selectors = {
            "homecenter": ["[itemprop='price']", ".price-0", ".price"],
            "sodimac": ["[itemprop='price']", ".price-0", ".price-box .price"],
            "vtex": ["span[class*='currencyContainer']", ".vtex-product-price-1-x-currencyContainer"],
            "promart": [".bestPrice", ".price"],
            "homedepot_mx": ["[itemprop='price']", ".price"],
            "ultracem": ["[data-price-amount]", ".price"],
            "degrescolombia": [".price", ".product-price", "[class*='price']"],
            "almacencanaima": [".precio", ".price", ".product__price", "[class*='precio']"],
            "gyj": [".precio", ".price_tag", "td.precio", ".articulo-precio", "[class*='prec']"],
        }
        for sel in selectors.get(platform, [".price"]):
            el = soup.select_one(sel)
            if el:
                v = _clean_price(el.get("content", el.text))
                if v: return v
        meta = soup.find("meta", itemprop="price")
        if meta and meta.get("content"):
            return float(meta["content"])
    except Exception as e:
        # No mostrar error al usuario, solo retornar None
        pass
    return None

def get_live_prices(pais, ref_prices):
    urls = PRICE_URLS.get(pais, {})
    result = dict(ref_prices)
    
    if "cemento" in urls:
        vals = []
        for name, url, plat in urls["cemento"]:
            v = _scrape_price(url, plat)
            if v and v > 0: vals.append(v)
        if vals: result["cemento"] = sum(vals)/len(vals)
    
    if "acero" in urls:
        vals = []
        for name, url, plat in urls["acero"]:
            v = _scrape_price(url, plat)
            if v and v > 0: vals.append(v)
        if vals:
            # Convert unit price to kg price
            len_var = 6 if pais == "Colombia" else 9 if pais == "Perú" else 12
            result["acero_kg"] = (sum(vals)/len(vals)) / (0.994 * len_var)

    if "teja_barro" in urls:
        vals = []
        for name, url, plat in urls["teja_barro"]:
            v = _scrape_price(url, plat)
            if v and v > 0: vals.append(v)
        if vals:
            result["teja_barro"] = sum(vals)/len(vals)

    if "ceramica" in urls:
        vals = []
        for name, url, plat in urls["ceramica"]:
            v = _scrape_price(url, plat)
            if v and v > 0: vals.append(v)
        if vals:
            result["ceramica"] = sum(vals)/len(vals)

    return result


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: EXCEL EXPORT (RESUMEN)
# ─────────────────────────────────────────────────────────────────────────────
def build_excel_resumen(rows, precios, R):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        wb = writer.book
        hdr   = wb.add_format({"bold":True,"bg_color":"#1e3a5f","font_color":"white","border":1,"align":"center"})
        cell  = wb.add_format({"border":1})
        num   = wb.add_format({"border":1,"num_format":"#,##0.00"})
        money = wb.add_format({"border":1,"num_format":"#,##0.00","bg_color":"#e8f5e9"})
        title_fmt = wb.add_format({"bold":True,"font_size":14,"font_color":"#1e3a5f"})
        ws = wb.add_worksheet("Resumen Konte")
        ws.set_column("A:A",35); ws.set_column("B:B",18); ws.set_column("C:C",18); ws.set_column("D:D",20); ws.set_column("E:E",20)
        ws.write("A1", "KONTE — Calculadora de Materiales", title_fmt)
        ws.write("A2", f"Generado: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.write("A3", f"Norma: {R.get('_norma','---')}  |  Pais: {R.get('pais','---')}  |  Moneda: {R.get('moneda','---')}")
        cols = ["Elemento / Material","Unidad","Cantidad",f"P. Unit. ({R.get('moneda','$')})",f"Total ({R.get('moneda','$')})"]
        for ci,c in enumerate(cols): ws.write(5,ci,c,hdr)
        total_costo=0
        for ri,row in enumerate(rows, start=6):
            ws.write(ri,0,row.get("elemento",""),cell)
            ws.write(ri,1,row.get("unidad",""),cell)
            ws.write_number(ri,2,float(row.get("cant",0)),num)
            p_unit=float(row.get("precio",0)); subtotal=float(row.get("cant",0))*p_unit; total_costo+=subtotal
            ws.write_number(ri,3,p_unit,money); ws.write_number(ri,4,subtotal,money)
        last=len(rows)+6; bold_total=wb.add_format({"bold":True,"bg_color":"#1e3a5f","font_color":"white","num_format":"#,##0.00","border":1})
        ws.write(last,3,"TOTAL",hdr); ws.write_number(last,4,total_costo,bold_total)
        ws2=wb.add_worksheet("Precios Referencia")
        ws2.write(0,0,"Material",hdr); ws2.write(0,1,"Precio",hdr); ws2.write(0,2,"Moneda",hdr)
        for pi,(k,v) in enumerate(precios.items(),start=1):
            ws2.write(pi,0,k,cell); ws2.write_number(pi,1,float(v),num); ws2.write(pi,2,R.get("moneda","$"),cell)
    output.seek(0)
    return output

# ─────────────────────────────────────────────────────────────────────────────
# HELPER: EXCEL EXPORT (PRESUPUESTO)
# ─────────────────────────────────────────────────────────────────────────────
def build_excel_presupuesto(df_pres, cliente, proyecto):
    # Try to load template, if not found create a basic workbook
    template_path = None
    # Search for template in common locations
    possible_paths = [
        pathlib.Path(__file__).parent / "plantillas" / "Formato Presupuestos.xlsx",
        pathlib.Path.cwd() / "plantillas" / "Formato Presupuestos.xlsx",
    ]
    for p in possible_paths:
        if p.exists():
            template_path = p
            break
    
    try:
        if template_path:
            wb = openpyxl.load_workbook(template_path)
            ws = wb['Presupuesto']
        else:
            # Fallback: create a new workbook with basic structure
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Presupuesto'
            ws['B2'] = "PRESUPUESTO"
            ws['B4'] = "CLIENTE"
            ws['D4'] = "EMPRESA"
            ws['B12'] = "N°"; ws['C12'] = "DESCRIPCIÓN"; ws['D12'] = "UDS."; ws['E12'] = "CANTIDAD"; ws['F12'] = "PU"; ws['G12'] = "TOTAL"
    except Exception as e:
        # Fallback to an empty workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Presupuesto'
        ws['B2'] = "PRESUPUESTO"
        ws['B4'] = "CLIENTE"
        ws['D4'] = "EMPRESA"
        ws['B12'] = "N°"; ws['C12'] = "DESCRIPCIÓN"; ws['D12'] = "UDS."; ws['E12'] = "CANTIDAD"; ws['F12'] = "PU"; ws['G12'] = "TOTAL"
    
    # Fill header
    ws['B5'] = f"Nombre: {cliente}"
    ws['B8'] = f"Proyecto: {proyecto}"
    ws['B10'] = f"Fecha Presupuesto: {datetime.datetime.now().strftime('%d/%m/%Y')}"
    
    # Clear existing rows from 14 downwards to avoid left-over template data
    start_row = 14
    total_val = 0.0
    for i, row in enumerate(df_pres.itertuples(), start=0):
        r = start_row + i
        qty = float(row.CANTIDAD) if pd.notnull(row.CANTIDAD) else 0.0
        pu = float(row.PU) if pd.notnull(row.PU) else 0.0
        tot = qty * pu
        total_val += tot
        
        ws.cell(row=r, column=2, value=row._1) # N°
        ws.cell(row=r, column=3, value=row.DESCRIPCIÓN)
        ws.cell(row=r, column=4, value=row.UDS)
        ws.cell(row=r, column=5, value=qty)
        ws.cell(row=r, column=6, value=pu)
        ws.cell(row=r, column=7, value=tot)
        
        # Format numbers
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7).number_format = '"$"#,##0.00'
        
    # Grand Total row
    r_final = start_row + len(df_pres) + 1
    ws.cell(row=r_final, column=6, value="GRAN TOTAL:")
    ws.cell(row=r_final, column=7, value=total_val)
    ws.cell(row=r_final, column=7).number_format = '"$"#,##0.00'
    ws.cell(row=r_final, column=6).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=r_final, column=7).font = openpyxl.styles.Font(bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ─────────────────────────────────────────────────────────────────────────────
# SALARIOS MÍNIMOS POR PAÍS (2026) — actualización manual cada año
# ─────────────────────────────────────────────────────────────────────────────
SALARIOS_MIN = {
    "Colombia": {
        "moneda": "COP $",
        "salario_base": 1_750_905,          # SMLMV 2026 • Decreto 2513 de 2025
        "auxilio_transporte": 249_095,      # Aux. transporte 2026
        "factor_prestacional": 1.5221,      # Prima, Cesantías, Vacac., Salud, Pensión, ARL, SENA, ICBF, Caja
        "dias_mes": 30,
        "horas_dia": 8,
        "nota": "SMLMV 2026 • $1,750,905 + Aux. Transporte $249,095 = $2,000,000 total",
        "prestaciones": {
            "Prima de servicios":       "8.33% s/salario",
            "Cesantías":               "8.33% s/salario",
            "Interés cesantías":       "1% s/cesantías",
            "Vacaciones":              "4.17% s/salario",
            "Salud (empleador 8.5%)": "8.5% s/salario",
            "Pensión (empleador 12%)": "12% s/salario",
            "ARL nivel I":             "0.522% s/salario",
            "SENA":                    "2% s/salario",
            "ICBF":                    "3% s/salario",
            "Caja Compensación":       "4% s/salario",
        }
    },
    "Perú": {
        "moneda": "PEN S/",
        "salario_base": 1_130,              # RMV 2026 • Decreto Supremo (incremento desde S/1,025)
        "auxilio_transporte": 0,
        "factor_prestacional": 1.2683,      # CTS, gratificaciones, EsSalud 9%, SCTR
        "dias_mes": 30,
        "horas_dia": 8,
        "nota": "RMV 2026 • Ministerio de Trabajo del Perú",
        "prestaciones": {
            "CTS":                     "8.33% s/salario",
            "Gratificaciones":         "16.67% s/salario (2 al año)",
            "EsSalud (9%)":           "9% s/salario",
            "SCTR (aprox)":           "1.83% s/salario",
        }
    },
    "México": {
        "moneda": "MXN $",
        "salario_base": 312.41 * 30,        # SM diario 2026 Zona General x30
        "salario_diario": 312.41,
        "auxilio_transporte": 0,
        "factor_prestacional": 1.35,        # IMSS, INFONAVIT, vacaciones, aguinaldo
        "dias_mes": 30,
        "horas_dia": 8,
        "nota": "SM Zona General 2026 • CONASAMI $312.41/día • ZLF: $470.07/día",
        "prestaciones": {
            "Aguinaldo":               "15 días/año",
            "Vacaciones":              "12 días primer año + 25%",
            "IMSS (aprox)":           "~8.5% s/salario",
            "INFONAVIT":              "5% s/salario",
        }
    },
    "Ecuador": {
        "moneda": "USD $",
        "salario_base": 470,                # SBU 2026 • Acuerdo Ministerial
        "auxilio_transporte": 0,
        "factor_prestacional": 1.3350,      # IESS 11.15%, fondos reserva, décimos
        "dias_mes": 30,
        "horas_dia": 8,
        "nota": "SBU 2026 • $470/mes • Ministerio de Trabajo Ecuador",
        "prestaciones": {
            "Décimo tercero":         "8.33% s/salario",
            "Décimo cuarto":          "SBU / 12",
            "Fondos de reserva":       "8.33% (desde el 2do año)",
            "Aporte IESS empleador":   "11.15% s/salario",
        }
    },
    "Venezuela": {
        "moneda": "USD $",
        "salario_base": 130,                # Referencia plataforma USD 2026 (muy volátil)
        "auxilio_transporte": 0,
        "factor_prestacional": 1.25,
        "dias_mes": 30,
        "horas_dia": 8,
        "nota": "⚠ Referencia estimada USD 2026 • Verificar con LOTTT y IVSS vigentes",
        "prestaciones": {
            "Utilidades":             "15 días mín/año",
            "Vacaciones":             "15 días + bono",
            "IVSS (empleador)":       "~9% s/salario",
        }
    },
    "Bolivia": {
        "moneda": "BOB Bs.",
        "salario_base": 2_750,              # SMN 2026 • Decreto Supremo (incremento desde Bs.2,500)
        "auxilio_transporte": 0,
        "factor_prestacional": 1.43,        # AFP, bono, aguinaldo
        "dias_mes": 30,
        "horas_dia": 8,
        "nota": "SMN 2026 • Decreto Supremo Bolivia",
        "prestaciones": {
            "Aguinaldo (13°)": "1 salario/año",
            "AFP patronal":     "3% s/salario",
            "Pro-Vivienda":     "2% s/salario",
        }
    },
    "Argentina": {
        "moneda": "ARS $",
        "salario_base": 1_400_000,          # SMVM 2026 ref. • Alta inflación – verificar INDEC
        "auxilio_transporte": 0,
        "factor_prestacional": 1.50,        # Cargas sociales ~21%, SAC, vacaciones
        "dias_mes": 30,
        "horas_dia": 8,
        "nota": "⚠ Ref. 2026 • Actualizar por alta inflación • SMVM INDEC",
        "prestaciones": {
            "SAC (Aguinaldo)":         "8.33% s/salario",
            "Vacaciones":             "8.33% s/salario",
            "Seg. Social empleador":   "~21% s/salario",
        }
    },
    "USA": {
        "moneda": "USD $",
        "salario_base": 7.25 * 8 * 5 * 4,
        "salario_diario": 7.25 * 8,
        "auxilio_transporte": 0,
        "factor_prestacional": 1.25,
        "dias_mes": 30,
        "horas_dia": 8,
        "nota": "Salario mínimo federal $7.25/hora (estimado mensual)",
        "prestaciones": {
            "Social Security": "6.2% s/salario",
            "Medicare": "1.45% s/salario",
            "Unemployment": "~2% s/salario",
        }
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# HISTÓRICO SMLMV COLOMBIA (últimos 5 años) — para liquidaciones laborales
# ─────────────────────────────────────────────────────────────────────────────
HIST_SAL = {
    "Colombia": {
        2026: {"salario_base": 1_750_905, "auxilio_transporte": 249_095, "nota": "Decreto 2513/2025 • Incremento 23.%"},
        2025: {"salario_base": 1_423_500, "auxilio_transporte": 200_000, "nota": "Decreto 2616/2024 • Incremento 9.54%"},
        2024: {"salario_base": 1_300_000, "auxilio_transporte": 162_000, "nota": "Decreto 2641/2023 • Incremento 12.07%"},
        2023: {"salario_base": 1_160_000, "auxilio_transporte": 140_606, "nota": "Decreto 2613/2022 • Incremento 16%"},
        2022: {"salario_base": 1_000_000, "auxilio_transporte": 117_172, "nota": "Decreto 1724/2021 • Incremento 10.07%"},
    },
    "Perú": {
        2026: {"salario_base": 1_130, "auxilio_transporte": 0, "nota": "RMV 2026"},
        2025: {"salario_base": 1_025, "auxilio_transporte": 0, "nota": "RMV 2025"},
        2024: {"salario_base": 1_025, "auxilio_transporte": 0, "nota": "RMV 2024 (sin cambio)"},
        2023: {"salario_base": 1_025, "auxilio_transporte": 0, "nota": "RMV 2023"},
        2022: {"salario_base": 930,   "auxilio_transporte": 0, "nota": "RMV 2022"},
    },
    "México": {
        2026: {"salario_base": 312.41*30, "auxilio_transporte": 0, "nota": "SM $312.41/día • ZLF $470.07/día"},
        2025: {"salario_base": 278.80*30, "auxilio_transporte": 0, "nota": "SM $278.80/día • ZLF $419.88/día"},
        2024: {"salario_base": 248.93*30, "auxilio_transporte": 0, "nota": "SM $248.93/día"},
        2023: {"salario_base": 207.44*30, "auxilio_transporte": 0, "nota": "SM $207.44/día"},
        2022: {"salario_base": 172.87*30, "auxilio_transporte": 0, "nota": "SM $172.87/día"},
    },
    "Ecuador": {
        2026: {"salario_base": 470, "auxilio_transporte": 0, "nota": "SBU 2026"},
        2025: {"salario_base": 460, "auxilio_transporte": 0, "nota": "SBU 2025"},
        2024: {"salario_base": 450, "auxilio_transporte": 0, "nota": "SBU 2024"},
        2023: {"salario_base": 450, "auxilio_transporte": 0, "nota": "SBU 2023"},
        2022: {"salario_base": 425, "auxilio_transporte": 0, "nota": "SBU 2022"},
    },
    "USA": {
        2026: {"salario_base": 7.25*8*5*4, "auxilio_transporte": 0, "nota": "Federal minimum wage $7.25/h"},
        2025: {"salario_base": 7.25*8*5*4, "auxilio_transporte": 0, "nota": "Same as 2026"},
        2024: {"salario_base": 7.25*8*5*4, "auxilio_transporte": 0, "nota": "Same as 2026"},
        2023: {"salario_base": 7.25*8*5*4, "auxilio_transporte": 0, "nota": "Same as 2026"},
        2022: {"salario_base": 7.25*8*5*4, "auxilio_transporte": 0, "nota": "Same as 2026"},
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# STATE INIT
# ─────────────────────────────────────────────────────────────────────────────
if "kc_rows" not in st.session_state:
    st.session_state.kc_rows = []
    load_state()
else:
    if "state_loaded" not in st.session_state:
        load_state()
        st.session_state.state_loaded = True
if "kc_precios" not in st.session_state: st.session_state.kc_precios={}

# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="background:linear-gradient(135deg,#0d1b2a 0%,#1e3a5f 60%,#2e6da4 100%);
    padding:28px 36px;border-radius:16px;margin-bottom:18px;box-shadow:0 8px 32px rgba(0,0,0,0.4);">
  <div style="display:flex;align-items:center;gap:18px;">
    <span style="font-size:52px;"></span>
    <div>
      <h1 style="color:white;margin:0;font-size:2.2rem;font-weight:800;letter-spacing:-0.5px;">Konte — Calculadora de Materiales</h1>
      <p style="color:#90caf9;margin:4px 0 0;font-size:1rem;">Multinorma &nbsp;·&nbsp; Terminología Regional &nbsp;·&nbsp; Precios en Tiempo Real</p>
    </div>
  </div>
</div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR - CONFIGURACIÓN KONTE
# ─────────────────────────────────────────────────────────────────────────────
st.sidebar.markdown(f"## ⚙ {_t('Configuración Konte', 'Konte Settings')}")

# ── Selector de Normativa ──────────────────────────────────────────────────
NORMAS_LIST = list(REGIONAL.keys())
_norma_prev = st.session_state.get("norma_sel", "NSR-10 (Colombia)")
_norma_idx  = NORMAS_LIST.index(_norma_prev) if _norma_prev in NORMAS_LIST else 0

norma_sel = st.sidebar.selectbox(
    f" {_t('Normativa / País', 'Code / Country')}",
    NORMAS_LIST,
    index=_norma_idx,
    key="sel_norma_kc_ui",
)

if norma_sel != _norma_prev:
    st.session_state["norma_sel"] = norma_sel
    st.session_state.pop("auto_unit_kc", None)   # reset auto-imperial flag
    # Forzar reinicio de precios al nuevo país
    _temp_R = REGIONAL.get(norma_sel, REGIONAL["NSR-10 (Colombia)"])
    ref_nuevo = PRICE_URLS.get(_temp_R["pais"], {}).get("precios_ref", {"cemento":30, "arena":50, "grava":60, "acero_kg":3})
    st.session_state.kc_precios = dict(ref_nuevo)
    st.session_state.kc_pais_prev = _temp_R["pais"]
    st.rerun()

R = REGIONAL.get(norma_sel, REGIONAL["NSR-10 (Colombia)"])
R["_norma"] = norma_sel
pais = R["pais"]
moneda = R["moneda"]
PAIS_ISO = {"Colombia":"co","USA":"us","Ecuador":"ec","Perú":"pe","México":"mx","Venezuela":"ve","Bolivia":"bo","Argentina":"ar"}
iso = PAIS_ISO.get(pais, "un")

# Badge de norma activa
st.sidebar.markdown(
    f'<div style="background:#1e3a1e;border-radius:8px;padding:8px 14px;margin-top:4px;margin-bottom:8px;">'
    f'<img src="https://flagpedia.net/data/flags/mini/{iso}.png" style="vertical-align:middle;margin-right:8px;">'
    f'<span style="color:#7ec87e;font-weight:600;">{norma_sel}</span></div>',
    unsafe_allow_html=True)
st.sidebar.caption(f" **{_t('País', 'Country')}:** {pais}  |  **{_t('Moneda', 'Currency')}:** {moneda}")


# Toggles for Language & Units
col_l, col_u = st.sidebar.columns(2)
new_lang = col_l.selectbox(
    f" {_t('Idioma', 'Language')}", 
    ["Español", "English"], 
    index=0 if st.session_state["idioma"]=="Español" else 1,
    key="sel_lang_kc"
)
new_unit = col_u.selectbox(
    f" {_t('Unidad', 'Units')}", 
    ["Métrico", "Imperial"], 
    index=0 if st.session_state["unidades"]=="Métrico" else 1,
    key="sel_unit_kc"
)

if R.get("imperial", False) and "auto_unit_kc" not in st.session_state:
    st.session_state["unidades"] = "Imperial"
    st.session_state["auto_unit_kc"] = True
    new_unit = "Imperial"

if new_lang != st.session_state["idioma"] or new_unit != st.session_state["unidades"]:
    st.session_state["idioma"] = new_lang
    st.session_state["unidades"] = new_unit
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.markdown(f"###  {_t('Precios de Materiales', 'Material Prices')}")
ref=PRICE_URLS.get(pais,{}).get("precios_ref",{"cemento":30,"arena":50,"grava":60,"acero_kg":3})
if "kc_pais_prev" not in st.session_state or st.session_state.kc_pais_prev!=pais:
    st.session_state.kc_precios=dict(ref); st.session_state.kc_pais_prev=pais
if st.sidebar.button(f" {_t('Consultar Precios en Vivo', 'Get Live Prices')}", use_container_width=True):
    with st.sidebar:
        with st.spinner(_t("Consultando precios...", "Fetching prices...")):
            live=get_live_prices(pais,ref); st.session_state.kc_precios=live
    st.sidebar.success(_t(" Precios actualizados", " Prices updated"))
p=st.session_state.kc_precios
p["cemento"] =st.sidebar.number_input(f" {R['cemento']} [{moneda}/bolsa]",value=float(p.get("cemento",ref.get("cemento",30))),min_value=0.0,step=100.0,key="kc_p_cem",format="%.2f")
p["arena"]   =st.sidebar.number_input(f" {R['arena']} [{moneda}/m³]",value=float(p.get("arena",ref.get("arena",50))),min_value=0.0,step=100.0,key="kc_p_are",format="%.2f")
p["grava"]   =st.sidebar.number_input(f" {R['grava']} [{moneda}/m³]",value=float(p.get("grava",ref.get("grava",60))),min_value=0.0,step=100.0,key="kc_p_gra",format="%.2f")
p["acero_kg"]=st.sidebar.number_input(f" {R['varilla']} [{moneda}/kg]",value=float(p.get("acero_kg",ref.get("acero_kg",3))),min_value=0.0,step=10.0,key="kc_p_ace",format="%.2f")
p["bloque"]  =st.sidebar.number_input(f" {R['bloque']} (1 und) [{moneda}]",value=float(p.get("bloque",3700 if pais=="Colombia" else 2.0)),min_value=0.0,step=10.0,key="kc_p_blq",format="%.2f")
p["ladrillo"]=st.sidebar.number_input(f" {R['ladrillo']} (1 und) [{moneda}]",value=float(p.get("ladrillo",1800 if pais=="Colombia" else 1.2)),min_value=0.0,step=10.0,key="kc_p_lad",format="%.2f")
p["pintura"] =st.sidebar.number_input(f" {R['pintura']} [{moneda}/galón]",value=float(p.get("pintura",55000 if pais=="Colombia" else 40.0)),min_value=0.0,step=1.0,key="kc_p_pin",format="%.2f")
p["ceramica"]=st.sidebar.number_input(f" {R['ceramica']} [{moneda}/m²]",value=float(p.get("ceramica",18000 if pais=="Colombia" else 15.0)),min_value=0.0,step=0.5,key="kc_p_cer",format="%.2f")
st.session_state.kc_precios=p

if st.sidebar.button(f" {_t('Restablecer precios por defecto', 'Reset default prices')}", use_container_width=True):
    st.session_state.kc_precios = dict(ref)
    st.rerun()

st.sidebar.markdown("---")

# ── Resumen de precios vigentes (siempre visible) ──────────────────────
st.sidebar.markdown("###  Precios vigentes")
st.sidebar.markdown(f"""
<div style="background:#0d1b2a;border:1px solid #1e4d8c;border-radius:8px;padding:10px 12px;font-size:12px;line-height:1.8;">
  <div> Cemento: <b>{moneda} {p.get('cemento',0):,.0f}</b> / bolsa</div>
  <div> Arena: <b>{moneda} {p.get('arena',0):,.0f}</b> / m³</div>
  <div> Grava: <b>{moneda} {p.get('grava',0):,.0f}</b> / m³</div>
  <div> Varilla: <b>{moneda} {p.get('acero_kg',0):,.0f}</b> / kg</div>
  <div> Bloque: <b>{moneda} {p.get('bloque',0):,.0f}</b> / und</div>
  <div> Ladrillo: <b>{moneda} {p.get('ladrillo',0):,.0f}</b> / und</div>
  <div> Pintura: <b>{moneda} {p.get('pintura',0):,.0f}</b> / galón</div>
  <div> Cerámica: <b>{moneda} {p.get('ceramica',0):,.0f}</b> / m²</div>
</div>
<div style="margin-top:6px;font-size:10px;color:#ff8f00;">
  ⚠ Si los valores parecen incorrectos (ej: 30 en vez de 30.000), cambia los precios arriba.
</div>
""", unsafe_allow_html=True)
st.sidebar.markdown("---")

st.sidebar.markdown('<div style="text-align:center;color:gray;font-size:10px;">© 2026 Konte — <br><i>⚠ Uso profesional exclusivo</i></div>',unsafe_allow_html=True)

RENDIMIENTOS_MO = [
    {'Actividad': 'Excavación Zanja Drenaje a negras PVC', 'Cantidad': 1.75, 'Unidad': 'm³/día/peón'},
    {'Actividad': 'Excavación Zanja Drenaje Tubería Polipropileno', 'Cantidad': 1.5, 'Unidad': 'm³/día/peón'},
    {'Actividad': 'Colocación Tubería Concreto y rajas a negras', 'Cantidad': 2, 'Unidad': 'm³/día/albañil + ½ peón'},
    {'Actividad': 'Colocación Tubería Drenaje PVC Ø 2”', 'Cantidad': 30.4, 'Unidad': 'mL/día/plomero'},
    {'Actividad': 'Colocación Tubería Drenaje PVC Ø 3”', 'Cantidad': 28.5, 'Unidad': 'mL/día/plomero'},
    {'Actividad': 'Colocación Tubería Drenaje PVC Ø 4”', 'Cantidad': 24, 'Unidad': 'mL/día/plomero'},
    {'Actividad': 'Colocación Tubería Drenaje PVC Ø 6”', 'Cantidad': 18, 'Unidad': 'mL/día/plomero'},
    {'Actividad': 'Colocación Tubería Drenaje PVC Ø 8”', 'Cantidad': 12, 'Unidad': 'mL/día/plomero'},
    {'Actividad': 'Construcción cimiento corrido, incluye armado calz. Y fundición C.C - 1', 'Cantidad': 16, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Construcción cimiento corrido, incluye armado calz. Y fundición C.C - 2 ', 'Cantidad': 14, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Construcción cimiento corrido, incluye armado calz. Y fundición C.C - 3', 'Cantidad': 12, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Construcción Zapatas incluye levantado y fundición (1.00*1.00*0.25)', 'Cantidad': 2, 'Unidad': 'U/día Alb. + Ayud.'},
    {'Actividad': 'Construcción Zapatas incluye levantado y fundición (1.20*1.20*0.25)', 'Cantidad': 1.5, 'Unidad': 'U/día Alb. + Ayud.'},
    {'Actividad': 'Construcción Zapatas incluye levantado y fundición (1.50*1.50*0.30)', 'Cantidad': 1, 'Unidad': 'U/día Alb. + Ayud.'},
    {'Actividad': 'Columnetas amarre (armar, Encofrar, Fundir, Desenc.)', 'Cantidad': 9, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Solera de humedad (armar, encofrar, fundir, Desc)(.15*.20)', 'Cantidad': 12, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Levantado muretes hasta nivel piso', 'Cantidad': 7, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Viga de amarre 0.20 * 0.40', 'Cantidad': 3, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Viga de corona o collar - 15 * 20 (incluye todo)', 'Cantidad': 10, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Viga de amarre 15 * 30 (incluye todo)', 'Cantidad': 8, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Alzado de pared de bloque (0.15 * 0.20 * 0.40)', 'Cantidad': 12, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Alzado de pared de bloque (0.20 * 0.20 * 0.40)', 'Cantidad': 10, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Alzado de pared de bloque Vist. (0.15 * 0.20 * 0.40) 1 cara', 'Cantidad': 8, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Alzado de pared de Ladrillo Tubular', 'Cantidad': 12, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Alzado de pared de Ladrillo Tayuyo de punta', 'Cantidad': 8, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Fundición de techo Losa tradicional E= 10', 'Cantidad': 20, 'Unidad': 'm²/día Alb. + ayud'},
    {'Actividad': 'Armado de Acero Losa + Electrotubos', 'Cantidad': 16, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Colocación Casetón o Bovedilla Losa Pre.', 'Cantidad': 30, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Encofrado Columnas .20*.20 (.30*.30)', 'Cantidad': 3, 'Unidad': 'Col/día Carpintero'},
    {'Actividad': 'Fundición Columna .20*.20  a  .30*.30 ', 'Cantidad': 5, 'Unidad': 'Col/día Alb+2ayu'},
    {'Actividad': 'Armado de Columna  4 a 6 varillas 3/8 est, 1/4 ', 'Cantidad': 12, 'Unidad': 'H/día Armador'},
    {'Actividad': 'Colocación Tubería Eléctrica para losas 20 m2', 'Cantidad': 25, 'Unidad': 'm²/día Electricista'},
    {'Actividad': 'Losa Prefabricada Colocación (incluye encofrado y apuntalado)', 'Cantidad': 25, 'Unidad': 'm²/día Alb. + Peón'},
    {'Actividad': 'Instalación de cubierta asbesto E. Fibrocemento', 'Cantidad': 18, 'Unidad': 'm²/día Techero + Ayu'},
    {'Actividad': 'Instalación de Cubierta ZincAlum', 'Cantidad': 25, 'Unidad': 'm²/día Techero + Ayu'},
    {'Actividad': 'Colocación Tubería d=1/2 P/AP', 'Cantidad': 44.8, 'Unidad': 'm/día Plomero + Ayu'},
    {'Actividad': 'Colocación Tubería d=3/4 P/AP', 'Cantidad': 40, 'Unidad': 'm/día Plomero + Ayu'},
    {'Actividad': 'Colocación Tubería d=1 P/AP', 'Cantidad': 36, 'Unidad': 'm/día Plomero + Ayu'},
    {'Actividad': 'Instalación Inodoro c/tanque bajo', 'Cantidad': 2.5, 'Unidad': 'juego/día Plomero'},
    {'Actividad': 'Instalación lavamanos c/pedestal', 'Cantidad': 1.5, 'Unidad': 'juego/día Plomero'},
    {'Actividad': 'Instalación lavamanos para colgar', 'Cantidad': 2, 'Unidad': 'juego/día Plomero'},
    {'Actividad': 'Instalación urinario c/fluxómetro', 'Cantidad': 3.5, 'Unidad': 'juego/día Plomero'},
    {'Actividad': 'Instalación regaderas completas', 'Cantidad': 5, 'Unidad': 'juego/día Plomero'},
    {'Actividad': 'Instalación bidé completo', 'Cantidad': 2, 'Unidad': 'juego/día Plomero'},
    {'Actividad': 'Instalación batea prefabricada de concreto simple', 'Cantidad': 4, 'Unidad': 'U/día plomero'},
    {'Actividad': 'Instalación batea prefabricada de concreto doble', 'Cantidad': 3, 'Unidad': 'U/día Plomero'},
    {'Actividad': 'Instalación llave de chorro', 'Cantidad': 16, 'Unidad': 'U/día Plomero'},
    {'Actividad': 'Instalación mueble (pantry de doble depósito) (A. m. de chorros)', 'Cantidad': 2, 'Unidad': 'U/día Plomero'},
    {'Actividad': 'Instalación accesorios baño completo (M)', 'Cantidad': 4, 'Unidad': 'Jueg/día Plomero'},
    {'Actividad': 'Instalación caja registro mampostería c/tapa 40x40 cms a=50 cm', 'Cantidad': 2, 'Unidad': 'U/día plomero'},
    {'Actividad': 'Instalación caja ladrillo p/valv 30x30x20', 'Cantidad': 4, 'Unidad': 'U/día Alb + Ayud'},
    {'Actividad': 'Instalación caja ladrillo p/reja trans 50x50x40', 'Cantidad': 2.5, 'Unidad': 'U/día Alb + Ayud'},
    {'Actividad': 'Instalación de acometida 120 v', 'Cantidad': 0.5, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de acometida 220 v', 'Cantidad': 0.3, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de contador', 'Cantidad': 4, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de caja socket rh', 'Cantidad': 1, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de varilla de cobre', 'Cantidad': 8, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de tablero', 'Cantidad': 0.6, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de lámpara sencilla', 'Cantidad': 10, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de lámpara 3w', 'Cantidad': 8, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de lámpara 2*40', 'Cantidad': 5, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de tomacorriente 110 v', 'Cantidad': 5, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de tomacorriente polar', 'Cantidad': 5, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de intercomunicador 3U', 'Cantidad': 3, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de calentador', 'Cantidad': 1, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de calentador 110', 'Cantidad': 1, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de tomacorriente 220 v', 'Cantidad': 2, 'Unidad': 'U/día Electricista'},
    {'Actividad': 'Instalación de poliducto', 'Cantidad': 50, 'Unidad': 'mL/día Electricista'},
    {'Actividad': 'Ensabietado', 'Cantidad': 26, 'Unidad': 'm²/día '},
    {'Actividad': 'Repello de paredes con maestra', 'Cantidad': 16, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Repello de cielo con maestras', 'Cantidad': 11, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Cernido vertical en paredes', 'Cantidad': 23, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Cernido vertical en cielo', 'Cantidad': 20, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Cernido vertical exterior', 'Cantidad': 18, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Cernido final en pisos y terraza', 'Cantidad': 16, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Blanqueado de paredes', 'Cantidad': 16, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Blanqueado de cielos', 'Cantidad': 14, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Blanqueado de cenefa', 'Cantidad': 28, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Alisado', 'Cantidad': 10, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Gota resaltada y hundida hacer', 'Cantidad': 10, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Colocar sardineles y dinteles incluido todo', 'Cantidad': 8, 'Unidad': 'mL/día Alb. + Ayud.'},
    {'Actividad': 'Colocación de azulejo estucado', 'Cantidad': 12, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Colocación de mosaico', 'Cantidad': 10, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Cortes de fachaleta a mano', 'Cantidad': 120, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Alizado de cemento', 'Cantidad': 18, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Granceado en cielos', 'Cantidad': 15, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Granceado en paredes', 'Cantidad': 19, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Granito lavado', 'Cantidad': 14, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Martelinado', 'Cantidad': 8, 'Unidad': 'm²/día Alb. + Ayud.'},
    {'Actividad': 'Tallado de cajas eléctricas', 'Cantidad': 20, 'Unidad': 'U/día Alb. + Ayud.'},
    {'Actividad': 'Instalación de piso de granito o cerámico', 'Cantidad': 12, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Cernido de torta de concreto (0.10 esp.)', 'Cantidad': 12, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Hacer sisa tallada para piso de concreto', 'Cantidad': 40, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Instalación de piso vinílico', 'Cantidad': 30, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Instalación de zócalo plástico 3"', 'Cantidad': 50, 'Unidad': '    mL/día Alb. + Ayud.'},
    {'Actividad': 'Colocación de baldoseta estucada y sisada', 'Cantidad': 12, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Colocación de adoquín', 'Cantidad': 12, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Colocación de piso de barro cocido', 'Cantidad': 8, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Colocación de piso de cemento líquido', 'Cantidad': 20, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Fundición de grada de granito pulida', 'Cantidad': 8, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Mezclón hasta 0.05 de espesor para losa final', 'Cantidad': 75, 'Unidad': '    m²/día Alb. + Ayud.'},
    {'Actividad': 'Aplicación pintura aceite 2 manos', 'Cantidad': 22, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Aplicación pintura látex 2 manos', 'Cantidad': 22, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Aplicación pintura 3 manos', 'Cantidad': 12, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Aplicación de 1 fondo + 2 manos aceite en puerta', 'Cantidad': 10, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Aplicación pintura puerta de metal', 'Cantidad': 16, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Aplicación pintura molduras 6"', 'Cantidad': 16, 'Unidad': '    mL/día Pintor'},
    {'Actividad': 'Aplicación pintura ventana de hierro', 'Cantidad': 15, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Barnizado de paredes de ladrillo', 'Cantidad': 18, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Barnizado de puertas de madera fondo', 'Cantidad': 20, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Barnizado de puertas de madera finish', 'Cantidad': 16, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Impermeabilizado de baldosa de ladrillo', 'Cantidad': 14, 'Unidad': '    m²/día Pintor'},
    {'Actividad': 'Lavado de baldosa de piso con ácido', 'Cantidad': 12, 'Unidad': '    m²/día Peón'},
    {'Actividad': 'Colocación de puertas de madera (incluido chapa)', 'Cantidad': 1.5, 'Unidad': '    U/día Calificado'},
    {'Actividad': 'Colocación de puertas de MDF (incluido chapa)', 'Cantidad': 2, 'Unidad': '    U/día Calificado'},
    {'Actividad': 'Colocación de puertas de metal (incluido chapa)', 'Cantidad': 1, 'Unidad': '    U/día Calificado'},
    {'Actividad': 'Colocación de puertas de aluminio + vidrio', 'Cantidad': 5, 'Unidad': '    m²/día Calificado'},
    {'Actividad': 'Colocación de puerta de madera abatible', 'Cantidad': 5, 'Unidad': '    m²/día Calificado'},
    {'Actividad': 'Colocación de ventanas marco de aluminio', 'Cantidad': 5, 'Unidad': '    m²/día Calificado'},
    {'Actividad': 'Colocación de ventanas de hierro', 'Cantidad': 8, 'Unidad': '    m²/día Calificado'},
    {'Actividad': 'Colocación de ventanas de caoba', 'Cantidad': 6, 'Unidad': '    m²/día Calificado'},
    {'Actividad': 'Colocación de balcones de hierro', 'Cantidad': 8, 'Unidad': '    mL/día Calificado'},
    {'Actividad': 'Colocación de baranda de metal', 'Cantidad': 6, 'Unidad': '    mL/día Calificado'},
    {'Actividad': 'Instalación de puerta tipo Deko', 'Cantidad': 1, 'Unidad': '    U/día Calificado'},
    {'Actividad': 'Engramado con tepe', 'Cantidad': 10, 'Unidad': '    m²/día Calificado'},
    {'Actividad': 'Limpieza general', 'Cantidad': 60, 'Unidad': '    m²/día Calificado'},
    {'Actividad': 'Alineación Municipal', 'Cantidad': 0.3, 'Unidad': '    U/día Calificado'},
    {'Actividad': 'Licencia Municipal', 'Cantidad': 0.05, 'Unidad': '    U/día Calificado'},
    {'Actividad': 'Conexión domiciliar de drenajes', 'Cantidad': 0.1, 'Unidad': '    U/día Calificado'},
    {'Actividad': 'Inspección de la Empresa Eléctrica', 'Cantidad': 0.3, 'Unidad': '    U/día Calificado'},
]

import io
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# ─────────────────────────────────────────────────────────────────────────────
# MAIN TABS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
div[data-testid="stTabs"] div[role="tablist"] {
    flex-wrap: wrap;
    gap: 4px;
}
div[data-testid="stTabs"] button[role="tab"] {
    padding: 2px 10px;
    font-size: 0.85rem;
}
</style>
""", unsafe_allow_html=True)

tabs=st.tabs([f"{R['concreto']}","Mampostería","Columna","Viga","Losa","Cimiento","Muro Contención","Techo","Pisos","Cielo Raso",f"{R['varilla']}",
              "Pintura","Cubierta","Importar Excel","Salario Mínimo / Liq.","Configuración","Rendimientos MO","Presupuesto","Resumen y Exportar"])

# ══════════ TAB 1 — CONCRETO ══════════
with tabs[0]:
    st.subheader(f" Calculadora de {R['concreto']}")
    st.caption(f"Dosificaciones CM-V3.0 | Norma: {norma_sel}")
    geo_col, elem_col = st.columns([2, 3])
    with geo_col:
        geo_mode = st.radio("Tipo de geometría:", [" Cubo", " Cilindro", " Volumen"],
                            horizontal=True, key="kc_geo_mode")
    with elem_col:
        elemento_conc = st.selectbox("Elemento estructural:",
            [R["columna"], R["viga"], R["placa"], R["zapata"], R["cimiento"],
             "Muro de contención", "Entrepiso macizo"], key="kc_elem_conc")
    st.markdown("---")
    inp_col, fig_col = st.columns([2, 1])
    with inp_col:
        st.markdown("#####  Dimensiones")
        if "Cubo" in geo_mode:
            d1, d2 = st.columns(2)
            with d1:
                dim_a = st.number_input("Largo  a [m]", min_value=0.0, value=0.50, step=0.05, key="kc_da", format="%.3f")
                dim_b = st.number_input("Ancho  b [m]", min_value=0.0, value=0.50, step=0.05, key="kc_db", format="%.3f")
            with d2:
                dim_h = st.number_input("Altura  h [m]", min_value=0.0, value=3.00, step=0.05, key="kc_dh", format="%.3f")
                cantidad = st.number_input("Cantidad (und)", min_value=1, value=1, step=1, key="kc_qty")
            vol_geo = dim_a * dim_b * dim_h * cantidad
        elif "Cilindro" in geo_mode:
            d1, d2 = st.columns(2)
            with d1:
                dim_r = st.number_input("Radio  r [m]", min_value=0.0, value=0.20, step=0.01, key="kc_dr", format="%.3f")
            with d2:
                dim_h2 = st.number_input("Altura  h [m]", min_value=0.0, value=3.00, step=0.05, key="kc_dhc", format="%.3f")
                cantidad = st.number_input("Cantidad (und)", min_value=1, value=1, step=1, key="kc_qty2")
            vol_geo = math.pi * dim_r**2 * dim_h2 * cantidad
        else:
            vol_geo = st.number_input("Volumen [m³]", min_value=0.0, value=1.00, step=0.10, key="kc_vol_dir", format="%.4f")
            cantidad = st.number_input("Cantidad (und)", min_value=1, value=1, step=1, key="kc_qty3")
            vol_geo *= cantidad
        st.caption(f" Volumen neto calculado: **{vol_geo:.4f} m³**")
        st.markdown("#####  Concreto")
        mode_conc = st.radio("Seleccionar por:", ["Dosificación", "Resistencia f\'c"], horizontal=True, key="kc_mode_conc")
        if mode_conc == "Dosificación":
            dos_labels = [f"{m['dos']}  →  {m['fc_kgcm2']} kg/cm² ({m['fc_mpa']} MPa)" for m in MIX_DESIGNS]
            dos_sel = st.selectbox(" Dosificación C:A:G", dos_labels, index=4, key="kc_dos_sel")
            mix = MIX_DESIGNS[[m["dos"] for m in MIX_DESIGNS].index(dos_sel.split("  →")[0])]
        else:
            fc_kgcm2_opts = sorted(set(m["fc_kgcm2"] for m in MIX_DESIGNS))
            fc_kgcm2_sel = st.select_slider("Resistencia [kg/cm²]", options=fc_kgcm2_opts, value=200, key="kc_fc_kgcm2")
            mix = min(MIX_DESIGNS, key=lambda m: abs(m["fc_kgcm2"] - fc_kgcm2_sel))
            st.caption(f"Dosificación: **{mix['dos']}** | f\'c = {mix['fc_mpa']} MPa")
        desp_pct = st.slider("Desperdicio [%]", 0, 20, 5, key="kc_desp_conc")
        desperdicio = desp_pct / 100
        desc_elem = st.text_input("Descripción (opcional)", placeholder="Ej: Columna eje A-1", key="kc_desc_elem")
    with fig_col:
        if "Cubo" in geo_mode:
            a_v, b_v, h_v = max(dim_a, 0.01), max(dim_b, 0.01), max(dim_h, 0.01)
            verts = [[0,0,0],[a_v,0,0],[a_v,b_v,0],[0,b_v,0],[0,0,h_v],[a_v,0,h_v],[a_v,b_v,h_v],[0,b_v,h_v]]
            fig3d = go.Figure(go.Mesh3d(
                x=[v[0] for v in verts], y=[v[1] for v in verts], z=[v[2] for v in verts],
                i=[0, 0, 4, 4, 0, 0, 3, 3, 0, 0, 1, 1],
                j=[1, 2, 5, 6, 1, 5, 2, 6, 3, 7, 2, 6],
                k=[2, 3, 6, 7, 5, 4, 6, 7, 7, 4, 6, 5],
                color="#1e4d8c", opacity=0.85, flatshading=True))
            ann = f"a={dim_a}m · b={dim_b}m · h={dim_h}m<br>V={vol_geo:.2f}m³"
        elif "Cilindro" in geo_mode:
            theta = np.linspace(0, 2*np.pi, 40)
            r_val = max(dim_r, 0.01); h_val = max(dim_h2, 0.01)
            xc = r_val * np.cos(theta); yc = r_val * np.sin(theta)
            fig3d = go.Figure(go.Mesh3d(
                x=np.concatenate([xc, xc]), y=np.concatenate([yc, yc]),
                z=np.concatenate([np.zeros(40), np.full(40, h_val)]),
                alphahull=0, color="#1e4d8c", opacity=0.85))
            ann = f"r={dim_r}m · h={dim_h2}m<br>V={vol_geo:.2f}m³"
        else:
            fig3d = go.Figure(go.Scatter3d(x=[0,1],y=[0,0],z=[0,max(vol_geo,0.01)],
                mode="lines+markers", line=dict(color="#2196f3", width=12),
                marker=dict(size=8, color="#2196f3")))
            ann = f"V = {vol_geo:.4f} m³"
        fig3d.update_layout(scene=dict(aspectmode="data",
            xaxis=dict(showticklabels=False,title=""),
            yaxis=dict(showticklabels=False,title=""),
            zaxis=dict(showticklabels=False,title=""),
            bgcolor="rgba(10,20,40,1)"),
            paper_bgcolor="rgba(10,20,40,1)", margin=dict(l=0,r=0,t=0,b=0), height=260)
        st.plotly_chart(fig3d, use_container_width=True)
        st.markdown(f'<div style="text-align:center;color:#90caf9;font-size:12px;">{ann}</div>', unsafe_allow_html=True)
    st.markdown("---")
    vol_t = vol_geo * (1 + desperdicio)
    cem_kg = mix["cem_kg"] * vol_t; bolsas = math.ceil(cem_kg / R["peso_bolsa"])
    arena_c = mix["arena_m3"] * vol_t; grava_c = mix["grava_m3"] * vol_t; agua_c = mix["agua_lt"] * vol_t
    costo_c = bolsas * p["cemento"] + arena_c * p["arena"] + grava_c * p["grava"]
    st.markdown("####  Resumen de Resultados")
    rc1,rc2,rc3,rc4,rc5,rc6 = st.columns(6)
    rc1.metric(f" Cemento", f"{cem_kg:.1f} kg", f"{bolsas} bolsas")
    rc2.metric(f" Arena", f"{arena_c:.2f}m³")
    rc3.metric(f" Grava", f"{grava_c:.2f}m³")
    rc4.metric(f" Agua", f"{agua_c:.0f} lt")
    rc5.metric(" Dosif.", mix["dos"])
    with rc6:
        st.markdown("<p style='font-size:0.78em;color:#aaa;margin-bottom:0px'> Costo</p>", unsafe_allow_html=True)
        st.markdown(f"<p style='font-size:1.1em;font-weight:700;margin:0;'>{moneda} {costo_c:,.0f}</p>", unsafe_allow_html=True)
    st.markdown(
        f'<div style="background:#0d2137;border-radius:8px;padding:10px 18px;margin:8px 0;">'
        f'<span style="color:#7ec87e;font-weight:600;"> Vol. neto: {vol_geo:.4f} m³</span> &nbsp;'
        f'<span style="color:#90caf9;">+ {desp_pct}% desperdicio = <b>{vol_t:.4f} m³</b></span>'
        f'&nbsp;&nbsp; <span style="color:#ffcc80;">Dosificación: <b>{mix["dos"]}</b> | f\'c = <b>{mix["fc_kgcm2"]} kg/cm²</b> ({mix["fc_mpa"]} MPa)</span>'
        f'</div>', unsafe_allow_html=True)
    # ── Panel de precios aplicados (diagnóstico rápido) ──────────────────
    with st.expander(" Precios unitarios aplicados en este cálculo", expanded=False):
        _pr = st.session_state.kc_precios
        st.markdown(f"""
<div style="display:flex;flex-wrap:wrap;gap:10px;padding:4px 0;">
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:8px 16px;min-width:160px;">
    <div style="color:#90caf9;font-size:11px;margin-bottom:2px;"> Cemento / bolsa</div>
    <div style="font-size:1.15em;font-weight:700;color:#fff;">{moneda} {_pr.get('cemento', 0):,.0f}</div>
    <div style="font-size:10px;color:#7ec87e;">Bolsas usadas: {bolsas}</div>
    <div style="font-size:10px;color:#ffcc80;">Subtotal: {moneda} {bolsas * _pr.get('cemento', 0):,.0f}</div>
  </div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:8px 16px;min-width:160px;">
    <div style="color:#90caf9;font-size:11px;margin-bottom:2px;"> Arena / m³</div>
    <div style="font-size:1.15em;font-weight:700;color:#fff;">{moneda} {_pr.get('arena', 0):,.0f}</div>
    <div style="font-size:10px;color:#7ec87e;">m³ usados: {arena_c:.3f}</div>
    <div style="font-size:10px;color:#ffcc80;">Subtotal: {moneda} {arena_c * _pr.get('arena', 0):,.0f}</div>
  </div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:8px 16px;min-width:160px;">
    <div style="color:#90caf9;font-size:11px;margin-bottom:2px;"> Grava / m³</div>
    <div style="font-size:1.15em;font-weight:700;color:#fff;">{moneda} {_pr.get('grava', 0):,.0f}</div>
    <div style="font-size:10px;color:#7ec87e;">m³ usados: {grava_c:.3f}</div>
    <div style="font-size:10px;color:#ffcc80;">Subtotal: {moneda} {grava_c * _pr.get('grava', 0):,.0f}</div>
  </div>
  <div style="background:#1a1a0d;border:1px solid #6d6000;border-radius:8px;padding:8px 16px;min-width:160px;align-self:center;">
    <div style="color:#ffcc80;font-size:11px;margin-bottom:2px;"> TOTAL MATERIALES</div>
    <div style="font-size:1.3em;font-weight:800;color:#ffdd57;">{moneda} {costo_c:,.0f}</div>
    <div style="font-size:10px;color:#aaa;">Norma: {norma_sel}</div>
  </div>
</div>
<div style="margin-top:8px;padding:6px 10px;background:#1a1000;border-left:3px solid #ff8f00;border-radius:4px;font-size:11px;color:#ffcc80;">
  ⚠ Si los precios se ven muy bajos o muy altos, verifica la moneda en el sidebar (<b>{moneda}</b>).
  Los precios se pueden ajustar manualmente en <b> Precios de Materiales</b> (panel izquierdo).
</div>
""", unsafe_allow_html=True)
    if st.button(f" Agregar {elemento_conc} al Resumen", key="kc_add_conc", type="primary"):
        label = desc_elem if desc_elem else elemento_conc
        st.session_state.kc_rows.extend([
            {"elemento": f"{label} — {R['cemento']}", "unidad": "bultos", "cant": round(bolsas, 0), "precio": p["cemento"]},
            {"elemento": f"{label} — {R['arena']}", "unidad": "m³", "cant": round(arena_c, 3), "precio": p["arena"]},
            {"elemento": f"{label} — {R['grava']}", "unidad": "m³", "cant": round(grava_c, 3), "precio": p["grava"]},
            {"elemento": f"{label} — Agua", "unidad": "lt", "cant": round(agua_c, 0), "precio": 0},
        ])
        save_state()
        st.success(f" {label} — {bolsas} bolsas | {arena_c:.2f}m³ arena | {grava_c:.2f}m³ grava | {agua_c:.0f} lt")

# ══════════ TAB 2 — PARED / MAMPOSTERÍA (7 modos) ══════════
with tabs[1]:
    st.subheader(" Calculadora de Pared")
    st.caption(f"CM-V3.0 | 7 tipos | {norma_sel}")

    MODOS_PARED = [" Block"," Ladrillo"," Ciclópeo"," Por % (CA)"," Dimensión (CA)"," Panel Yeso"," Pintura"]
    modo_pared = st.radio("Tipo:", MODOS_PARED, horizontal=True, key="kc_pared_modo")
    st.markdown("---")

    # ─── datos mortar helper ───
    DOS_MORT = {"1:3":{"ratio":3,"kg_cem_m3":490},"1:4":{"ratio":4,"kg_cem_m3":370},
                "1:5":{"ratio":5,"kg_cem_m3":300},"1:6":{"ratio":6,"kg_cem_m3":250}}
    def mortero_mats(dos_key, vol_m3, peso_bolsa):
        d = DOS_MORT[dos_key]
        kg = d["kg_cem_m3"] * vol_m3
        bolsas = math.ceil(kg / peso_bolsa)
        arena = vol_m3 * d["ratio"] / (d["ratio"]+1)
        agua = vol_m3 * 200
        return bolsas, round(arena,3), round(agua,0)

    # helper precio
    def _precio(key, default=0):
        return float(p.get(key, default))

    desc_pared = st.text_input("Descripción del elemento", placeholder="Ej: Muro eje A", key="kc_pared_desc")

    # SVG Diagram
    if modo_pared not in [" Panel Yeso", " Pintura"]:
        html_capas = """
        <div style="display:flex; flex-direction:column; align-items:center; margin-bottom:20px; font-family:sans-serif; color:#ddd; font-size:12px; background:#1e1e1e; padding:15px; border-radius:8px; border:1px solid #333;">
            <div style="display:flex; align-items:center; width:100%; max-width:500px; margin-bottom:4px;">
                <span style="width:120px; text-align:right; margin-right:12px; color:#ff8a8a; font-weight:bold;">A1 (Afinado Ext)</span>
                <div style="flex-grow:1; height:6px; background-color:#8c1c1c; border-radius:2px;"></div>
            </div>
            <div style="display:flex; align-items:center; width:100%; max-width:500px; margin-bottom:4px;">
                <span style="width:120px; text-align:right; margin-right:12px; color:#b0bec5; font-weight:bold;">R1 (Repello Ext)</span>
                <div style="flex-grow:1; height:16px; background-color:#7a7a7a; border-radius:2px;"></div>
            </div>
            <div style="display:flex; align-items:center; width:100%; max-width:500px; margin-bottom:4px;">
                <span style="width:120px; text-align:right; margin-right:12px; color:#fff; font-weight:bold;">Muro Base</span>
                <div style="flex-grow:1; height:45px; background-color:#4B566A; border-radius:4px; display:flex; align-items:center; justify-content:center; border:2px dashed #6f809a;">Bloque / Ladrillo / Ciclópeo / CA</div>
            </div>
            <div style="display:flex; align-items:center; width:100%; max-width:500px; margin-bottom:4px;">
                <span style="width:120px; text-align:right; margin-right:12px; color:#b0bec5; font-weight:bold;">R2 (Repello Int)</span>
                <div style="flex-grow:1; height:16px; background-color:#7a7a7a; border-radius:2px;"></div>
            </div>
            <div style="display:flex; align-items:center; width:100%; max-width:500px; margin-bottom:4px;">
                <span style="width:120px; text-align:right; margin-right:12px; color:#ff8a8a; font-weight:bold;">A2 (Afinado Int)</span>
                <div style="flex-grow:1; height:6px; background-color:#8c1c1c; border-radius:2px;"></div>
            </div>
        </div>
        """
        st.markdown(html_capas, unsafe_allow_html=True)
        
        # Helpers for finishes
        def ui_acabados(key_prefix, default_inc=None):
            st.markdown("** Capas y Acabados**")
            ac1, ac2 = st.columns(2)
            with ac1:
                dos_rep  = st.selectbox("Mortero Repello:",list(DOS_MORT.keys()),index=1,key=f"{key_prefix}_rep")
                esp_rep  = st.select_slider("Espesor Repello [cm]",[0.5,1.0,1.5,2.0,2.5,3.0],value=1.5,key=f"{key_prefix}_esrep")
                esp_afin = st.select_slider("Espesor Afinado [mm]",[3,4,5,6,7,8],value=5,key=f"{key_prefix}_esafin")
            with ac2:
                opciones_capas = [
                    "A1 R1 R2 A2 (Todas)",
                    "Muro Base (Ninguna)",
                    "A1", "A2", "R1", "R2",
                    "A1 R1 A2", "A1 R1 R2",
                    "A1 R1", "A1 R2",
                    "R1 A2", "R2 A2",
                    "R1 R2", "A1 A2",
                    "A1 R2 A2", "R1 R2 A2"
                ]
                default_idx = 0
                if default_inc:
                    # Buscar la opción que mejor coincida con default_inc
                    for i, op in enumerate(opciones_capas):
                        if all(d.split(" ")[0] in op for d in default_inc):
                            default_idx = i; break
                incluir_sel = st.selectbox("Incluir capas:", opciones_capas, index=default_idx, key=f"{key_prefix}_inc")
                incluir = []
                if "A1" in incluir_sel and "Ninguna" not in incluir_sel: incluir.append("A1")
                if "A2" in incluir_sel and "Ninguna" not in incluir_sel: incluir.append("A2")
                if "R1" in incluir_sel and "Ninguna" not in incluir_sel: incluir.append("R1")
                if "R2" in incluir_sel and "Ninguna" not in incluir_sel: incluir.append("R2")
                
                desp_acab = st.slider("Desperdicio acabados [%]",0,15,5,key=f"{key_prefix}_dacab")/100
            return dos_rep, esp_rep, esp_afin, incluir, desp_acab

        def calc_acabados(area_t, dos_rep, esp_rep, esp_afin, incluir, desp_acab):
            n_rep = sum(1 for c in incluir if "R" in c)
            mort_rep_vol = (esp_rep/100) * area_t * n_rep
            bol_rep, arena_rep, agua_rep = mortero_mats(dos_rep, mort_rep_vol*(1+desp_acab), R["peso_bolsa"])
            
            n_afin = sum(1 for c in incluir if "A" in c)
            afin_vol = (esp_afin/1000) * area_t * n_afin
            bol_afin, arena_afin, agua_afin = mortero_mats("1:3", afin_vol*(1+desp_acab), R["peso_bolsa"])
            return bol_rep, arena_rep, agua_rep, bol_afin, arena_afin, agua_afin

    # ========= BLOCK =========
    if modo_pared == " Block":
        TIPOS_BLOQUE = {"B-10x20x40 cm":{"und_m2":12.5,"mort_pega_m3_m2":0.012},"B-12x20x40 cm":{"und_m2":12.5,"mort_pega_m3_m2":0.014},
                        "B-15x20x40 cm":{"und_m2":12.5,"mort_pega_m3_m2":0.016},"B-20x20x40 cm":{"und_m2":12.5,"mort_pega_m3_m2":0.019}}
        i1,i2,i3=st.columns(3)
        with i1:
            area_b = st.number_input("Área [m²]",min_value=0.0,value=10.0,step=0.5,key="kc_b_area")
            tipo_b = st.selectbox("Bloque:",list(TIPOS_BLOQUE.keys()),key="kc_b_tipo")
            bkd = TIPOS_BLOQUE[tipo_b]
        with i2:
            cantidad_b = st.number_input("Cantidad elementos",min_value=1,value=1,key="kc_b_qty")
            dos_pega = st.selectbox("Mortero de Pega:",list(DOS_MORT.keys()),index=0,key="kc_b_pega")
        with i3:
            desp_blq  = st.slider("Desperdicio bloques [%]",0,15,5,key="kc_b_dblq")/100
            
        dos_rep, esp_rep, esp_afin, incluir, desp_acab = ui_acabados("kc_b")
        
        area_t = area_b * cantidad_b
        und_blq = math.ceil(bkd["und_m2"] * area_t * (1+desp_blq))
        mort_pega_vol = bkd["mort_pega_m3_m2"] * area_t
        bol_pega, arena_pega, agua_pega = mortero_mats(dos_pega, mort_pega_vol*(1+desp_blq), R["peso_bolsa"])
        
        bol_rep, arena_rep, agua_rep, bol_afin, arena_afin, agua_afin = calc_acabados(area_t, dos_rep, esp_rep, esp_afin, incluir, desp_acab)

        st.markdown("---")
        und_bol = "bulto" if R.get("pais") == "Colombia" else "bol"
        nom_rep = R.get("pañete", "Repello").split("/")[0].strip()
        costo_base = und_blq*_precio("bloque") + bol_pega*_precio("cemento") + arena_pega*_precio("arena") + agua_pega*_precio("agua",0)
        costo_acabados = (bol_rep+bol_afin)*_precio("cemento") + (arena_rep+arena_afin)*_precio("arena") + (agua_rep+agua_afin)*_precio("agua",0)
        
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"** Muro Base** ({moneda} {costo_base:,.0f})")
            m1,m2,m3,m4 = st.columns(4)
            m1.metric("Bloques", f"{und_blq:,} und"); m2.metric("Cem. Pega", f"{bol_pega} {und_bol}"); m3.metric("Arena", f"{arena_pega:.2f}m³"); m4.metric("Agua", f"{agua_pega:.0f} lt")
            st.caption(f"Dosificación: {dos_pega}")
            
        with c2:
            st.markdown(f"** Acabados** ({moneda} {costo_acabados:,.0f})")
            a1,a2,a3,a4 = st.columns(4)
            a1.metric(f"Cem. {nom_rep}", f"{bol_rep} {und_bol}"); a2.metric("Cem. Afinado", f"{bol_afin} {und_bol}")
            a3.metric("Arena", f"{arena_rep+arena_afin:.2f}m³"); a4.metric("Agua", f"{agua_rep+agua_afin:.0f} lt")
            
        st.info(f" Costo Total Estimado: **{moneda} {costo_base+costo_acabados:,.0f}**")
        with st.expander(" Precios unitarios aplicados", expanded=False):
            _pr = st.session_state.kc_precios
            st.markdown(f"""<div style="display:flex;flex-wrap:wrap;gap:8px;">
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#90caf9;font-size:11px;"> Cemento/bolsa</div>
    <b>{moneda} {_pr.get('cemento',0):,.0f}</b>
  </div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#90caf9;font-size:11px;"> Arena/m³</div>
    <b>{moneda} {_pr.get('arena',0):,.0f}</b>
  </div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#90caf9;font-size:11px;"> Bloque/und</div>
    <b>{moneda} {_pr.get('bloque',0):,.0f}</b>
  </div>
  <div style="background:#1a1a0d;border:1px solid #6d6000;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#ffcc80;font-size:11px;"> TOTAL</div>
    <b style="color:#ffdd57;">{moneda} {costo_base+costo_acabados:,.0f}</b>
  </div>
</div><div style="margin-top:5px;font-size:10px;color:#ff8f00;">⚠ Moneda activa: <b>{moneda}</b> | Norma: {norma_sel}</div>""", unsafe_allow_html=True)
        if st.button(" Agregar al Resumen",key="kc_add_b",type="primary"):
            lb = desc_pared or f"Muro {tipo_b}"
            st.session_state.kc_rows.extend([
                {"elemento":f"{lb} — {R['bloque']}","unidad":"und","cant":und_blq,"precio":_precio("bloque")},
                {"elemento":f"{lb} — Cemento Base","unidad":und_bol,"cant":bol_pega,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Cemento Acabados","unidad":und_bol,"cant":bol_rep+bol_afin,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Arena","unidad":"m³","cant":round(arena_pega+arena_rep+arena_afin,3),"precio":_precio("arena")},
            ]); st.success(f" Añadido: {und_blq} bloques + {bol_pega} base + {bol_rep+bol_afin} acabados")

    # ========= LADRILLO =========
    elif modo_pared == " Ladrillo":
        TIPOS_LAD = {"L-A (6x12x25 cm)":{"und_m2":33,"mort_m3_m2":0.010},"L-B (8x12x25 cm)":{"und_m2":33,"mort_m3_m2":0.012},
                     "L-C (10x14x28 cm)":{"und_m2":25,"mort_m3_m2":0.015},"L-Bloque (12x25x40)":{"und_m2":10,"mort_m3_m2":0.018}}
        i1,i2,i3=st.columns(3)
        with i1:
            area_l = st.number_input("Área [m²]",min_value=0.0,value=10.0,step=0.5,key="kc_l_area")
            tipo_l = st.selectbox("Ladrillo:",list(TIPOS_LAD.keys()),key="kc_l_tipo"); lkd=TIPOS_LAD[tipo_l]
        with i2:
            cant_l = st.number_input("Cantidad elementos",min_value=1,value=1,key="kc_l_qty")
            dos_l_pega = st.selectbox("Mortero Pega:",list(DOS_MORT.keys()),index=2,key="kc_l_pega")
        with i3:
            desp_l_lad  = st.slider("Desperdicio ladrillo [%]",0,15,7,key="kc_l_dlad")/100
            
        dos_rep, esp_rep, esp_afin, incluir, desp_acab = ui_acabados("kc_l", default_inc=["R1 (repello ext)","R2 (repello int)"])
        
        area_lt = area_l*cant_l
        und_l = math.ceil(lkd["und_m2"]*area_lt*(1+desp_l_lad))
        vp = lkd["mort_m3_m2"]*area_lt; bp,ap,agp = mortero_mats(dos_l_pega,vp*(1+desp_l_lad),R["peso_bolsa"])
        
        bol_rep, arena_rep, agua_rep, bol_afin, arena_afin, agua_afin = calc_acabados(area_lt, dos_rep, esp_rep, esp_afin, incluir, desp_acab)
        
        st.markdown("---")
        und_bol = "bulto" if R.get("pais") == "Colombia" else "bol"
        nom_rep = R.get("pañete", "Repello").split("/")[0].strip()
        costo_base = und_l*_precio("ladrillo") + bp*_precio("cemento") + ap*_precio("arena") + agp*_precio("agua",0)
        costo_acabados = (bol_rep+bol_afin)*_precio("cemento") + (arena_rep+arena_afin)*_precio("arena") + (agua_rep+agua_afin)*_precio("agua",0)
        
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"** Muro Base** ({moneda} {costo_base:,.0f})")
            m1,m2,m3,m4 = st.columns(4)
            m1.metric("Ladrillos", f"{und_l:,} und"); m2.metric("Cem. Pega", f"{bp} {und_bol}"); m3.metric("Arena", f"{ap:.2f}m³"); m4.metric("Agua", f"{agp:.0f} lt")
            st.caption(f"Dosificación: {dos_l_pega}")
            
        with c2:
            st.markdown(f"** Acabados** ({moneda} {costo_acabados:,.0f})")
            a1,a2,a3,a4 = st.columns(4)
            a1.metric(f"Cem. {nom_rep}", f"{bol_rep} {und_bol}"); a2.metric("Cem. Afinado", f"{bol_afin} {und_bol}")
            a3.metric("Arena", f"{arena_rep+arena_afin:.2f}m³"); a4.metric("Agua", f"{agua_rep+agua_afin:.0f} lt")
            
        st.info(f" Costo Total Estimado: **{moneda} {costo_base+costo_acabados:,.0f}**")
        with st.expander(" Precios unitarios aplicados", expanded=False):
            _pr = st.session_state.kc_precios
            st.markdown(f"""<div style="display:flex;flex-wrap:wrap;gap:8px;">
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#90caf9;font-size:11px;"> Cemento/bolsa</div>
    <b>{moneda} {_pr.get('cemento',0):,.0f}</b>
  </div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#90caf9;font-size:11px;"> Arena/m³</div>
    <b>{moneda} {_pr.get('arena',0):,.0f}</b>
  </div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#90caf9;font-size:11px;"> Ladrillo/und</div>
    <b>{moneda} {_pr.get('ladrillo',0):,.0f}</b>
  </div>
  <div style="background:#1a1a0d;border:1px solid #6d6000;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#ffcc80;font-size:11px;"> TOTAL</div>
    <b style="color:#ffdd57;">{moneda} {costo_base+costo_acabados:,.0f}</b>
  </div>
</div><div style="margin-top:5px;font-size:10px;color:#ff8f00;">⚠ Moneda activa: <b>{moneda}</b> | Norma: {norma_sel}</div>""", unsafe_allow_html=True)
        if st.button(" Agregar al Resumen",key="kc_add_l",type="primary"):
            lb=desc_pared or f"Muro {tipo_l}"
            st.session_state.kc_rows.extend([
                {"elemento":f"{lb} — Ladrillos","unidad":"und","cant":und_l,"precio":_precio("bloque")*0.6},
                {"elemento":f"{lb} — Cemento Base","unidad":und_bol,"cant":bp,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Cemento Acabados","unidad":und_bol,"cant":bol_rep+bol_afin,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Arena","unidad":"m³","cant":round(ap+arena_rep+arena_afin,3),"precio":_precio("arena")},
            ]); st.success(f" Añadido: {und_l} ladrillos + {bp} base + {bol_rep+bol_afin} acabados")

    # ========= CICLÓPEO =========
    elif modo_pared == " Ciclópeo":
        i1,i2=st.columns(2)
        with i1:
            alto_c=st.number_input("Alto [m]",0.0,value=2.0,step=0.1,key="kc_c_alt")
            largo_c=st.number_input("Largo [m]",0.0,value=5.0,step=0.5,key="kc_c_lar")
            esp_c=st.number_input("Espesor [m]",0.0,value=0.30,step=0.05,key="kc_c_esp")
            cant_c=st.number_input("Cantidad",min_value=1,value=1,key="kc_c_qty")
        with i2:
            dos_c = st.selectbox("Dosificación mortero:",list(DOS_MORT.keys()),index=0,key="kc_c_dos")
            pct_mortero = st.selectbox("% Mortero:",[25,30,35,40,45,50],index=2,key="kc_c_mort")
            pct_piedra  = 100-pct_mortero
            desp_p = st.slider("Desperdicio piedra [%]",0,15,4,key="kc_c_despp")/100
            desp_m = st.slider("Desperdicio mortero [%]",0,15,6,key="kc_c_despm")/100
            
        dos_rep, esp_rep, esp_afin, incluir, desp_acab = ui_acabados("kc_c")
        
        vol_total_c = alto_c*largo_c*esp_c*cant_c
        area_c_t = alto_c * largo_c * cant_c
        vol_piedra = vol_total_c*(pct_piedra/100)*(1+desp_p)
        vol_mort_c = vol_total_c*(pct_mortero/100)*(1+desp_m)
        bol_c,arena_c_val,agua_c_val = mortero_mats(dos_c,vol_mort_c,R["peso_bolsa"])
        
        bol_rep, arena_rep, agua_rep, bol_afin, arena_afin, agua_afin = calc_acabados(area_c_t, dos_rep, esp_rep, esp_afin, incluir, desp_acab)
        
        st.markdown("---")
        und_bol = "bulto" if R.get("pais") == "Colombia" else "bol"
        nom_rep = R.get("pañete", "Repello").split("/")[0].strip()
        costo_base = bol_c*_precio("cemento") + arena_c_val*_precio("arena") + agua_c_val*_precio("agua",0) + vol_piedra*_precio("grava")*0.8
        costo_acabados = (bol_rep+bol_afin)*_precio("cemento") + (arena_rep+arena_afin)*_precio("arena") + (agua_rep+agua_afin)*_precio("agua",0)
        
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"** Muro Ciclópeo (Base)** ({moneda} {costo_base:,.0f})")
            m1,m2,m3,m4 = st.columns(4)
            m1.metric("Piedra", f"{vol_piedra:.2f}m³"); m2.metric("Cemento", f"{bol_c} {und_bol}"); m3.metric("Arena", f"{arena_c_val:.2f}m³"); m4.metric("Agua", f"{agua_c_val:.0f} lt")
            st.caption(f"Dim: {vol_total_c:.2f}m³ | Piedra: {pct_piedra}% | Mortero: {pct_mortero}%")
            
        with c2:
            st.markdown(f"** Acabados** ({moneda} {costo_acabados:,.0f})")
            a1,a2,a3,a4 = st.columns(4)
            a1.metric(f"Cem. {nom_rep}", f"{bol_rep} {und_bol}"); a2.metric("Cem. Afinado", f"{bol_afin} {und_bol}")
            a3.metric("Arena", f"{arena_rep+arena_afin:.2f}m³"); a4.metric("Agua", f"{agua_rep+agua_afin:.0f} lt")

        st.info(f" Costo Total Estimado: **{moneda} {costo_base+costo_acabados:,.0f}**")
        with st.expander(" Precios unitarios aplicados", expanded=False):
            _pr = st.session_state.kc_precios
            st.markdown(f"""<div style="display:flex;flex-wrap:wrap;gap:8px;">
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#90caf9;font-size:11px;"> Cemento/bolsa</div>
    <b>{moneda} {_pr.get('cemento',0):,.0f}</b>
  </div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#90caf9;font-size:11px;"> Arena/m³</div>
    <b>{moneda} {_pr.get('arena',0):,.0f}</b>
  </div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#90caf9;font-size:11px;"> Grava(Piedra)/m³</div>
    <b>{moneda} {_pr.get('grava',0):,.0f}</b>
  </div>
  <div style="background:#1a1a0d;border:1px solid #6d6000;border-radius:8px;padding:6px 12px;min-width:140px;">
    <div style="color:#ffcc80;font-size:11px;"> TOTAL</div>
    <b style="color:#ffdd57;">{moneda} {costo_base+costo_acabados:,.0f}</b>
  </div>
</div><div style="margin-top:5px;font-size:10px;color:#ff8f00;">⚠ Moneda activa: <b>{moneda}</b></div>""", unsafe_allow_html=True)
        if st.button(" Agregar al Resumen",key="kc_add_c",type="primary"):
            lb=desc_pared or "Muro Ciclópeo"
            st.session_state.kc_rows.extend([
                {"elemento":f"{lb} — Piedra","unidad":"m³","cant":round(vol_piedra,3),"precio":_precio("grava")*0.8},
                {"elemento":f"{lb} — Cemento Base","unidad":und_bol,"cant":bol_c,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Cemento Acabados","unidad":und_bol,"cant":bol_rep+bol_afin,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Arena","unidad":"m³","cant":round(arena_c_val+arena_rep+arena_afin,3),"precio":_precio("arena")},
            ]); st.success(f" Añadido: {vol_piedra:.2f}m³ piedra + {bol_c} base + {bol_rep+bol_afin} acabados")

    # ========= POR % (Concreto Armado) =========
    elif modo_pared == " Por % (CA)":
        i1,i2=st.columns(2)
        with i1:
            alt_pca=st.number_input("Alto [m]",0.0,value=2.0,step=0.1,key="kc_pca_alt")
            lar_pca=st.number_input("Largo [m]",0.0,value=5.0,step=0.5,key="kc_pca_lar")
            esp_pca=st.number_input("Espesor [m]",0.0,value=0.15,step=0.05,key="kc_pca_esp")
            cant_pca=st.number_input("Cantidad",min_value=1,value=1,key="kc_pca_qty")
        with i2:
            fc_pca = st.selectbox("Resistencia concreto:",[m["fc_kgcm2"] for m in MIX_DESIGNS],index=4,key="kc_pca_fc")
            mix_pca = next(m for m in MIX_DESIGNS if m["fc_kgcm2"]==fc_pca)
            pct_acero = st.selectbox("% Acero en volumen:",[0.5,1.0,1.5,2.0,2.5,3.0,3.5,4.0],index=1,key="kc_pca_pctace")
            var_pca = st.selectbox("Diámetro varilla:",[v["nombre"] for v in VARILLAS],index=3,key="kc_pca_var")
            vkd_pca = next(v for v in VARILLAS if v["nombre"]==var_pca)
            desp_conc_pca = st.slider("Desperdicio concreto [%]",0,15,5,key="kc_pca_dc")/100
            desp_ace_pca  = st.slider("Desperdicio acero [%]",0,15,5,key="kc_pca_da")/100
            
        dos_rep, esp_rep, esp_afin, incluir, desp_acab = ui_acabados("kc_pca")
        
        area_pca_t = alt_pca * lar_pca * cant_pca
        vol_pca = area_pca_t * esp_pca * (1+desp_conc_pca)
        cem_pca = mix_pca["cem_kg"]*vol_pca; bol_pca=math.ceil(cem_pca/R["peso_bolsa"])
        arena_pca=mix_pca["arena_m3"]*vol_pca; grava_pca=mix_pca["grava_m3"]*vol_pca; agua_pca=mix_pca["agua_lt"]*vol_pca
        vol_ace = alt_pca*lar_pca*esp_pca*cant_pca*(pct_acero/100)
        kg_ace = vol_ace*7850*(1+desp_ace_pca)
        
        bol_rep, arena_rep, agua_rep, bol_afin, arena_afin, agua_afin = calc_acabados(area_pca_t, dos_rep, esp_rep, esp_afin, incluir, desp_acab)
        
        st.markdown("---")
        und_bol = "bulto" if R.get("pais") == "Colombia" else "bol"
        nom_rep = R.get("pañete", "Repello").split("/")[0].strip()
        costo_base = bol_pca*_precio("cemento") + arena_pca*_precio("arena") + grava_pca*_precio("grava") + agua_pca*_precio("agua",0) + kg_ace*_precio("acero_kg")
        costo_acabados = (bol_rep+bol_afin)*_precio("cemento") + (arena_rep+arena_afin)*_precio("arena") + (agua_rep+agua_afin)*_precio("agua",0)
        
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"** Concreto Armado (Base)** ({moneda} {costo_base:,.0f})")
            m1,m2,m3,m4,m5 = st.columns(5)
            m1.metric("Cemento", f"{bol_pca} {und_bol}"); m2.metric("Arena", f"{arena_pca:.2f}m³"); m3.metric("Grava", f"{grava_pca:.2f}m³")
            m4.metric("Agua", f"{agua_pca:.0f} lt"); m5.metric("Acero", f"{kg_ace:.1f} kg")
            st.caption(f"Concreto: {vol_pca:.2f}m³ ({fc_pca} kg/cm²) | Acero: {pct_acero}% vert/horiz")
            
        with c2:
            st.markdown(f"** Acabados** ({moneda} {costo_acabados:,.0f})")
            a1,a2,a3,a4 = st.columns(4)
            a1.metric(f"Cem. {nom_rep}", f"{bol_rep} {und_bol}"); a2.metric("Cem. Afinado", f"{bol_afin} {und_bol}")
            a3.metric("Arena", f"{arena_rep+arena_afin:.2f}m³"); a4.metric("Agua", f"{agua_rep+agua_afin:.0f} lt")

        st.info(f" Costo Total Estimado: **{moneda} {costo_base+costo_acabados:,.0f}**")
        with st.expander(" Precios unitarios aplicados", expanded=False):
            _pr = st.session_state.kc_precios
            st.markdown(f"""<div style="display:flex;flex-wrap:wrap;gap:8px;">
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Cemento/bolsa</div><b>{moneda} {_pr.get('cemento',0):,.0f}</b></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Arena/m³</div><b>{moneda} {_pr.get('arena',0):,.0f}</b></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Grava/m³</div><b>{moneda} {_pr.get('grava',0):,.0f}</b></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Acero/kg</div><b>{moneda} {_pr.get('acero_kg',0):,.0f}</b></div>
  <div style="background:#1a1a0d;border:1px solid #6d6000;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#ffcc80;font-size:11px;"> TOTAL</div><b style="color:#ffdd57;">{moneda} {costo_base+costo_acabados:,.0f}</b></div>
</div><div style="margin-top:5px;font-size:10px;color:#ff8f00;">⚠ Moneda: <b>{moneda}</b></div>""", unsafe_allow_html=True)
        if st.button(" Agregar al Resumen",key="kc_add_pca",type="primary"):
            lb=desc_pared or f"Muro CA Por%"
            st.session_state.kc_rows.extend([
                {"elemento":f"{lb} — Cemento Base","unidad":und_bol,"cant":bol_pca,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Cemento Acabados","unidad":und_bol,"cant":bol_rep+bol_afin,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Arena","unidad":"m³","cant":round(arena_pca+arena_rep+arena_afin,3),"precio":_precio("arena")},
                {"elemento":f"{lb} — Grava (Base)","unidad":"m³","cant":round(grava_pca,3),"precio":_precio("grava")},
                {"elemento":f"{lb} — Acero {var_pca}","unidad":"kg","cant":round(kg_ace,1),"precio":_precio("acero_kg")},
            ]); st.success(f" Añadido: {bol_pca} bultos base + {bol_rep+bol_afin} acabados + {kg_ace:.1f} kg acero")

    # ========= DIMENSIÓN (Concreto Armado detallado) =========
    elif modo_pared == " Dimensión (CA)":
        d1,d2,d3=st.columns(3)
        with d1:
            st.markdown("**Acero Vertical**")
            sep_v = st.number_input("Separación [m]",0.0,value=0.40,step=0.05,key="kc_dim_sepv")
            anc_sup = st.number_input("Anclaje superior [m]",0.0,value=0.40,step=0.05,key="kc_dim_ancsup")
            anc_inf = st.number_input("Anclaje inferior [m]",0.0,value=0.40,step=0.05,key="kc_dim_ancinf")
            var_v = st.selectbox("Ø Vertical:",[v["nombre"] for v in VARILLAS],index=4,key="kc_dim_varv")
            st.markdown("**Acero Horizontal**")
            sep_h = st.number_input("Separación [m]",0.0,value=0.40,step=0.05,key="kc_dim_seph")
            anc_ext1 = st.number_input("Anclaje extremo 1 [m]",0.0,value=0.40,step=0.05,key="kc_dim_ance1")
            anc_ext2 = st.number_input("Anclaje extremo 2 [m]",0.0,value=0.40,step=0.05,key="kc_dim_ance2")
            var_h = st.selectbox("Ø Horizontal:",[v["nombre"] for v in VARILLAS],index=2,key="kc_dim_varh")
        with d2:
            st.markdown("**Concreto**")
            alt_dim = st.number_input("Alto [m]",0.0,value=2.0,step=0.1,key="kc_dim_alt")
            lar_dim = st.number_input("Largo [m]",0.0,value=5.0,step=0.5,key="kc_dim_lar")
            esp_dim = st.number_input("Espesor [m]",0.0,value=0.20,step=0.05,key="kc_dim_esp")
            fc_dim  = st.selectbox("Resistencia [kg/cm²]:",[m["fc_kgcm2"] for m in MIX_DESIGNS],index=3,key="kc_dim_fc")
            rec_dim = st.number_input("Recubrimiento [m]",0.0,value=0.05,step=0.01,key="kc_dim_rec")
            cant_dim = st.number_input("Cantidad",min_value=1,value=1,key="kc_dim_qty")
        with d3:
            st.markdown("**Traslapes**")
            trasl_v = st.radio("Traslape vertical:",["NO","SI"],horizontal=True,key="kc_dim_tv")
            long_tv = st.number_input("Longitud traslape V [m]",0.0,value=0.50,key="kc_dim_ltv") if trasl_v=="SI" else 0
            trasl_h = st.radio("Traslape horizontal:",["NO","SI"],horizontal=True,key="kc_dim_th")
            long_th = st.number_input("Longitud traslape H [m]",0.0,value=0.50,key="kc_dim_lth") if trasl_h=="SI" else 0
            st.markdown("**Desperdicio**")
            desp_conc_dim = st.slider("Concreto [%]",0,15,5,key="kc_dim_dc")/100
            desp_ace_dim  = st.slider("Acero [%]",0,15,5,key="kc_dim_da")/100
            
        dos_rep, esp_rep, esp_afin, incluir, desp_acab = ui_acabados("kc_dim")
        
        # Concrete
        area_dim_t = alt_dim * lar_dim * cant_dim
        mix_dim = next(m for m in MIX_DESIGNS if m["fc_kgcm2"]==fc_dim)
        vol_dim = area_dim_t * esp_dim * (1+desp_conc_dim)
        bol_dim = math.ceil(mix_dim["cem_kg"]*vol_dim/R["peso_bolsa"])
        arena_dim=mix_dim["arena_m3"]*vol_dim; grava_dim=mix_dim["grava_m3"]*vol_dim; agua_dim=mix_dim["agua_lt"]*vol_dim
        # Rebar vertical
        vkd_v = next(v for v in VARILLAS if v["nombre"]==var_v)
        vkd_h = next(v for v in VARILLAS if v["nombre"]==var_h)
        n_barras_v = math.ceil(lar_dim / sep_v) + 1
        long_v_bar = alt_dim + anc_sup + anc_inf + long_tv
        kg_v = n_barras_v * long_v_bar * vkd_v["kg_m"] * cant_dim * (1+desp_ace_dim)
        n_barras_h = math.ceil(alt_dim / sep_h) + 1
        long_h_bar = lar_dim + anc_ext1 + anc_ext2 + long_th
        kg_h = n_barras_h * long_h_bar * vkd_h["kg_m"] * cant_dim * (1+desp_ace_dim)
        kg_total_dim = kg_v + kg_h
        
        bol_rep, arena_rep, agua_rep, bol_afin, arena_afin, agua_afin = calc_acabados(area_dim_t, dos_rep, esp_rep, esp_afin, incluir, desp_acab)
        
        st.markdown("---")
        und_bol = "bulto" if R.get("pais") == "Colombia" else "bol"
        nom_rep = R.get("pañete", "Repello").split("/")[0].strip()
        costo_base = bol_dim*_precio("cemento") + arena_dim*_precio("arena") + grava_dim*_precio("grava") + agua_dim*_precio("agua",0) + kg_total_dim*_precio("acero_kg")
        costo_acabados = (bol_rep+bol_afin)*_precio("cemento") + (arena_rep+arena_afin)*_precio("arena") + (agua_rep+agua_afin)*_precio("agua",0)
        
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"** Concreto Armado (Base)** ({moneda} {costo_base:,.0f})")
            m1,m2,m3,m4,m5 = st.columns(5)
            m1.metric("Cemento", f"{bol_dim} {und_bol}"); m2.metric("Arena", f"{arena_dim:.2f}m³"); m3.metric("Grava", f"{grava_dim:.2f}m³")
            m4.metric("Agua", f"{agua_dim:.0f} lt"); m5.metric("Acero", f"{kg_total_dim:.1f} kg")
            st.caption(f"V: {n_barras_v} barras × {long_v_bar:.2f}m ({kg_v:.1f} kg) | H: {n_barras_h} barras × {long_h_bar:.2f}m ({kg_h:.1f} kg)")
            
        with c2:
            st.markdown(f"** Acabados** ({moneda} {costo_acabados:,.0f})")
            a1,a2,a3,a4 = st.columns(4)
            a1.metric(f"Cem. {nom_rep}", f"{bol_rep} {und_bol}"); a2.metric("Cem. Afinado", f"{bol_afin} {und_bol}")
            a3.metric("Arena", f"{arena_rep+arena_afin:.2f}m³"); a4.metric("Agua", f"{agua_rep+agua_afin:.0f} lt")
            
        st.info(f" Costo Total Estimado: **{moneda} {costo_base+costo_acabados:,.0f}**")
        with st.expander(" Precios unitarios aplicados", expanded=False):
            _pr = st.session_state.kc_precios
            st.markdown(f"""<div style="display:flex;flex-wrap:wrap;gap:8px;">
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Cemento/bolsa</div><b>{moneda} {_pr.get('cemento',0):,.0f}</b></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Arena/m³</div><b>{moneda} {_pr.get('arena',0):,.0f}</b></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Grava/m³</div><b>{moneda} {_pr.get('grava',0):,.0f}</b></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Acero/kg</div><b>{moneda} {_pr.get('acero_kg',0):,.0f}</b></div>
  <div style="background:#1a1a0d;border:1px solid #6d6000;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#ffcc80;font-size:11px;"> TOTAL</div><b style="color:#ffdd57;">{moneda} {costo_base+costo_acabados:,.0f}</b></div>
</div><div style="margin-top:5px;font-size:10px;color:#ff8f00;">⚠ Moneda: <b>{moneda}</b></div>""", unsafe_allow_html=True)
        if st.button(" Agregar al Resumen",key="kc_add_dim",type="primary"):
            lb=desc_pared or "Muro CA Dimensión"
            st.session_state.kc_rows.extend([
                {"elemento":f"{lb} — Cemento Base","unidad":und_bol,"cant":bol_dim,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Cemento Acabados","unidad":und_bol,"cant":bol_rep+bol_afin,"precio":_precio("cemento")},
                {"elemento":f"{lb} — Arena","unidad":"m³","cant":round(arena_dim+arena_rep+arena_afin,3),"precio":_precio("arena")},
                {"elemento":f"{lb} — Grava (Base)","unidad":"m³","cant":round(grava_dim,3),"precio":_precio("grava")},
                {"elemento":f"{lb} — Acero vert. {var_v}","unidad":"kg","cant":round(kg_v,1),"precio":_precio("acero_kg")},
                {"elemento":f"{lb} — Acero horiz. {var_h}","unidad":"kg","cant":round(kg_h,1),"precio":_precio("acero_kg")},
            ]); st.success(f" Añadido: {bol_dim} bultos base + {bol_rep+bol_afin} acabados + {kg_total_dim:.1f} kg acero")

    # ========= PANEL YESO =========
    elif modo_pared == " Panel Yeso":
        CANAL_DIMS = {"Canal 2.44m (std)":2.44,"Canal 1.22m":1.22,"Canal 3.00m":3.00}
        PANEL_TIPOS = {"Panel 1.22×2.44m":{"area":2.98,"ancho":1.22},"Panel 1.22×3.00m":{"area":3.66,"ancho":1.22},
                       "Panel 1.20×2.40m":{"area":2.88,"ancho":1.20}}
        py1,py2=st.columns(2)
        with py1:
            alt_py = st.number_input("Alto [m]",0.0,value=2.44,step=0.1,key="kc_py_alt")
            lar_py = st.number_input("Largo [m]",0.0,value=5.0,step=0.5,key="kc_py_lar")
            cant_py = st.number_input("Cantidad muros",min_value=1,value=1,key="kc_py_qty")
            canal_tipo = st.selectbox("Dimensión Canal:",list(CANAL_DIMS.keys()),key="kc_py_canal")
            canal_long = CANAL_DIMS[canal_tipo]
        with py2:
            sep_parante = st.selectbox("Separación Parantes [m]",[0.40,0.48,0.60,1.22],index=2,key="kc_py_sepp")
            panel_tipo  = st.selectbox("Tipo de panel:",list(PANEL_TIPOS.keys()),key="kc_py_panel"); pd = PANEL_TIPOS[panel_tipo]
            doble = st.radio("Panel:",["Simple","Doble"],index=1,horizontal=True,key="kc_py_doble")
            desp_py  = st.slider("Desperdicio panel [%]",0,20,4,key="kc_py_dp")/100
            desp_pyo = st.slider("Desperdicio otros [%]",0,20,4,key="kc_py_do")/100
        area_py = alt_py*lar_py*cant_py; perim_py = (alt_py+lar_py)*2*cant_py
        # Paneles
        n_paneles = math.ceil(area_py/pd["area"]*(2 if doble=="Doble" else 1)*(1+desp_py))
        # Canales (piso+techo = 2 * largo / longitud_canal)
        n_canales = math.ceil(2*lar_py*cant_py/canal_long*(1+desp_pyo))
        # Parantes (altura / separacion)
        n_parantes = math.ceil((lar_py*cant_py/sep_parante+1)*(1+desp_pyo))
        # Masilla: 1 cubeta / 20 m²
        n_masilla = math.ceil(area_py/20*(2 if doble=="Doble" else 1)*(1+desp_pyo))
        # Tornillos estructura: 5/m², tornillos panel: 3/m²
        t_estr = math.ceil(area_py*5*(1+desp_pyo)); t_pan = math.ceil(area_py*3*(1+desp_pyo))
        costo_py = n_paneles*_precio("ceramica")*2 + n_canales*_precio("grava")*0.1 + n_masilla*_precio("pintura")*0.5
        st.markdown("---")
        rpy1,rpy2,rpy3,rpy4,rpy5,rpy6=st.columns(6)
        rpy1.metric(" Paneles",f"{n_paneles} und"); rpy2.metric("⚙ Canales",f"{n_canales} und"); rpy3.metric(" Parantes",f"{n_parantes} und")
        rpy4.metric(" Masilla",f"{n_masilla} cubetas"); rpy5.metric(" T. Estructura",f"{t_estr} und"); rpy6.metric(" T. Panel",f"{t_pan} und")
        st.info(f" Área: {area_py:.2f} m² | Perímetro: {perim_py:.2f} m | Tipo: {doble}")
        if st.button(" Agregar Panel Yeso al Resumen",key="kc_add_py",type="primary"):
            lb=desc_pared or "Panel Yeso"
            st.session_state.kc_rows.extend([
                {"elemento":f"{lb} — Paneles ({panel_tipo})","unidad":"und","cant":n_paneles,"precio":_precio("ceramica")*2},
                {"elemento":f"{lb} — Canales","unidad":"und","cant":n_canales,"precio":_precio("grava")*0.1},
                {"elemento":f"{lb} — Parantes","unidad":"und","cant":n_parantes,"precio":_precio("grava")*0.08},
                {"elemento":f"{lb} — Masilla","unidad":"cubeta","cant":n_masilla,"precio":_precio("pintura")*0.5},
                {"elemento":f"{lb} — Tornillos estructura","unidad":"und","cant":t_estr,"precio":50},
                {"elemento":f"{lb} — Tornillos panel","unidad":"und","cant":t_pan,"precio":30},
            ]); st.success(f" {n_paneles} paneles + {n_canales} canales + {n_parantes} parantes")

    # ========= PINTURA =========
    elif modo_pared == " Pintura":
        pp1,pp2=st.columns(2)
        with pp1:
            area_pp = st.number_input("Área [m²]",min_value=0.0,value=0.0,step=1.0,key="kc_pp_area")
            rend_pp = st.number_input("Rendimiento [m²/lt]",min_value=1.0,value=8.0,step=0.5,key="kc_pp_rend")
            manos_pp = st.selectbox("Manos:",[1,2,3,4],index=2,key="kc_pp_manos")
            lados_pp = st.selectbox("Lados:",[1,2],index=0,key="kc_pp_lados")
            desp_pp  = st.selectbox("Desperdicio [%]",[1,2,3,4,5,7,10],index=2,key="kc_pp_desp")
            cant_pp  = st.number_input("Cantidad elementos",min_value=1,value=1,key="kc_pp_qty")
        with pp2:
            st.markdown("**Otros Materiales (libre)**")
            extras = []
            for ei in range(1,7):
                exc,eqc,epc = st.columns([3,1,1])
                en = exc.text_input(f"Extra {ei}",key=f"kc_pp_en{ei}",label_visibility="collapsed",placeholder=f"Extra {ei}")
                eq = eqc.number_input("Cant",min_value=0.0,value=0.0,key=f"kc_pp_eq{ei}",label_visibility="collapsed")
                epp = epc.number_input("PU",min_value=0.0,value=0.0,key=f"kc_pp_ep{ei}",label_visibility="collapsed")
                if en and eq>0: extras.append({"nombre":en,"cant":eq,"pu":epp})
        lt_pp = math.ceil(area_pp*manos_pp*lados_pp*cant_pp/rend_pp*(1+desp_pp/100))
        costo_pp = lt_pp*_precio("pintura")/4 + sum(e["cant"]*e["pu"] for e in extras)
        st.markdown("---")
        rpp1,rpp2,rpp3=st.columns(3)
        rpp1.metric(f" {R['pintura']}",f"{lt_pp} lt",f"{manos_pp} manos · {lados_pp} lado(s)")
        rpp2.metric(" Área total",f"{area_pp*cant_pp:.1f} m²"); rpp3.metric(" Costo",f"{moneda} {costo_pp:,.0f}")
        if st.button(" Agregar Pintura al Resumen",key="kc_add_pp",type="primary"):
            lb=desc_pared or "Pintura"
            rows_pp=[{"elemento":f"{lb} — {R['pintura']}","unidad":"lt","cant":lt_pp,"precio":_precio("pintura")/4}]
            for e in extras: rows_pp.append({"elemento":f"{lb} — {e['nombre']}","unidad":"und","cant":e["cant"],"precio":e["pu"]})
            st.session_state.kc_rows.extend(rows_pp); st.success(f" {lt_pp} lt de pintura + {len(extras)} extras")


# ══════════ TAB 3 — COLUMNA (7 secciones CM-V3.0) ══════════
with tabs[2]:
    st.subheader(" Calculadora de Columna")
    st.caption(f"CM-V3.0 | 7 tipos de sección | {norma_sel}")

    SECCION_TIPOS = [
        "■ Sec.1  V1·E.1","■■ Sec.2  V1+V2·E.1","■■ Sec.3  V1+V2·E.1+E.2",
        "■■ Sec.4  V1+V2·E.1 cruzado","■■■ Sec.5  V1+V2·E.1+E.2+E.3",
        "n  Ilimitado","□ Castillo",
    ]
    # Imágenes SVG inline para cada tipo de sección
    _SEC_SVG = {
        "■ Sec.1  V1·E.1": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <circle cx='9' cy='9' r='3.5' fill='#ff6b35'/><circle cx='45' cy='9' r='3.5' fill='#ff6b35'/>
          <circle cx='9' cy='45' r='3.5' fill='#ff6b35'/><circle cx='45' cy='45' r='3.5' fill='#ff6b35'/>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='4,3'/>
        </svg>""",
        "■■ Sec.2  V1+V2·E.1": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <circle cx='9' cy='9' r='3.5' fill='#ff6b35'/><circle cx='45' cy='9' r='3.5' fill='#ff6b35'/>
          <circle cx='9' cy='45' r='3.5' fill='#ff6b35'/><circle cx='45' cy='45' r='3.5' fill='#ff6b35'/>
          <circle cx='9' cy='27' r='3.5' fill='#ffd700'/><circle cx='45' cy='27' r='3.5' fill='#ffd700'/>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='4,3'/>
        </svg>""",
        "■■ Sec.3  V1+V2·E.1+E.2": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <circle cx='9' cy='9' r='3.5' fill='#ff6b35'/><circle cx='45' cy='9' r='3.5' fill='#ff6b35'/>
          <circle cx='9' cy='45' r='3.5' fill='#ff6b35'/><circle cx='45' cy='45' r='3.5' fill='#ff6b35'/>
          <circle cx='9' cy='27' r='3.5' fill='#ffd700'/><circle cx='45' cy='27' r='3.5' fill='#ffd700'/>
          <line x1='27' y1='9' x2='27' y2='45' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='3,2'/>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='4,3'/>
        </svg>""",
        "■■ Sec.4  V1+V2·E.1 cruzado": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <circle cx='9' cy='9' r='3.5' fill='#ff6b35'/><circle cx='45' cy='9' r='3.5' fill='#ff6b35'/>
          <circle cx='9' cy='45' r='3.5' fill='#ff6b35'/><circle cx='45' cy='45' r='3.5' fill='#ff6b35'/>
          <circle cx='9' cy='27' r='3.5' fill='#ffd700'/><circle cx='45' cy='27' r='3.5' fill='#ffd700'/>
          <circle cx='27' cy='9' r='3.5' fill='#ffd700'/><circle cx='27' cy='45' r='3.5' fill='#ffd700'/>
          <line x1='27' y1='9' x2='27' y2='45' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='3,2'/>
          <line x1='9' y1='27' x2='45' y2='27' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='3,2'/>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='4,3'/>
        </svg>""",
        "■■■ Sec.5  V1+V2·E.1+E.2+E.3": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <circle cx='9' cy='9' r='3.5' fill='#ff6b35'/><circle cx='45' cy='9' r='3.5' fill='#ff6b35'/>
          <circle cx='9' cy='45' r='3.5' fill='#ff6b35'/><circle cx='45' cy='45' r='3.5' fill='#ff6b35'/>
          <circle cx='9' cy='27' r='3.5' fill='#ffd700'/><circle cx='45' cy='27' r='3.5' fill='#ffd700'/>
          <circle cx='27' cy='9' r='3.5' fill='#ffd700'/><circle cx='27' cy='45' r='3.5' fill='#ffd700'/>
          <polygon points='27,9 45,27 27,45 9,27' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='3,2'/>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='4,3'/>
        </svg>""",
        "n  Ilimitado": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='2' y='2' width='50' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <text x='27' y='34' text-anchor='middle' font-size='22' font-weight='bold' fill='#90caf9' font-family='serif'>n</text>
        </svg>""",
        "□ Castillo": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='6' y='6' width='42' height='42' rx='3' fill='none' stroke='#4a9eff' stroke-width='2.5'/>
          <rect x='12' y='12' width='30' height='30' rx='2' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='4,3'/>
          <circle cx='13' cy='13' r='3' fill='#ff6b35'/><circle cx='41' cy='13' r='3' fill='#ff6b35'/>
          <circle cx='13' cy='41' r='3' fill='#ff6b35'/><circle cx='41' cy='41' r='3' fill='#ff6b35'/>
        </svg>""",
    }
    # Selector visual con SVG
    st.markdown("**Tipo de sección:**")
    _cols_sec = st.columns(len(SECCION_TIPOS))
    if "kc_col_sec" not in st.session_state:
        st.session_state["kc_col_sec"] = SECCION_TIPOS[0]
    for _i, _stype in enumerate(SECCION_TIPOS):
        with _cols_sec[_i]:
            _selected = st.session_state.get("kc_col_sec") == _stype
            _border = "3px solid #4a9eff" if _selected else "1px solid #333"
            _bg = "#0d1f3c" if _selected else "#111"
            _lbl = _stype.split()[1] if len(_stype.split()) > 1 else _stype
            st.markdown(
                f"""<div style='text-align:center;background:{_bg};border:{_border};border-radius:8px;padding:4px 2px;cursor:pointer;'>
                {_SEC_SVG[_stype]}<br>
                <span style='font-size:0.65rem;color:#ccc;'>{_lbl}</span></div>""",
                unsafe_allow_html=True
            )
            if st.button("" if _selected else "○", key=f"kc_sec_btn_{_i}",
                         help=_stype, use_container_width=True):
                st.session_state["kc_col_sec"] = _stype
                st.rerun()
    sec_tipo = st.session_state.get("kc_col_sec", SECCION_TIPOS[0])
    st.markdown("---")
    
    # ======= GRÁFICO DE ELEVACIÓN DINÁMICO =======
    _html_elevation = f"""
    <div style="background:#0a192f; padding: 20px; border-radius: 8px; border: 1px solid #1e3a5f; margin: 15px 0; display:flex; justify-content:space-around; align-items:center; flex-wrap:wrap; box-shadow: 0 4px 15px rgba(0,0,0,0.3);">
       <div style="text-align:center; min-width: 150px;">
          <h4 style="color:#90caf9; margin-bottom:20px; font-size:1.05rem; letter-spacing: 0.5px;">Sección Transversal</h4>
          <div style="transform: scale(2.8); transform-origin: top center; margin-bottom: 95px;">
             {_SEC_SVG[sec_tipo]}
          </div>
       </div>
       <div style="text-align:center; min-width: 250px;">
          <h4 style="color:#90caf9; margin-bottom:15px; font-size:1.05rem; letter-spacing: 0.5px;">Elevación ({'Uniforme' if sec_tipo == '□ Castillo' else 'NSR-10 / ACI'})</h4>
          <svg width="280" height="340" viewBox="0 0 280 340" xmlns="http://www.w3.org/2000/svg">
             <!-- Losas superior e inferior -->
             <rect x="80" y="10" width="120" height="30" fill="#2d4a6e" opacity="0.6"/>
             <rect x="80" y="300" width="120" height="30" fill="#2d4a6e" opacity="0.6"/>
             <!-- Concreto columna -->
             <rect x="100" y="10" width="80" height="320" fill="#1e3a5f" stroke="#4a9eff" stroke-width="2"/>
             <!-- Acero Longitudinal Base -->
             <line x1="112" y1="2" x2="112" y2="338" stroke="#ff6b35" stroke-width="2.5"/>
             <line x1="168" y1="2" x2="168" y2="338" stroke="#ff6b35" stroke-width="2.5"/>
"""
    if "V2" in sec_tipo or sec_tipo.startswith("■■■"):
        _html_elevation += '<line x1="140" y1="2" x2="140" y2="338" stroke="#ffd700" stroke-width="2.5"/>'
    if sec_tipo == "□ Castillo":
        for y in range(40, 300, 15):
            _html_elevation += f'<line x1="112" y1="{y}" x2="168" y2="{y}" stroke="#00d4ff" stroke-width="1.5"/>'
        _html_elevation += '<text x="70" y="175" font-size="12" fill="#ccc" text-anchor="end">Estribos uniformes</text>'
        _html_elevation += '<path d="M75 170 L95 170" stroke="#ccc" stroke-width="1"/>'
    else:
        # Nudo Sup
        for y in range(15, 40, 6): _html_elevation += f'<line x1="112" y1="{y}" x2="168" y2="{y}" stroke="#00d4ff" stroke-width="1.5" opacity="0.5"/>'
        _html_elevation += '<text x="75" y="25" font-size="12" fill="#fff" text-anchor="end">Nudo Sup.</text>'
        # ZC Sup
        for y in range(45, 95, 8): _html_elevation += f'<line x1="112" y1="{y}" x2="168" y2="{y}" stroke="#00d4ff" stroke-width="1.5"/>'
        _html_elevation += '<text x="75" y="72" font-size="12" fill="#fff" text-anchor="end">ZC (Confinada)</text>'
        _html_elevation += '<path d="M 80 68 L 95 68" stroke="#fff" stroke-width="1" opacity="0.5"/>'
        # Zona Central
        for y in range(105, 210, 15): _html_elevation += f'<line x1="112" y1="{y}" x2="168" y2="{y}" stroke="#00d4ff" stroke-width="1.5"/>'
        _html_elevation += '<text x="75" y="160" font-size="12" fill="#fff" text-anchor="end">Zona Central</text>'
        _html_elevation += '<path d="M 80 156 L 95 156" stroke="#fff" stroke-width="1" opacity="0.5"/>'
        # Traslapes (Splice)
        _html_elevation += '<path d="M 168 150 L 175 155 L 175 180" stroke="#ff6b35" stroke-width="2" fill="none"/>'
        _html_elevation += '<text x="195" y="166" font-size="12" fill="#ffb74d">Traslape</text>'
        _html_elevation += '<line x1="172" y1="162" x2="190" y2="162" stroke="#ffb74d" stroke-dasharray="2,2" stroke-width="1"/>'
        # ZC Inf
        for y in range(215, 265, 8): _html_elevation += f'<line x1="112" y1="{y}" x2="168" y2="{y}" stroke="#00d4ff" stroke-width="1.5"/>'
        _html_elevation += '<text x="75" y="240" font-size="12" fill="#fff" text-anchor="end">ZC (Confinada)</text>'
        _html_elevation += '<path d="M 80 236 L 95 236" stroke="#fff" stroke-width="1" opacity="0.5"/>'
        # Nudo Inf
        for y in range(270, 295, 6): _html_elevation += f'<line x1="112" y1="{y}" x2="168" y2="{y}" stroke="#00d4ff" stroke-width="1.5" opacity="0.5"/>'
        _html_elevation += '<text x="75" y="285" font-size="12" fill="#fff" text-anchor="end">Nudo Inf.</text>'
        # Anclajes Gancho
        _html_elevation += '<path d="M 112 2 L 130 2" stroke="#ff6b35" stroke-width="2.5" fill="none"/>'
        _html_elevation += '<path d="M 168 338 L 150 338" stroke="#ff6b35" stroke-width="2.5" fill="none"/>'
        _html_elevation += '<text x="185" y="15" font-size="12" fill="#ffb74d">Anclaje Sup.</text>'
        _html_elevation += '<text x="185" y="335" font-size="12" fill="#ffb74d">Anclaje Inf.</text>'
    _html_elevation += """
          </svg>
       </div>
    </div>
    """
    st.markdown(_html_elevation, unsafe_allow_html=True)
    
    desc_col = st.text_input("Descripción", placeholder="Ej: Columna C1", key="kc_col_desc")

    _N_BARS = {
        "■ Sec.1  V1·E.1":(4,0),"■■ Sec.2  V1+V2·E.1":(4,2),
        "■■ Sec.3  V1+V2·E.1+E.2":(4,2),"■■ Sec.4  V1+V2·E.1 cruzado":(4,4),
        "■■■ Sec.5  V1+V2·E.1+E.2+E.3":(4,4),"n  Ilimitado":(0,0),"□ Castillo":(4,0),
    }
    n_v1_default, n_v2_default = _N_BARS[sec_tipo]

    col_i, col_r, col_z = st.columns(3)
    with col_i:
        st.markdown("** Dimensiones Columna**")
        alt_col  = st.number_input("Alto [m]",0.0,value=7.0,step=0.1,key="kc_col_alt")
        a_col    = st.number_input("a [m]",0.0,value=0.30,step=0.05,key="kc_col_a")
        b_col    = st.number_input("b [m]",0.0,value=0.30,step=0.05,key="kc_col_b")
        fc_col   = st.selectbox("Resistencia [kg/cm²]:",[m["fc_kgcm2"] for m in MIX_DESIGNS],index=3,key="kc_col_fc")
        cant_col = st.number_input("Cantidad",min_value=1,value=1,key="kc_col_qty")
        mix_col  = next(m for m in MIX_DESIGNS if m["fc_kgcm2"]==fc_col)
        vol_col_neto = a_col*b_col*alt_col
        st.caption(f"Vol neto: {vol_col_neto:.2f}m³ | Dosif: {mix_col['dos']}")
        st.markdown("** Desperdicios**")
        desp_conc_col = st.select_slider("Concreto [%]",[1,2,3,4,5,6,7,8,10],value=5,key="kc_col_dc")
        desp_long_col = st.select_slider("Acero Long. [%]",[1,2,3,4,5,6,7,8,10],value=5,key="kc_col_dl")
        desp_est_col  = st.select_slider("Acero Estrib. [%]",[1,2,3,4,5,6,7,8,10],value=3,key="kc_col_de")

    with col_r:
        st.markdown("** Acero Longitudinal**")
        lt_col   = st.number_input("Longitud barra [m]",0.0,value=alt_col,step=0.1,key="kc_col_lt")
        trasl_col = st.radio("Traslapes:",["NO","SI"],horizontal=True,key="kc_col_trasl")
        long_trasl = st.number_input("Long. traslape [m]",0.0,value=0.35,step=0.05,key="kc_col_ltrasl") if trasl_col=="SI" else 0.0
        if sec_tipo=="n  Ilimitado":
            n_v1 = st.number_input("N° barras V1",min_value=1,value=4,step=1,key="kc_col_nv1n"); n_v2=0
        else:
            n_v1=n_v1_default; n_v2=n_v2_default
        var_v1 = st.selectbox("Ø V1:",[v["nombre"] for v in VARILLAS],index=3,key="kc_col_v1"); vkd_v1=next(v for v in VARILLAS if v["nombre"]==var_v1)
        if n_v2>0:
            var_v2=st.selectbox("Ø V2:",[v["nombre"] for v in VARILLAS],index=3,key="kc_col_v2"); vkd_v2=next(v for v in VARILLAS if v["nombre"]==var_v2)
        else:
            var_v2=None; vkd_v2=None
        rec_col = st.number_input("Recubrimiento [m]",0.0,value=0.05,step=0.01,key="kc_col_rec")

    with col_z:
        if sec_tipo!="□ Castillo":
            st.markdown("** Zonas**")
            zona_conf = st.number_input("Confinada ZC [m]",0.0,value=2.0,step=0.1,key="kc_col_zc")
            zona_cent = st.number_input("Central [m]",0.0,value=3.0,step=0.1,key="kc_col_zce")
            nudo_sup  = st.number_input("Nudo sup. [m]",0.0,value=0.50,step=0.05,key="kc_col_ns")
            nudo_inf  = st.number_input("Nudo inf. [m]",0.0,value=0.50,step=0.05,key="kc_col_ni")
        st.markdown("**⚙ Estribos E.1**")
        long_e1 = st.number_input("Long. E.1 [m]",0.0,value=round(2*(a_col+b_col)-8*max(rec_col,0.05)+0.30,3),step=0.01,key="kc_col_le1")
        var_e1  = st.selectbox("Ø E.1:",[v["nombre"] for v in VARILLAS],index=1,key="kc_col_ve1"); vkd_e1=next(v for v in VARILLAS if v["nombre"]==var_e1)
        if sec_tipo=="□ Castillo":
            sep_cast=st.number_input("Separación [m]",0.0,value=0.12,step=0.01,key="kc_col_scast")
        else:
            sep_ZC   = st.number_input("Sep. ZC [m]",0.0,value=0.12,step=0.01,key="kc_col_szc")
            sep_cent = st.number_input("Sep. Centro [m]",0.0,value=0.12,step=0.01,key="kc_col_scen")
            sep_nudo = st.number_input("Sep. Nudo [m]",0.0,value=0.12,step=0.01,key="kc_col_snud")
        if "E.2" in sec_tipo:
            st.markdown("**⚙ Estribos E.2**")
            long_e2=st.number_input("Long. E.2 [m]",0.0,value=round((a_col-2*max(rec_col,0.05))+0.20,3),step=0.01,key="kc_col_le2")
            var_e2=st.selectbox("Ø E.2:",[v["nombre"] for v in VARILLAS],index=1,key="kc_col_ve2"); vkd_e2=next(v for v in VARILLAS if v["nombre"]==var_e2)
        if "E.3" in sec_tipo:
            st.markdown("**⚙ Estribos E.3**")
            long_e3=st.number_input("Long. E.3 [m]",0.0,value=round((b_col-2*max(rec_col,0.05))+0.20,3),step=0.01,key="kc_col_le3")
            var_e3=st.selectbox("Ø E.3:",[v["nombre"] for v in VARILLAS],index=0,key="kc_col_ve3"); vkd_e3=next(v for v in VARILLAS if v["nombre"]==var_e3)

    # Calculations
    vol_conc = vol_col_neto*cant_col*(1+desp_conc_col/100)
    bol_col  = math.ceil(mix_col["cem_kg"]*vol_conc/R["peso_bolsa"])
    arena_col= mix_col["arena_m3"]*vol_conc; grava_col=mix_col["grava_m3"]*vol_conc; agua_col=mix_col["agua_lt"]*vol_conc
    long_bar = lt_col+long_trasl
    kg_v1 = n_v1*long_bar*vkd_v1["kg_m"]*cant_col*(1+desp_long_col/100)
    kg_v2 = (n_v2*long_bar*vkd_v2["kg_m"]*cant_col*(1+desp_long_col/100)) if vkd_v2 else 0
    kg_long = kg_v1+kg_v2
    if sec_tipo=="□ Castillo":
        n_est_e1 = math.floor(alt_col/max(sep_cast,0.01))+1
    else:
        n_est_e1 = math.floor(zona_conf/max(sep_ZC,0.01))*2 + math.floor(zona_cent/max(sep_cent,0.01))
    kg_e1 = n_est_e1*long_e1*vkd_e1["kg_m"]*cant_col*(1+desp_est_col/100)
    kg_e2 = (n_est_e1*long_e2*vkd_e2["kg_m"]*cant_col*(1+desp_est_col/100)) if "E.2" in sec_tipo else 0
    kg_e3 = (n_est_e1*long_e3*vkd_e3["kg_m"]*cant_col*(1+desp_est_col/100)) if "E.3" in sec_tipo else 0
    kg_est = kg_e1+kg_e2+kg_e3; kg_ace_total=kg_long+kg_est
    costo_col = bol_col*float(p.get("cemento",0))+arena_col*float(p.get("arena",0))+grava_col*float(p.get("grava",0))+kg_ace_total*float(p.get("acero_kg",0))

    # ── Costos unitarios ──
    _cp_cem  = float(p.get("cemento", 0))
    _cp_are  = float(p.get("arena", 0))
    _cp_gra  = float(p.get("grava", 0))
    _cp_ace  = float(p.get("acero_kg", 0))
    _costo_cem  = bol_col * _cp_cem
    _costo_are  = arena_col * _cp_are
    _costo_gra  = grava_col * _cp_gra
    _costo_long = kg_long * _cp_ace
    _costo_est  = kg_est  * _cp_ace

    st.markdown("---")
    st.markdown("####  Resumen de Resultados")
    rc1, rc2, rc3, rc4 = st.columns(4)

    with rc1:
        st.markdown(
            "<div style='text-align:center;font-size:1.8rem;'></div>"
            "<div style='text-align:center;font-weight:700;color:#90caf9;margin-bottom:6px;'>Materiales Concreto</div>",
            unsafe_allow_html=True)
        st.metric(f" {R.get('cemento','Cemento')}", f"{bol_col} bultos")
        st.metric(" Arena", f"{arena_col:.2f}m³")
        st.metric(" Grava", f"{grava_col:.2f}m³")
        st.metric(" Agua",  f"{agua_col:.0f} lt")

    with rc2:
        st.markdown(
            "<div style='text-align:center;font-size:1.8rem;'></div>"
            "<div style='text-align:center;font-weight:700;color:#ffcc80;margin-bottom:6px;'>Costos Concreto</div>",
            unsafe_allow_html=True)
        st.metric(" Cemento",  f"{moneda} {_costo_cem:,.0f}")
        st.metric(" Arena",   f"{moneda} {_costo_are:,.0f}")
        st.metric(" Grava",   f"{moneda} {_costo_gra:,.0f}")
        st.metric(" Dosif.",   mix_col["dos"])

    with rc3:
        st.markdown(
            "<div style='text-align:center;font-size:1.8rem;'></div>"
            "<div style='text-align:center;font-weight:700;color:#90caf9;margin-bottom:6px;'>Material Acero</div>",
            unsafe_allow_html=True)
        st.metric(f"V1 ({n_v1} br)", f"{kg_v1:.2f} kg")
        if n_v2: st.metric(f"V2 ({n_v2} br)", f"{kg_v2:.2f} kg")
        st.metric("Long. total",  f"{kg_long:.2f} kg")
        st.metric("N° E.1",       f"{n_est_e1}")
        st.metric("Estrib. total",f"{kg_est:.2f} kg")

    with rc4:
        st.markdown(
            "<div style='text-align:center;font-size:1.8rem;'></div>"
            "<div style='text-align:center;font-weight:700;color:#ffcc80;margin-bottom:6px;'>Costos Acero</div>",
            unsafe_allow_html=True)
        st.metric("Longitudinal",  f"{moneda} {_costo_long:,.0f}")
        st.metric("Estribos",      f"{moneda} {_costo_est:,.0f}")
        st.metric(" Vol. concreto", f"{vol_col_neto*cant_col:.2f}m³")
        st.metric(" COSTO TOTAL",   f"{moneda} {costo_col:,.0f}")

    st.markdown(
        f'<div style="background:#0d2137;border-radius:8px;padding:8px 16px;">'
        f'<span style="color:#ffcc80;">f\'c={fc_col} kg/cm² | {a_col}×{b_col} m | '
        f'h={alt_col} m | V1:{n_v1} brs + V2:{n_v2} brs | E.1:{n_est_e1} estrib.</span></div>',
        unsafe_allow_html=True)

    with st.expander(" Precios unitarios aplicados en Columna", expanded=False):
        _pr = st.session_state.kc_precios
        _cp_cem2 = float(_pr.get('cemento', 0))
        _cp_are2 = float(_pr.get('arena', 0))
        _cp_gra2 = float(_pr.get('grava', 0))
        _cp_ace2 = float(_pr.get('acero_kg', 0))
        st.markdown(f"""<div style="display:flex;flex-wrap:wrap;gap:8px;">
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Cemento/bolsa</div>
    <b>{moneda} {_cp_cem2:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{bol_col} bultos = {moneda} {bol_col*_cp_cem2:,.0f}</span></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Arena/m³</div>
    <b>{moneda} {_cp_are2:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{arena_col:.2f}m³ = {moneda} {arena_col*_cp_are2:,.0f}</span></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Grava/m³</div>
    <b>{moneda} {_cp_gra2:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{grava_col:.2f}m³ = {moneda} {grava_col*_cp_gra2:,.0f}</span></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Acero/kg</div>
    <b>{moneda} {_cp_ace2:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{kg_ace_total:.1f} kg = {moneda} {kg_ace_total*_cp_ace2:,.0f}</span></div>
  <div style="background:#1a1a0d;border:1px solid #6d6000;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#ffcc80;font-size:11px;"> COSTO TOTAL</div>
    <b style="color:#ffdd57;font-size:1.1em;">{moneda} {costo_col:,.0f}</b></div>
</div><div style="margin-top:5px;font-size:10px;color:#ff8f00;">⚠ Moneda: <b>{moneda}</b> | Norma: {norma_sel}</div>""", unsafe_allow_html=True)

    if st.button(" Agregar Columna al Resumen", key="kc_add_col", type="primary"):
        lb = desc_col or "Columna"
        st.session_state.kc_rows.extend([
            {"elemento": f"{lb} — {R.get('cemento','Cemento')}", "unidad": "bultos", "cant": bol_col,       "precio": _cp_cem},
            {"elemento": f"{lb} — Arena",                        "unidad": "m³",     "cant": round(arena_col,3), "precio": _cp_are},
            {"elemento": f"{lb} — Grava",                        "unidad": "m³",     "cant": round(grava_col,3), "precio": _cp_gra},
            {"elemento": f"{lb} — Acero long. {var_v1}",         "unidad": "kg",     "cant": round(kg_long,2),   "precio": _cp_ace},
            {"elemento": f"{lb} — Estribos {var_e1}",            "unidad": "kg",     "cant": round(kg_est,2),    "precio": _cp_ace},
        ])
        save_state()
        st.success(f" {lb}: {bol_col} bultos cemento | {kg_long:.1f} kg long. | {kg_est:.1f} kg estrib. | {moneda} {costo_col:,.0f}")


# ══════════ TAB 4 — VIGA (4 secciones CM-V3.0) ══════════
with tabs[3]:
    st.subheader(" Calculadora de Viga")
    st.caption(f"CM-V3.0 | 4 tipos | Acero +/- | {norma_sel}")

    SEC_VIGA = ["■ Sec.1  A+:2 A-:2", "■■ Sec.2  A+:3 A-:2", "■■■ Sec.3  A+:4 A-:2", "n  Ilimitado"]
    _VIGA_SVG = {
        "■ Sec.1  A+:2 A-:2": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='8' y='2' width='38' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <circle cx='15' cy='9' r='3.5' fill='#ff6b35'/><circle cx='39' cy='9' r='3.5' fill='#ff6b35'/>
          <circle cx='15' cy='45' r='3.5' fill='#ff6b35'/><circle cx='39' cy='45' r='3.5' fill='#ff6b35'/>
          <rect x='8' y='2' width='38' height='50' rx='3' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='4,3'/>
        </svg>""",
        "■■ Sec.2  A+:3 A-:2": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='8' y='2' width='38' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <circle cx='15' cy='9' r='3.5' fill='#ff6b35'/><circle cx='39' cy='9' r='3.5' fill='#ff6b35'/>
          <circle cx='15' cy='45' r='3.5' fill='#ff6b35'/><circle cx='27' cy='45' r='3.5' fill='#ff6b35'/><circle cx='39' cy='45' r='3.5' fill='#ff6b35'/>
          <rect x='8' y='2' width='38' height='50' rx='3' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='4,3'/>
        </svg>""",
        "■■■ Sec.3  A+:4 A-:2": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='8' y='2' width='38' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <circle cx='15' cy='9' r='3.5' fill='#ff6b35'/><circle cx='39' cy='9' r='3.5' fill='#ff6b35'/>
          <circle cx='13' cy='45' r='3' fill='#ff6b35'/><circle cx='22' cy='45' r='3' fill='#ff6b35'/><circle cx='32' cy='45' r='3' fill='#ff6b35'/><circle cx='41' cy='45' r='3' fill='#ff6b35'/>
          <rect x='8' y='2' width='38' height='50' rx='3' fill='none' stroke='#00d4ff' stroke-width='1.5' stroke-dasharray='4,3'/>
        </svg>""",
        "n  Ilimitado": """<svg width='54' height='54' viewBox='0 0 54 54' xmlns='http://www.w3.org/2000/svg'>
          <rect x='8' y='2' width='38' height='50' rx='3' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2.5'/>
          <text x='27' y='34' text-anchor='middle' font-size='22' font-weight='bold' fill='#90caf9' font-family='serif'>n</text>
        </svg>""",
    }

    st.markdown("**Tipo de sección:**")
    _cols_vsec = st.columns(len(SEC_VIGA))
    if "kc_viga_sec" not in st.session_state:
        st.session_state["kc_viga_sec"] = SEC_VIGA[0]
    for _i, _stype in enumerate(SEC_VIGA):
        with _cols_vsec[_i]:
            _selected = st.session_state.get("kc_viga_sec") == _stype
            _border = "3px solid #4a9eff" if _selected else "1px solid #333"
            _bg = "#0d1f3c" if _selected else "#111"
            _lbl = _stype.split()[1] if len(_stype.split()) > 1 else _stype
            st.markdown(
                f"""<div style='text-align:center;background:{_bg};border:{_border};border-radius:8px;padding:4px 2px;cursor:pointer;'>
                {_VIGA_SVG[_stype]}<br>
                <span style='font-size:0.65rem;color:#ccc;'>{_lbl}</span></div>""",
                unsafe_allow_html=True
            )
            if st.button("" if _selected else "○", key=f"kc_vsec_btn_{_i}",
                         help=_stype, use_container_width=True):
                st.session_state["kc_viga_sec"] = _stype
                st.rerun()
    sec_viga = st.session_state.get("kc_viga_sec", SEC_VIGA[0])
    st.markdown("---")

    # ======= GRÁFICO DE ELEVACIÓN DINÁMICO =======
    _html_viga_elevation = f"""
    <div style="background:#0a192f; padding: 20px; border-radius: 8px; border: 1px solid #1e3a5f; margin: 15px 0; display:flex; justify-content:space-around; align-items:center; flex-wrap:wrap; box-shadow: 0 4px 15px rgba(0,0,0,0.3);">
       <div style="text-align:center; min-width: 150px;">
          <h4 style="color:#90caf9; margin-bottom:20px; font-size:1.05rem; letter-spacing: 0.5px;">Sección Transversal</h4>
          <div style="transform: scale(2.8); transform-origin: top center; margin-bottom: 95px;">
             {_VIGA_SVG[sec_viga]}
          </div>
       </div>
       <div style="text-align:center; min-width: 350px;">
          <h4 style="color:#90caf9; margin-bottom:15px; font-size:1.05rem; letter-spacing: 0.5px;">Elevación (NSR-10 / ACI)</h4>
          <svg width="350" height="200" viewBox="0 0 350 200" xmlns="http://www.w3.org/2000/svg">
             <!-- Soportes / Columnas -->
             <rect x="20" y="70" width="30" height="80" fill="#2d4a6e" opacity="0.6"/>
             <rect x="300" y="70" width="30" height="80" fill="#2d4a6e" opacity="0.6"/>
             
             <!-- Concreto Viga -->
             <rect x="50" y="80" width="250" height="40" fill="#1e3a5f" stroke="#4a9eff" stroke-width="2"/>
             
             <!-- Acero negativo (top) -->
             <line x1="30" y1="88" x2="320" y2="88" stroke="#ff6b35" stroke-width="2.5"/>
             <path d="M 30 88 L 30 115" stroke="#ff6b35" stroke-width="2.5" fill="none"/>
             <path d="M 320 88 L 320 115" stroke="#ff6b35" stroke-width="2.5" fill="none"/>
             <text x="210" y="80" font-size="12" fill="#ff6b35">Acero-</text>
             
             <!-- Acero positivo (bottom) -->
             <line x1="38" y1="112" x2="312" y2="112" stroke="#ff6b35" stroke-width="2.5"/>
             <path d="M 38 112 L 38 85" stroke="#ff6b35" stroke-width="2.5" fill="none"/>
             <path d="M 312 112 L 312 85" stroke="#ff6b35" stroke-width="2.5" fill="none"/>
             <text x="210" y="128" font-size="12" fill="#ff6b35">Acero+</text>
             
             <!-- Traslape -->
             <path d="M 110 112 L 120 116 L 140 116" stroke="#ffb74d" stroke-width="2" fill="none"/>
             <text x="145" y="119" font-size="10" fill="#ffb74d">Traslape</text>
             
             <!-- Zonas -->
             <line x1="50" y1="55" x2="110" y2="55" stroke="#fff" stroke-width="1"/>
             <line x1="50" y1="50" x2="50" y2="60" stroke="#fff" stroke-width="1"/>
             <line x1="110" y1="50" x2="110" y2="60" stroke="#fff" stroke-width="1"/>
             <text x="80" y="45" font-size="11" fill="#fff" text-anchor="middle">ZC</text>
             
             <line x1="110" y1="55" x2="240" y2="55" stroke="#fff" stroke-width="1"/>
             <line x1="240" y1="50" x2="240" y2="60" stroke="#fff" stroke-width="1"/>
             <text x="175" y="45" font-size="11" fill="#fff" text-anchor="middle">Zona Central</text>
             
             <line x1="240" y1="55" x2="300" y2="55" stroke="#fff" stroke-width="1"/>
             <line x1="300" y1="50" x2="300" y2="60" stroke="#fff" stroke-width="1"/>
             <text x="270" y="45" font-size="11" fill="#fff" text-anchor="middle">ZC</text>
             
             <!-- Largo -->
             <line x1="50" y1="160" x2="300" y2="160" stroke="#ccc" stroke-width="1"/>
             <line x1="50" y1="155" x2="50" y2="165" stroke="#ccc" stroke-width="1"/>
             <line x1="300" y1="155" x2="300" y2="165" stroke="#ccc" stroke-width="1"/>
             <text x="175" y="175" font-size="11" fill="#ccc" text-anchor="middle">Largo</text>
             <line x1="54" y1="80" x2="54" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="62" y1="80" x2="62" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="70" y1="80" x2="70" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="78" y1="80" x2="78" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="86" y1="80" x2="86" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="94" y1="80" x2="94" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="102" y1="80" x2="102" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="116" y1="80" x2="116" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="131" y1="80" x2="131" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="146" y1="80" x2="146" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="161" y1="80" x2="161" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="176" y1="80" x2="176" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="191" y1="80" x2="191" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="206" y1="80" x2="206" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="221" y1="80" x2="221" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="236" y1="80" x2="236" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="244" y1="80" x2="244" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="252" y1="80" x2="252" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="260" y1="80" x2="260" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="268" y1="80" x2="268" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="276" y1="80" x2="276" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="284" y1="80" x2="284" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
             <line x1="292" y1="80" x2="292" y2="120" stroke="#00d4ff" stroke-width="1.5"/>
          </svg>
       </div>
    </div>
    """
    import streamlit.components.v1 as _stc
    _stc.html(_html_viga_elevation, height=280, scrolling=False)
    desc_viga = st.text_input("Descripción", placeholder="Ej: Viga V-101", key="kc_viga_desc")

    _NBV = {"■ Sec.1  A+:2 A-:2":(2,2),"■■ Sec.2  A+:3 A-:2":(3,2),"■■■ Sec.3  A+:4 A-:2":(4,2),"n  Ilimitado":(0,0)}
    nb_pos_def, nb_neg_def = _NBV[sec_viga]

    vi1, vi2, vi3 = st.columns(3)

    with vi1:
        st.markdown("** Dimensiones Viga**")
        largo_viga = st.number_input("Largo [m]", 0.0, value=4.0, step=0.1, key="kc_viga_L")
        a_viga = st.number_input("a (ancho) [m]", 0.0, value=0.30, step=0.05, key="kc_viga_a")
        b_viga = st.number_input("b (alto) [m]", 0.0, value=0.30, step=0.05, key="kc_viga_b")
        fc_viga = st.selectbox("Resistencia [kg/cm²]:", [m["fc_kgcm2"] for m in MIX_DESIGNS], index=3, key="kc_viga_fc")
        cant_viga = st.number_input("Cantidad", min_value=1, value=1, key="kc_viga_qty")
        mix_viga = next(m for m in MIX_DESIGNS if m["fc_kgcm2"] == fc_viga)
        rec_viga = st.number_input("Recubrimiento [m]", 0.0, value=0.05, step=0.01, key="kc_viga_rec")
        vol_viga_neto = a_viga * b_viga * largo_viga
        st.caption(f"Vol neto: {vol_viga_neto:.2f}m³ | Dosif: {mix_viga['dos']}")
        st.markdown("** Desperdicios**")
        desp_conc_v = st.select_slider("Concreto [%]", [1,2,3,4,5,6,8,10], value=1, key="kc_viga_dc")
        desp_long_v = st.select_slider("Acero Long. [%]", [1,2,3,4,5,6,8,10], value=2, key="kc_viga_dl")
        desp_est_v  = st.select_slider("Acero Estrib. [%]", [1,2,3,4,5,6,8,10], value=3, key="kc_viga_de")

    with vi2:
        st.markdown("** Acero Longitudinal**")
        lt_viga = st.number_input("Longitud barra [m]", 0.0, value=largo_viga + 0.60, step=0.1, key="kc_viga_lt")
        trasl_viga = st.radio("Traslapes:", ["NO", "SI"], horizontal=True, key="kc_viga_trasl")
        long_trasl_v = st.number_input("Long. traslape [m]", 0.0, value=0.35, step=0.05, key="kc_viga_lt2") if trasl_viga == "SI" else 0.0
        if sec_viga == "n  Ilimitado":
            nb_pos = st.number_input("N° barras Acero+", min_value=1, value=2, step=1, key="kc_viga_nbp")
            nb_neg = st.number_input("N° barras Acero-", min_value=1, value=2, step=1, key="kc_viga_nbn")
        else:
            nb_pos = nb_pos_def; nb_neg = nb_neg_def
        st.markdown(f"**Acero+ (positivo — {nb_pos} brs)**")
        var_pos = st.selectbox("Ø Acero+:", [v["nombre"] for v in VARILLAS], index=3, key="kc_viga_vpos")
        vkd_pos = next(v for v in VARILLAS if v["nombre"] == var_pos)
        st.markdown(f"**Acero− (negativo — {nb_neg} brs)**")
        var_neg = st.selectbox("Ø Acero−:", [v["nombre"] for v in VARILLAS], index=3, key="kc_viga_vneg")
        vkd_neg = next(v for v in VARILLAS if v["nombre"] == var_neg)

    with vi3:
        st.markdown("** Zonas**")
        zona_conf_v = st.number_input("Confinada ZC [m]", 0.0, value=1.0, step=0.1, key="kc_viga_zc")
        zona_cent_v = st.number_input("Central [m]", 0.0, value=max(largo_viga - 2*zona_conf_v, 0.0), step=0.1, key="kc_viga_zce")
        st.markdown("**⚙ Estribos E.1**")
        long_e1_v = st.number_input("Long. E.1 [m]", 0.0, value=round(2*(a_viga + b_viga) - 8*max(rec_viga,0.05) + 0.30, 3), step=0.01, key="kc_viga_le1")
        var_e1_v = st.selectbox("Ø E.1:", [v["nombre"] for v in VARILLAS], index=1, key="kc_viga_ve1")
        vkd_e1_v = next(v for v in VARILLAS if v["nombre"] == var_e1_v)
        sep_ZC_v   = st.number_input("Sep. ZC [m]", 0.0, value=0.12, step=0.01, key="kc_viga_szc")
        sep_cent_v = st.number_input("Sep. Centro [m]", 0.0, value=0.12, step=0.01, key="kc_viga_scen")

    # Calculations
    vol_conc_v = vol_viga_neto * cant_viga * (1 + desp_conc_v/100)
    bol_viga   = math.ceil(mix_viga["cem_kg"] * vol_conc_v / R["peso_bolsa"])
    arena_v = mix_viga["arena_m3"] * vol_conc_v
    grava_v = mix_viga["grava_m3"] * vol_conc_v
    agua_v  = mix_viga["agua_lt"]  * vol_conc_v

    long_bar_v = lt_viga + long_trasl_v
    kg_pos = nb_pos * long_bar_v * vkd_pos["kg_m"] * cant_viga * (1 + desp_long_v/100)
    kg_neg = nb_neg * long_bar_v * vkd_neg["kg_m"] * cant_viga * (1 + desp_long_v/100)
    kg_long_v = kg_pos + kg_neg

    n_zc_v  = math.floor(zona_conf_v / max(sep_ZC_v, 0.01)) * 2
    n_cen_v = math.floor(zona_cent_v / max(sep_cent_v, 0.01))
    n_est_v = n_zc_v + n_cen_v
    kg_est_v = n_est_v * long_e1_v * vkd_e1_v["kg_m"] * cant_viga * (1 + desp_est_v/100)
    kg_ace_v = kg_long_v + kg_est_v
    costo_v = bol_viga*float(p.get("cemento",0)) + arena_v*float(p.get("arena",0)) + grava_v*float(p.get("grava",0)) + kg_ace_v*float(p.get("acero_kg",0))

    # Results
    st.markdown("---"); st.markdown("####  Resumen de Resultados")
    rv1,rv2,rv3,rv4 = st.columns(4)
    with rv1:
        st.markdown("**Materiales Concreto**")
        st.metric(" Cemento", f"{bol_viga} bultos"); st.metric(" Arena", f"{arena_v:.2f}m³")
        st.metric(" Grava", f"{grava_v:.2f}m³"); st.metric(" Agua", f"{agua_v:.0f} lt")
    with rv2:
        st.markdown("**Acero Longitudinal**")
        st.metric(f" Acero+ ({nb_pos} brs)", f"{kg_pos:.2f} kg")
        st.metric(f" Acero− ({nb_neg} brs)", f"{kg_neg:.2f} kg")
        st.metric(" Total long.", f"{kg_long_v:.2f} kg")
    with rv3:
        st.markdown("**Estribos E.1**")
        st.metric("N° estribos ZC×2", f"{n_zc_v}")
        st.metric("N° estribos centro", f"{n_cen_v}")
        st.metric("N° total", f"{n_est_v}")
        st.metric("⚙ Peso E.1", f"{kg_est_v:.2f} kg")
    with rv4:
        st.markdown("**Totales**")
        st.metric(" Acero total", f"{kg_ace_v:.2f} kg")
        st.metric(" Vol. concreto", f"{vol_viga_neto*cant_viga:.2f}m³")
        st.metric(" Dosif.", mix_viga["dos"])
        st.metric(" Costo", f"{moneda} {costo_v:,.0f}")

    st.markdown(
        f'<div style="background:#0d2137;border-radius:8px;padding:8px 16px;">'
        f'<span style="color:#ffcc80;">f\'c={fc_viga} kg/cm² | {a_viga}×{b_viga}m | '
        f'L={largo_viga}m | A+:{nb_pos}brs {var_pos} | A-:{nb_neg}brs {var_neg} | E.1:{n_est_v}estr.</span></div>',
        unsafe_allow_html=True)
    with st.expander(" Precios unitarios aplicados en Viga", expanded=False):
        _pr = st.session_state.kc_precios
        _pv_cem = float(_pr.get('cemento', 0))
        _pv_are = float(_pr.get('arena', 0))
        _pv_gra = float(_pr.get('grava', 0))
        _pv_ace = float(_pr.get('acero_kg', 0))
        st.markdown(f"""<div style="display:flex;flex-wrap:wrap;gap:8px;">
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Cemento/bolsa</div>
    <b>{moneda} {_pv_cem:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{bol_viga} bultos = {moneda} {bol_viga*_pv_cem:,.0f}</span></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Arena/m³</div>
    <b>{moneda} {_pv_are:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{arena_v:.2f}m³ = {moneda} {arena_v*_pv_are:,.0f}</span></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Grava/m³</div>
    <b>{moneda} {_pv_gra:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{grava_v:.2f}m³ = {moneda} {grava_v*_pv_gra:,.0f}</span></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Acero/kg</div>
    <b>{moneda} {_pv_ace:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{kg_ace_v:.1f} kg = {moneda} {kg_ace_v*_pv_ace:,.0f}</span></div>
  <div style="background:#1a1a0d;border:1px solid #6d6000;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#ffcc80;font-size:11px;"> COSTO TOTAL</div>
    <b style="color:#ffdd57;font-size:1.1em;">{moneda} {costo_v:,.0f}</b></div>
</div><div style="margin-top:5px;font-size:10px;color:#ff8f00;">⚠ Moneda: <b>{moneda}</b></div>""", unsafe_allow_html=True)
    if st.button(" Agregar Viga al Resumen", key="kc_add_viga", type="primary"):
        lb = desc_viga or "Viga"
        st.session_state.kc_rows.extend([
            {"elemento":f"{lb} — {R['cemento']}","unidad":"bultos","cant":bol_viga,"precio":float(p.get("cemento",0))},
            {"elemento":f"{lb} — Arena","unidad":"m³","cant":round(arena_v,3),"precio":float(p.get("arena",0))},
            {"elemento":f"{lb} — Grava","unidad":"m³","cant":round(grava_v,3),"precio":float(p.get("grava",0))},
            {"elemento":f"{lb} — Acero Long. {var_pos}/{var_neg}","unidad":"kg","cant":round(kg_long_v,2),"precio":float(p.get("acero_kg",0))},
            {"elemento":f"{lb} — Estribos {var_e1_v}","unidad":"kg","cant":round(kg_est_v,2),"precio":float(p.get("acero_kg",0))},
        ])
        save_state()
        st.success(f" {lb}: {bol_viga} bultos cemento | {kg_long_v:.1f} kg long. | {kg_est_v:.1f} kg estrib. | Costo: {moneda} {costo_v:,.0f}")


# ══════════ TAB 5 — LOSA (CM-V3.0) ══════════
with tabs[4]:
    st.subheader(" Calculadora de Losa")
    st.caption(f"CM-V3.0 | Maciza · Nervada · Ilimitado | {norma_sel}")

    TIPOS_LOSA = ["■ Maciza (1 capa)", "■■ Maciza (2 capas)", "▤ Nervada", "⊞ Casetonada", "n Ilimitado"]

    _LOSA_SVG = {
        "■ Maciza (1 capa)": """<svg width='64' height='36' viewBox='0 0 64 36' xmlns='http://www.w3.org/2000/svg'>
          <rect x='4' y='6' width='56' height='22' rx='2' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2'/>
          <line x1='12' y1='22' x2='18' y2='22' stroke='#ff6b35' stroke-width='2'/>
          <line x1='24' y1='22' x2='30' y2='22' stroke='#ff6b35' stroke-width='2'/>
          <line x1='36' y1='22' x2='42' y2='22' stroke='#ff6b35' stroke-width='2'/>
          <line x1='48' y1='22' x2='54' y2='22' stroke='#ff6b35' stroke-width='2'/>
          <line x1='4' y1='22' x2='60' y2='22' stroke='#00d4ff' stroke-width='1' stroke-dasharray='3,3'/>
        </svg>""",
        "■■ Maciza (2 capas)": """<svg width='64' height='36' viewBox='0 0 64 36' xmlns='http://www.w3.org/2000/svg'>
          <rect x='4' y='6' width='56' height='22' rx='2' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2'/>
          <line x1='12' y1='12' x2='18' y2='12' stroke='#ff6b35' stroke-width='2'/>
          <line x1='24' y1='12' x2='30' y2='12' stroke='#ff6b35' stroke-width='2'/>
          <line x1='36' y1='12' x2='42' y2='12' stroke='#ff6b35' stroke-width='2'/>
          <line x1='48' y1='12' x2='54' y2='12' stroke='#ff6b35' stroke-width='2'/>
          <line x1='12' y1='22' x2='18' y2='22' stroke='#ff6b35' stroke-width='2'/>
          <line x1='24' y1='22' x2='30' y2='22' stroke='#ff6b35' stroke-width='2'/>
          <line x1='36' y1='22' x2='42' y2='22' stroke='#ff6b35' stroke-width='2'/>
          <line x1='48' y1='22' x2='54' y2='22' stroke='#ff6b35' stroke-width='2'/>
        </svg>""",
        "▤ Nervada": """<svg width='64' height='36' viewBox='0 0 64 36' xmlns='http://www.w3.org/2000/svg'>
          <rect x='4' y='6' width='56' height='8' rx='1' fill='#1e3a5f' stroke='#4a9eff' stroke-width='1.5'/>
          <rect x='8'  y='14' width='10' height='16' rx='1' fill='#1e3a5f' stroke='#4a9eff' stroke-width='1.5'/>
          <rect x='27' y='14' width='10' height='16' rx='1' fill='#1e3a5f' stroke='#4a9eff' stroke-width='1.5'/>
          <rect x='46' y='14' width='10' height='16' rx='1' fill='#1e3a5f' stroke='#4a9eff' stroke-width='1.5'/>
          <line x1='10' y1='26' x2='16' y2='26' stroke='#ff6b35' stroke-width='2'/>
          <line x1='29' y1='26' x2='35' y2='26' stroke='#ff6b35' stroke-width='2'/>
          <line x1='48' y1='26' x2='54' y2='26' stroke='#ff6b35' stroke-width='2'/>
        </svg>""",
        "n Ilimitado": """<svg width='64' height='36' viewBox='0 0 64 36' xmlns='http://www.w3.org/2000/svg'>
          <rect x='4' y='6' width='56' height='22' rx='2' fill='#1e3a5f' stroke='#4a9eff' stroke-width='2'/>
          <text x='32' y='23' text-anchor='middle' font-size='16' font-weight='bold' fill='#90caf9' font-family='serif'>n</text>
        </svg>""",
        "⊞ Casetonada": """<svg width='64' height='36' viewBox='0 0 64 36' xmlns='http://www.w3.org/2000/svg'>
          <!-- Chapa superior -->
          <rect x='4' y='6' width='56' height='7' rx='1' fill='#1e3a5f' stroke='#4a9eff' stroke-width='1.5'/>
          <!-- Nervios verticales -->
          <rect x='4'  y='13' width='8' height='17' rx='1' fill='#1e3a5f' stroke='#4a9eff' stroke-width='1.5'/>
          <rect x='20' y='13' width='8' height='17' rx='1' fill='#1e3a5f' stroke='#4a9eff' stroke-width='1.5'/>
          <rect x='36' y='13' width='8' height='17' rx='1' fill='#1e3a5f' stroke='#4a9eff' stroke-width='1.5'/>
          <rect x='52' y='13' width='8' height='17' rx='1' fill='#1e3a5f' stroke='#4a9eff' stroke-width='1.5'/>
          <!-- Casetones (vacíos) -->
          <rect x='12' y='13' width='8' height='17' rx='1' fill='#334' stroke='#91a' stroke-width='1' stroke-dasharray='2,1'/>
          <rect x='28' y='13' width='8' height='17' rx='1' fill='#334' stroke='#91a' stroke-width='1' stroke-dasharray='2,1'/>
          <rect x='44' y='13' width='8' height='17' rx='1' fill='#334' stroke='#91a' stroke-width='1' stroke-dasharray='2,1'/>
          <!-- Acero nervios -->
          <line x1='6'  y1='27' x2='10' y2='27' stroke='#ff6b35' stroke-width='2'/>
          <line x1='22' y1='27' x2='26' y2='27' stroke='#ff6b35' stroke-width='2'/>
          <line x1='38' y1='27' x2='42' y2='27' stroke='#ff6b35' stroke-width='2'/>
          <line x1='54' y1='27' x2='58' y2='27' stroke='#ff6b35' stroke-width='2'/>
        </svg>""",
    }

    st.markdown("**Tipo de losa:**")
    _cols_lsec = st.columns(len(TIPOS_LOSA))
    if "kc_losa_tipo" not in st.session_state:
        st.session_state["kc_losa_tipo"] = TIPOS_LOSA[0]
    for _i, _ltype in enumerate(TIPOS_LOSA):
        with _cols_lsec[_i]:
            _selected = st.session_state.get("kc_losa_tipo") == _ltype
            _border = "3px solid #4a9eff" if _selected else "1px solid #333"
            _bg = "#0d1f3c" if _selected else "#111"
            _lbl = _ltype.split(maxsplit=1)[1] if len(_ltype.split()) > 1 else _ltype
            st.markdown(
                f"""<div style='text-align:center;background:{_bg};border:{_border};border-radius:8px;padding:4px 2px;cursor:pointer;'>
                {_LOSA_SVG[_ltype]}<br>
                <span style='font-size:0.65rem;color:#ccc;'>{_lbl}</span></div>""",
                unsafe_allow_html=True
            )
            if st.button("" if _selected else "○", key=f"kc_lsec_btn_{_i}",
                         help=_ltype, use_container_width=True):
                st.session_state["kc_losa_tipo"] = _ltype
                st.rerun()
    tipo_losa = st.session_state.get("kc_losa_tipo", TIPOS_LOSA[0])

    # ======= DIAGRAMA SVG DE LOSA =======
    _is_nervada    = tipo_losa == "▤ Nervada"
    _is_2capas     = tipo_losa == "■■ Maciza (2 capas)"
    _is_casetonada = tipo_losa == "⊞ Casetonada"
    _html_losa_diag = f"""
    <div style="background:#0a192f; padding:16px; border-radius:8px; border:1px solid #1e3a5f;
                margin:12px 0; display:flex; justify-content:space-around; align-items:flex-start;
                flex-wrap:wrap; box-shadow:0 4px 15px rgba(0,0,0,0.3);">
      <!-- Sección transversal -->
      <div style="text-align:center; min-width:220px;">
        <h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Sección Transversal</h4>
        <svg width="220" height="140" viewBox="0 0 220 140" xmlns="http://www.w3.org/2000/svg">
          {'''
          <!-- casetonada -->
          <rect x="10" y="10" width="200" height="20" rx="1" fill="#1e3a5f" stroke="#4a9eff" stroke-width="2"/>
          <rect x="10" y="30" width="30" height="60" fill="#1e3a5f" stroke="#4a9eff" stroke-width="1.5"/>
          <rect x="88" y="30" width="30" height="60" fill="#1e3a5f" stroke="#4a9eff" stroke-width="1.5"/>
          <rect x="166" y="30" width="30" height="60" fill="#1e3a5f" stroke="#4a9eff" stroke-width="1.5"/>
          <rect x="40" y="30" width="48" height="60" fill="#1d1d3e" stroke="#9c27b0" stroke-width="1" stroke-dasharray="4,2"/>
          <rect x="118" y="30" width="48" height="60" fill="#1d1d3e" stroke="#9c27b0" stroke-width="1" stroke-dasharray="4,2"/>
          <line x1="18" y1="84" x2="32" y2="84" stroke="#ff6b35" stroke-width="2.5"/>
          <line x1="96" y1="84" x2="110" y2="84" stroke="#ff6b35" stroke-width="2.5"/>
          <line x1="174" y1="84" x2="188" y2="84" stroke="#ff6b35" stroke-width="2.5"/>
          <text x="64" y="65" font-size="9" fill="#9c27b0" text-anchor="middle">Casetón</text>
          <text x="142" y="65" font-size="9" fill="#9c27b0" text-anchor="middle">Casetón</text>
          ''' if _is_casetonada else
          ('''
          <!-- nervada -->
          <rect x="10" y="10" width="200" height="25" rx="2" fill="#1e3a5f" stroke="#4a9eff" stroke-width="2"/>
          <rect x="18" y="35" width="33" height="55" rx="1" fill="#1e3a5f" stroke="#4a9eff" stroke-width="1.5"/>
          <rect x="93" y="35" width="33" height="55" rx="1" fill="#1e3a5f" stroke="#4a9eff" stroke-width="1.5"/>
          <rect x="168" y="35" width="33" height="55" rx="1" fill="#1e3a5f" stroke="#4a9eff" stroke-width="1.5"/>
          <line x1="23" y1="82" x2="43" y2="82" stroke="#ff6b35" stroke-width="2.5"/>
          <line x1="98" y1="82" x2="118" y2="82" stroke="#ff6b35" stroke-width="2.5"/>
          <line x1="173" y1="82" x2="193" y2="82" stroke="#ff6b35" stroke-width="2.5"/>
          ''' if _is_nervada else
          f'<rect x="10" y="30" width="200" height="50" rx="2" fill="#1e3a5f" stroke="#4a9eff" stroke-width="2"/><line x1="20" y1="72" x2="200" y2="72" stroke="#ff6b35" stroke-width="2"/><line x1="20" y1="38" x2="200" y2="38" stroke="#ff6b35" stroke-width="2"/>' if _is_2capas else
          '<rect x="10" y="30" width="200" height="50" rx="2" fill="#1e3a5f" stroke="#4a9eff" stroke-width="2"/><line x1="20" y1="72" x2="200" y2="72" stroke="#ff6b35" stroke-width="2"/>')}
          <!-- Leyenda -->
          <circle cx="15" cy="130" r="4" fill="#ff6b35"/>
          <text x="22" y="134" font-size="10" fill="#ccc">Acero</text>
          <rect x="80" y="124" width="14" height="9" fill="#1e3a5f" stroke="#4a9eff" stroke-width="1"/>
          <text x="97" y="133" font-size="10" fill="#ccc">Concreto</text>
          {'<rect x="145" y="124" width="14" height="9" fill="#1d1d3e" stroke="#9c27b0" stroke-width="1"/><text x="162" y="133" font-size="10" fill="#ccc">Casetón</text>' if _is_casetonada else ''}
        </svg>
      </div>
      <!-- Vista en planta -->
      <div style="text-align:center; min-width:220px;">
        <h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Vista en Planta</h4>
        <svg width="220" height="140" viewBox="0 0 220 140" xmlns="http://www.w3.org/2000/svg">
          <rect x="10" y="10" width="200" height="120" fill="#1e3a5f" stroke="#4a9eff" stroke-width="2"/>
          {'''
          <rect x="32" y="28" width="34" height="34" fill="#1d1d3e" stroke="#9c27b0" stroke-width="1" stroke-dasharray="3,2"/>
          <rect x="88" y="28" width="34" height="34" fill="#1d1d3e" stroke="#9c27b0" stroke-width="1" stroke-dasharray="3,2"/>
          <rect x="144" y="28" width="34" height="34" fill="#1d1d3e" stroke="#9c27b0" stroke-width="1" stroke-dasharray="3,2"/>
          <rect x="32" y="78" width="34" height="34" fill="#1d1d3e" stroke="#9c27b0" stroke-width="1" stroke-dasharray="3,2"/>
          <rect x="88" y="78" width="34" height="34" fill="#1d1d3e" stroke="#9c27b0" stroke-width="1" stroke-dasharray="3,2"/>
          <rect x="144" y="78" width="34" height="34" fill="#1d1d3e" stroke="#9c27b0" stroke-width="1" stroke-dasharray="3,2"/>
          <line x1="10" y1="28" x2="210" y2="28" stroke="#ff6b35" stroke-width="1.5"/>
          <line x1="10" y1="62" x2="210" y2="62" stroke="#ff6b35" stroke-width="1.5"/>
          <line x1="10" y1="78" x2="210" y2="78" stroke="#ff6b35" stroke-width="1.5"/>
          <line x1="10" y1="112" x2="210" y2="112" stroke="#ff6b35" stroke-width="1.5"/>
          <line x1="32" y1="10" x2="32" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          <line x1="66" y1="10" x2="66" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          <line x1="88" y1="10" x2="88" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          <line x1="122" y1="10" x2="122" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          <line x1="144" y1="10" x2="144" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          <line x1="178" y1="10" x2="178" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          ''' if _is_casetonada else
          '''
          <line x1="10" y1="30" x2="210" y2="30" stroke="#ff6b35" stroke-width="1.5"/>
          <line x1="10" y1="55" x2="210" y2="55" stroke="#ff6b35" stroke-width="1.5"/>
          <line x1="10" y1="80" x2="210" y2="80" stroke="#ff6b35" stroke-width="1.5"/>
          <line x1="10" y1="105" x2="210" y2="105" stroke="#ff6b35" stroke-width="1.5"/>
          <line x1="45" y1="10" x2="45" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          <line x1="90" y1="10" x2="90" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          <line x1="135" y1="10" x2="135" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          <line x1="180" y1="10" x2="180" y2="130" stroke="#00d4ff" stroke-width="1.5"/>
          '''}
          <text x="15" y="8" font-size="9" fill="#90caf9">← Ancho →</text>
          <text x="213" y="70" font-size="9" fill="#90caf9" writing-mode="tb">Largo</text>
          <circle cx="15" cy="137" r="3" fill="#ff6b35"/><text x="21" y="140" font-size="9" fill="#ccc">{'Nervio' if _is_casetonada else 'Dir X'}</text>
          <circle cx="80" cy="137" r="3" fill="{'#9c27b0' if _is_casetonada else '#00d4ff'}"/><text x="86" y="140" font-size="9" fill="#ccc">{'Casetón' if _is_casetonada else 'Dir Y'}</text>
          {'<circle cx="145" cy="137" r="3" fill="#00d4ff"/><text x="151" y="140" font-size="9" fill="#ccc">Nervio Y</text>' if _is_casetonada else ''}
        </svg>
      </div>
    </div>
    """
    import streamlit.components.v1 as _stc_l
    _stc_l.html(_html_losa_diag, height=215, scrolling=False)
    st.markdown("---")
    desc_losa = st.text_input("Descripción", placeholder="Ej: Losa nivel +3.20", key="kc_losa_desc")

    lo1, lo2, lo3 = st.columns(3)
    with lo1:
        st.markdown("** Dimensiones Losa**")
        largo_l  = st.number_input("Largo [m]", 0.0, value=4.0, step=0.5, key="kc_losa_lar")
        ancho_l  = st.number_input("Ancho [m]", 0.0, value=4.0, step=0.5, key="kc_losa_anc")
        esp_l    = st.number_input("Espesor [m]", 0.0, value=0.12, step=0.01, key="kc_losa_esp")
        cant_l   = st.number_input("Cantidad", min_value=1, value=1, key="kc_losa_qty")
        fc_l     = st.selectbox("Resistencia [kg/cm²]:", [m["fc_kgcm2"] for m in MIX_DESIGNS], index=3, key="kc_losa_fc")
        mix_l    = next(m for m in MIX_DESIGNS if m["fc_kgcm2"] == fc_l)
        rec_l    = st.number_input("Recubrimiento [m]", 0.0, value=0.03, step=0.01, key="kc_losa_rec")
        if tipo_losa == "▤ Nervada":
            ancho_nrv = st.number_input("Ancho nervio [m]", 0.0, value=0.10, step=0.01, key="kc_losa_anrv")
            sep_nrv   = st.number_input("Separación nervios [m]", 0.0, value=0.60, step=0.05, key="kc_losa_snrv")
            esp_losa_nrv = st.number_input("Espesor capa superficial [m]", 0.0, value=0.05, step=0.01, key="kc_losa_esnrv")
        if tipo_losa == "⊞ Casetonada":
            st.markdown("** Casetón (Icopor / Madera)**")
            ancho_cas  = st.number_input("Dim. casetón [m]", 0.01, value=0.45, step=0.05, key="kc_losa_acas",
                                         help="Dimensión del casetón cuadrado (ej: 0.35, 0.45, 0.50 m)")
            sep_nrv_cas = st.number_input("Ancho nervio [m]", 0.01, value=0.10, step=0.01, key="kc_losa_snrvcas",
                                          help="Ancho del nervio de concreto entre casetones")
            esp_chapa  = st.number_input("Espesor chapa [m]", 0.01, value=0.05, step=0.01, key="kc_losa_echapa",
                                         help="Espesor de la capa de compresion superior")
            h_cas = max(esp_l - esp_chapa, 0.01)  # altura del casetonado
        else:
            ancho_cas = sep_nrv_cas = esp_chapa = h_cas = 0.0
        st.markdown("** Desperdicios**")
        desp_cl = st.select_slider("Concreto [%]", [1,2,3,4,5,6,8,10], value=5, key="kc_losa_dc")
        desp_al = st.select_slider("Acero [%]", [1,2,3,4,5,6,8,10], value=5, key="kc_losa_da")

    with lo2:
        st.markdown("** Acero Dirección X**")
        if tipo_losa == "n Ilimitado":
            n_barras_x = st.number_input("N° barras direc. X", min_value=1, value=10, step=1, key="kc_losa_nbx")
        else:
            sep_x = st.number_input("Separación X [m]", 0.0, value=0.15, step=0.01, key="kc_losa_sepx")
            n_barras_x = math.ceil(ancho_l / max(sep_x, 0.01)) + 1
        long_x = largo_l + 2*rec_l
        var_x  = st.selectbox("Ø acero X:", [v["nombre"] for v in VARILLAS], index=2, key="kc_losa_varx")
        vkd_x  = next(v for v in VARILLAS if v["nombre"] == var_x)
        if tipo_losa in ["■■ Maciza (2 capas)", "n Ilimitado"]:
            st.markdown("** Acero Dirección X (capa sup.)**")
            if tipo_losa == "n Ilimitado":
                n_barras_x2 = st.number_input("N° barras X superior", min_value=0, value=5, step=1, key="kc_losa_nbx2")
            else:
                sep_x2 = st.number_input("Separación X sup. [m]", 0.0, value=0.20, step=0.01, key="kc_losa_sepx2")
                n_barras_x2 = math.ceil(ancho_l / max(sep_x2, 0.01)) + 1
            var_x2 = st.selectbox("Ø X sup.:", [v["nombre"] for v in VARILLAS], index=2, key="kc_losa_varx2")
            vkd_x2 = next(v for v in VARILLAS if v["nombre"] == var_x2)
        else:
            n_barras_x2 = 0; vkd_x2 = None; var_x2 = None
        st.markdown("** Acero Temperatura**")
        sep_temp = st.number_input("Separación temp. [m]", 0.0, value=0.25, step=0.01, key="kc_losa_temp")
        var_temp = st.selectbox("Ø temperatura:", [v["nombre"] for v in VARILLAS], index=1, key="kc_losa_vart")
        vkd_temp = next(v for v in VARILLAS if v["nombre"] == var_temp)

    with lo3:
        st.markdown("** Acero Dirección Y**")
        if tipo_losa == "n Ilimitado":
            n_barras_y = st.number_input("N° barras direc. Y", min_value=1, value=10, step=1, key="kc_losa_nby")
        else:
            sep_y = st.number_input("Separación Y [m]", 0.0, value=0.15, step=0.01, key="kc_losa_sepy")
            n_barras_y = math.ceil(largo_l / max(sep_y, 0.01)) + 1
        long_y = ancho_l + 2*rec_l
        var_y  = st.selectbox("Ø acero Y:", [v["nombre"] for v in VARILLAS], index=2, key="kc_losa_vary")
        vkd_y  = next(v for v in VARILLAS if v["nombre"] == var_y)
        if tipo_losa in ["■■ Maciza (2 capas)", "n Ilimitado"]:
            st.markdown("** Acero Dirección Y (capa sup.)**")
            if tipo_losa == "n Ilimitado":
                n_barras_y2 = st.number_input("N° barras Y superior", min_value=0, value=5, step=1, key="kc_losa_nby2")
            else:
                sep_y2 = st.number_input("Separación Y sup. [m]", 0.0, value=0.20, step=0.01, key="kc_losa_sepy2")
                n_barras_y2 = math.ceil(largo_l / max(sep_y2, 0.01)) + 1
            var_y2 = st.selectbox("Ø Y sup.:", [v["nombre"] for v in VARILLAS], index=2, key="kc_losa_vary2")
            vkd_y2 = next(v for v in VARILLAS if v["nombre"] == var_y2)
        else:
            n_barras_y2 = 0; vkd_y2 = None; var_y2 = None

    # ── Calculations ──
    area_l = largo_l * ancho_l * cant_l
    if tipo_losa == "▤ Nervada":
        # Nervada 1 dirección: chapa sup + nervios
        n_nervios_x  = math.ceil(largo_l  / max(sep_nrv, 0.01)) + 1
        n_nervios_y  = math.ceil(ancho_l  / max(sep_nrv, 0.01)) + 1
        vol_neto = (largo_l * ancho_l * esp_losa_nrv + n_nervios_x * ancho_nrv * (esp_l - esp_losa_nrv) * ancho_l) * cant_l
    elif tipo_losa == "⊞ Casetonada":
        # Casetonada: volumen total − volumen de casetones
        paso_cas = ancho_cas + sep_nrv_cas           # modulo nervio+casetonado
        n_cas_x  = max(math.floor(largo_l / max(paso_cas, 0.01)), 1)
        n_cas_y  = max(math.floor(ancho_l / max(paso_cas, 0.01)), 1)
        vol_total_bruto = largo_l * ancho_l * esp_l
        vol_casetones   = n_cas_x * n_cas_y * ancho_cas * ancho_cas * h_cas
        vol_neto = max(vol_total_bruto - vol_casetones, vol_total_bruto * 0.3) * cant_l
    else:
        vol_neto = largo_l * ancho_l * esp_l * cant_l

    vol_conc_l = vol_neto * (1 + desp_cl/100)
    bol_l   = math.ceil(mix_l["cem_kg"] * vol_conc_l / R["peso_bolsa"])
    arena_l = mix_l["arena_m3"] * vol_conc_l
    grava_l = mix_l["grava_m3"] * vol_conc_l
    agua_l  = mix_l["agua_lt"]  * vol_conc_l

    kg_x   = n_barras_x * long_x * vkd_x["kg_m"] * cant_l * (1 + desp_al/100)
    kg_x2  = (n_barras_x2 * long_x * vkd_x2["kg_m"] * cant_l * (1 + desp_al/100)) if vkd_x2 else 0
    kg_y   = n_barras_y * long_y * vkd_y["kg_m"] * cant_l * (1 + desp_al/100)
    kg_y2  = (n_barras_y2 * long_y * vkd_y2["kg_m"] * cant_l * (1 + desp_al/100)) if vkd_y2 else 0
    n_temp_bars = math.ceil(largo_l / max(sep_temp, 0.01)) + 1
    kg_temp = n_temp_bars * ancho_l * vkd_temp["kg_m"] * cant_l * (1 + desp_al/100)
    kg_ace_l = kg_x + kg_x2 + kg_y + kg_y2 + kg_temp
    costo_l = bol_l*float(p.get("cemento",0)) + arena_l*float(p.get("arena",0)) + grava_l*float(p.get("grava",0)) + kg_ace_l*float(p.get("acero_kg",0))

    # ── Results ──
    st.markdown("---"); st.markdown("####  Resumen de Resultados")
    rl1,rl2,rl3,rl4 = st.columns(4)
    with rl1:
        st.markdown("**Concreto**")
        st.metric(" Cemento", f"{bol_l} bultos"); st.metric(" Arena", f"{arena_l:.2f}m³")
        st.metric(" Grava", f"{grava_l:.2f}m³"); st.metric(" Agua", f"{agua_l:.0f} lt")
    with rl2:
        st.markdown("**Acero X**")
        st.metric(f" X inf. ({n_barras_x} brs)", f"{kg_x:.2f} kg")
        if kg_x2: st.metric(f" X sup. ({n_barras_x2} brs)", f"{kg_x2:.2f} kg")
        st.metric(" Temperatura", f"{kg_temp:.2f} kg")
    with rl3:
        st.markdown("**Acero Y**")
        st.metric(f" Y inf. ({n_barras_y} brs)", f"{kg_y:.2f} kg")
        if kg_y2: st.metric(f" Y sup. ({n_barras_y2} brs)", f"{kg_y2:.2f} kg")
    with rl4:
        st.markdown("**Totales**")
        st.metric(" Acero total", f"{kg_ace_l:.2f} kg")
        st.metric(" Volumen", f"{vol_neto:.2f}m³")
        st.metric(" Área", f"{area_l:.2f} m²")
        st.metric(" Costo", f"{moneda} {costo_l:,.0f}")
    st.markdown(
        f'<div style="background:#0d2137;border-radius:8px;padding:8px 16px;">'
        f'<span style="color:#ffcc80;">{tipo_losa} | {largo_l}×{ancho_l}m | e={esp_l}m | '
        f'f\'c={fc_l}kg/cm² | {kg_ace_l:.1f}kg acero | Vol={vol_neto:.2f}m³</span></div>',
        unsafe_allow_html=True)
    with st.expander(" Precios unitarios aplicados en Losa", expanded=False):
        _pr = st.session_state.kc_precios
        _pl_cem = float(_pr.get('cemento', 0))
        _pl_are = float(_pr.get('arena', 0))
        _pl_gra = float(_pr.get('grava', 0))
        _pl_ace = float(_pr.get('acero_kg', 0))
        st.markdown(f"""<div style="display:flex;flex-wrap:wrap;gap:8px;">
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Cemento/bolsa</div>
    <b>{moneda} {_pl_cem:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{bol_l} bultos = {moneda} {bol_l*_pl_cem:,.0f}</span></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Arena/m³</div>
    <b>{moneda} {_pl_are:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{arena_l:.2f}m³ = {moneda} {arena_l*_pl_are:,.0f}</span></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Grava/m³</div>
    <b>{moneda} {_pl_gra:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{grava_l:.2f}m³ = {moneda} {grava_l*_pl_gra:,.0f}</span></div>
  <div style="background:#0d2137;border:1px solid #1e4d8c;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#90caf9;font-size:11px;"> Acero/kg</div>
    <b>{moneda} {_pl_ace:,.0f}</b><br><span style="font-size:10px;color:#7ec87e;">{kg_ace_l:.1f} kg = {moneda} {kg_ace_l*_pl_ace:,.0f}</span></div>
  <div style="background:#1a1a0d;border:1px solid #6d6000;border-radius:8px;padding:6px 12px;min-width:130px;">
    <div style="color:#ffcc80;font-size:11px;"> COSTO TOTAL</div>
    <b style="color:#ffdd57;font-size:1.1em;">{moneda} {costo_l:,.0f}</b></div>
</div><div style="margin-top:5px;font-size:10px;color:#ff8f00;">⚠ Moneda: <b>{moneda}</b></div>""", unsafe_allow_html=True)
    if st.button(" Agregar Losa al Resumen", key="kc_add_losa", type="primary"):
        lb = desc_losa or f"Losa {tipo_losa[:5]}"
        st.session_state.kc_rows.extend([
            {"elemento":f"{lb} — {R['cemento']}","unidad":"bultos","cant":bol_l,"precio":float(p.get("cemento",0))},
            {"elemento":f"{lb} — Arena","unidad":"m³","cant":round(arena_l,3),"precio":float(p.get("arena",0))},
            {"elemento":f"{lb} — Grava","unidad":"m³","cant":round(grava_l,3),"precio":float(p.get("grava",0))},
            {"elemento":f"{lb} — Acero total","unidad":"kg","cant":round(kg_ace_l,2),"precio":float(p.get("acero_kg",0))},
        ])
        save_state()
        st.success(f" {lb}: {bol_l} bultos cemento | {kg_ace_l:.1f} kg acero | Vol {vol_neto:.2f}m³ | Costo: {moneda} {costo_l:,.0f}")


# ══════════ TAB 6 — CIMIENTO (CM-V3.0 · 11 modos) ══════════
with tabs[5]:
    st.subheader(" Calculadora de Cimiento")
    st.caption(f"CM-V3.0 | 11 modos | {norma_sel}")

    # Mortar mix empirical constants (Colombia)
    _MORT = {
        "1:2": {"bol":11.0,"sand":0.60,"water":180},
        "1:3": {"bol":9.0, "sand":0.90,"water":200},
        "1:4": {"bol":7.3, "sand":1.16,"water":230},
        "1:5": {"bol":6.0, "sand":1.21,"water":240},
        "1:6": {"bol":5.0, "sand":1.25,"water":250},
    }
    _ELECTROMALLA_AREA = {"6×6 W1.4": 14.4, "6×6 W2.0": 14.4, "4×4 W1.4": 14.4, "6×6 W2.9": 14.4}

    MODOS_CIM = [
        " Ciclópeo Cúbico"," Ciclópeo Central"," Ciclópeo Lateral",
        "⬜ Zapata Central","⬜ Zapata Lateral","⬜ Zapata Esquinera",
        "% Losa Cim · Por %","D Losa Cim · Dimensión",
        "→ Zapata Corrida","⚡ Electromalla"," Suelo Cemento",
    ]
    modo_cim = st.radio("Modo:", MODOS_CIM, horizontal=True, key="kc_cim_modo")
    st.markdown("---")
    desc_cim = st.text_input("Descripción", placeholder="Ej: Cimiento C-1", key="kc_cim_desc")

    # ======= DIAGRAMA TÉCNICO PROFESIONAL DE CIMIENTO =======
    _is_ciclopeo = "Ciclópeo" in modo_cim
    _is_zapata   = "Zapata" in modo_cim and "Corrida" not in modo_cim
    _is_corrida  = "Corrida" in modo_cim
    _is_losa     = "Losa" in modo_cim
    _is_malla    = "Electromalla" in modo_cim
    _is_suelo    = "Suelo Cemento" in modo_cim

    _w_svg, _h_svg = 260, 160
    _svg_corte, _svg_planta = "", ""

    _defs = """<defs>
      <pattern id="soil_c" width="8" height="8" patternUnits="userSpaceOnUse" patternTransform="rotate(45)">
        <line x1="0" y1="0" x2="0" y2="8" stroke="#5d4037" stroke-width="0.8" opacity="0.4"/>
        <line x1="2" y1="0" x2="2" y2="8" stroke="#5d4037" stroke-width="0.3" opacity="0.2"/>
      </pattern>
      <pattern id="stone_c" width="16" height="16" patternUnits="userSpaceOnUse">
        <rect width="16" height="16" fill="#1e3a5f" opacity="0.9"/>
        <path d="M2,2 Q4,0 6,2 T5,6 Q2,8 1,5 Z m8,8 Q12,8 14,10 T13,14 Q10,16 9,13 Z m-6,4 Q4,10 6,12" fill="#7f8c8d" opacity="0.5"/>
      </pattern>
      <pattern id="conc_c" width="10" height="10" patternUnits="userSpaceOnUse">
        <rect width="10" height="10" fill="#1e3a5f" opacity="0.8"/>
        <circle cx="2" cy="2" r="0.8" fill="#4a9eff" opacity="0.5"/>
        <circle cx="7" cy="6" r="1.2" fill="#4a9eff" opacity="0.5"/>
      </pattern>
      <pattern id="suelo_cem_c" width="12" height="12" patternUnits="userSpaceOnUse">
        <rect width="12" height="12" fill="#3e2723" opacity="0.9"/>
        <circle cx="3" cy="3" r="1" fill="#8d6e63" opacity="0.7"/>
        <line x1="0" y1="12" x2="12" y2="0" stroke="#8d6e63" stroke-width="1" opacity="0.4"/>
      </pattern>
    </defs>"""

    _soil_bg = '<rect x="0" y="50" width="260" height="110" fill="url(#soil_c)"/><line x1="0" y1="50" x2="260" y2="50" stroke="#795548" stroke-width="2"/><path d="M 15 50 L 10 40 L 20 40 Z" fill="#795548"/>'

    if _is_ciclopeo:
        if "Cúbico" in modo_cim:
            _svg_corte += f'{_soil_bg}<rect x="70" y="50" width="120" height="80" fill="url(#stone_c)" stroke="#bdc3c7" stroke-width="1.5"/>'
            _svg_planta += '<rect x="50" y="30" width="160" height="100" fill="url(#stone_c)" stroke="#bdc3c7" stroke-width="1.5"/><text x="130" y="85" fill="#ecf0f1" font-size="12" font-weight="bold" text-anchor="middle">Largo L × Base B</text>'
        elif "Central" in modo_cim:
            _svg_corte += f'{_soil_bg}<polygon points="90,50 170,50 200,130 60,130" fill="url(#stone_c)" stroke="#bdc3c7" stroke-width="1.5"/>'
            _svg_planta += '<rect x="50" y="20" width="160" height="120" fill="url(#stone_c)" stroke="#bdc3c7" stroke-width="1.5"/><rect x="50" y="40" width="160" height="80" fill="none" stroke="#bdc3c7" stroke-dasharray="4,3"/>'
        elif "Lateral" in modo_cim:
            _svg_corte += f'{_soil_bg}<polygon points="70,50 150,50 190,130 70,130" fill="url(#stone_c)" stroke="#bdc3c7" stroke-width="1.5"/>'
            _svg_planta += '<rect x="50" y="20" width="160" height="120" fill="url(#stone_c)" stroke="#bdc3c7" stroke-width="1.5"/><rect x="50" y="40" width="160" height="100" fill="none" stroke="#bdc3c7" stroke-dasharray="4,3"/>'
        _svg_corte += '<line x1="30" y1="50" x2="30" y2="130" stroke="#ccc" stroke-width="1" stroke-dasharray="2,2"/><text x="25" y="95" fill="#ccc" font-size="10" transform="rotate(-90 25,95)" text-anchor="middle">Alt H</text>'

    elif _is_zapata:
        _col_w = 30
        _z_x1, _z_x2, _y_base = 60, 200, 130
        _cx = 130 if "Central" in modo_cim else 70 + _col_w/2
        if "Lateral" in modo_cim or "Esquinera" in modo_cim: _z_x1 = 70
        _z_w = _z_x2 - _z_x1

        _svg_corte += f'{_soil_bg}'
        _svg_corte += f'<rect x="{_z_x1}" y="{_y_base-25}" width="{_z_w}" height="25" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="1.5"/>'
        _svg_corte += f'<rect x="{_cx - _col_w/2}" y="10" width="{_col_w}" height="{_y_base-10}" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="1.5" stroke-dasharray="0,0,30,0"/>'
        _svg_corte += f'<line x1="{_z_x1+6}" y1="{_y_base-6}" x2="{_z_x2-6}" y2="{_y_base-6}" stroke="#ff6b35" stroke-width="2.5"/>'
        _svg_corte += f'<path d="M{_z_x1+6},{_y_base-6} L{_z_x1+6},{_y_base-18} M{_z_x2-6},{_y_base-6} L{_z_x2-6},{_y_base-18}" stroke="#ff6b35" stroke-width="2.5" fill="none"/>'
        for i in range(7): _svg_corte += f'<circle cx="{_z_x1 + 16 + i*(_z_w-32)/6}" cy="{_y_base-10}" r="2" fill="#00d4ff"/>'
            
        _svg_planta += '<rect x="40" y="20" width="180" height="120" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="2"/>'
        for i in range(8): _svg_planta += f'<line x1="46" y1="{28 + i*14}" x2="214" y2="{28 + i*14}" stroke="#ff6b35" stroke-width="1.5"/>'
        for i in range(12): _svg_planta += f'<line x1="{48 + i*14}" y1="26" x2="{48 + i*14}" y2="134" stroke="#00d4ff" stroke-width="1.5"/>'
        _p_cx, _p_cy = 130, 80
        if "Lateral" in modo_cim: _p_cx = 55
        if "Esquinera" in modo_cim: _p_cx, _p_cy = 55, 35
        _svg_planta += f'<rect x="{_p_cx-15}" y="{_p_cy-15}" width="30" height="30" fill="#2c3e50" stroke="#fff" stroke-width="2"/>'
        
    elif _is_corrida:
        _svg_corte += f'{_soil_bg}<rect x="80" y="100" width="100" height="30" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="1.5"/><rect x="110" y="20" width="40" height="80" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="1.5" stroke-dasharray="0,0,40,0"/>'
        _svg_corte += '<line x1="86" y1="124" x2="174" y2="124" stroke="#ff6b35" stroke-width="2.5"/>'
        for i in range(5): _svg_corte += f'<circle cx="{92 + i*19}" cy="120" r="2" fill="#00d4ff"/>'
        _svg_planta += '<rect x="20" y="40" width="220" height="80" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="2"/><rect x="20" y="65" width="220" height="30" fill="#2c3e50" stroke="#ccc" stroke-dasharray="4,2"/>'
        for i in range(4): _svg_planta += f'<line x1="20" y1="{45 + i*23}" x2="240" y2="{45 + i*23}" stroke="#00d4ff" stroke-width="1.5" stroke-dasharray="10,2"/>'

    elif _is_losa:
        _svg_corte += f'{_soil_bg}<rect x="30" y="90" width="200" height="40" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="1.5"/>'
        _svg_corte += '<line x1="36" y1="98" x2="224" y2="98" stroke="#ff6b35" stroke-width="2.5"/><line x1="36" y1="122" x2="224" y2="122" stroke="#ff6b35" stroke-width="2.5"/>'
        for i in range(10): _svg_corte += f'<circle cx="{42 + i*19}" cy="94" r="2" fill="#00d4ff"/><circle cx="{42 + i*19}" cy="118" r="2" fill="#00d4ff"/>'
        _svg_planta += '<rect x="30" y="20" width="200" height="120" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="2"/>'
        for i in range(8): _svg_planta += f'<line x1="30" y1="{28 + i*14}" x2="230" y2="{28 + i*14}" stroke="#ff6b35" stroke-width="1.5"/>'
        for i in range(13): _svg_planta += f'<line x1="{38 + i*15}" y1="20" x2="{38 + i*15}" y2="140" stroke="#00d4ff" stroke-width="1.5"/>'

    elif _is_malla:
        _svg_corte += f'{_soil_bg}<rect x="30" y="100" width="200" height="20" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="1.5"/>'
        _svg_corte += '<line x1="30" y1="110" x2="230" y2="110" stroke="#00d4ff" stroke-width="1.5" stroke-dasharray="2,2"/>'
        _svg_planta += '<rect x="30" y="20" width="200" height="120" fill="url(#conc_c)" stroke="#4a9eff" stroke-width="2"/>'
        for i in range(10): 
            _svg_planta += f'<line x1="30" y1="{25 + i*12}" x2="230" y2="{25 + i*12}" stroke="#00d4ff" stroke-width="1" stroke-dasharray="4,2"/>'
        for i in range(14):
            _svg_planta += f'<line x1="{38 + i*14}" y1="20" x2="{38 + i*14}" y2="140" stroke="#00d4ff" stroke-width="1" stroke-dasharray="4,2"/>'
            
    elif _is_suelo:
        _svg_corte += f'{_soil_bg}<rect x="50" y="50" width="160" height="90" fill="url(#suelo_cem_c)" stroke="#8d6e63" stroke-width="1.5"/>'
        _svg_planta += '<rect x="40" y="20" width="180" height="120" fill="url(#suelo_cem_c)" stroke="#8d6e63" stroke-width="2"/>'
        _svg_planta += '<text x="130" y="80" fill="#fff" font-size="14" font-weight="bold" text-anchor="middle">SUELO CEMENTO</text>'

    _html_cim_diag = f'''
    <div style="background:#0a192f; padding:16px; border-radius:8px; border:1px solid #1e3a5f; margin:12px 0 24px 0; display:flex; justify-content:space-around; align-items:center; flex-wrap:wrap; box-shadow:0 4px 15px rgba(0,0,0,0.3);">
      <div style="text-align:center; min-width:270px;">
        <h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Corte (Elevación)</h4>
        <svg width="{_w_svg}" height="{_h_svg}" viewBox="0 0 {_w_svg} {_h_svg}" xmlns="http://www.w3.org/2000/svg">
          {_defs}{_svg_corte}
        </svg>
      </div>
      <div style="text-align:center; min-width:270px; border-left:1px dashed #1e3a5f;">
        <h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Vista en Planta</h4>
        <svg width="{_w_svg}" height="{_h_svg}" viewBox="0 0 {_w_svg} {_h_svg}" xmlns="http://www.w3.org/2000/svg">
          {_defs}{_svg_planta}
        </svg>
      </div>
    </div>
    '''
    import streamlit.components.v1 as _stc_c
    _stc_c.html(_html_cim_diag, height=230, scrolling=False)

    ci1, ci2, ci3 = st.columns(3)


    # ─────── CICLÓPEO ───────
    if modo_cim in [" Ciclópeo Cúbico"," Ciclópeo Central"," Ciclópeo Lateral"]:
        with ci1:
            st.markdown("** Dimensiones**")
            if modo_cim == " Ciclópeo Cúbico":
                base_c  = st.number_input("Base B [m]",0.0,value=1.0,step=0.1,key="kc_cim_B_1")
                alt_c   = st.number_input("Altura H [m]",0.0,value=1.0,step=0.1,key="kc_cim_H_1")
                lar_c   = st.number_input("Largo [m]",0.0,value=1.0,step=0.1,key="kc_cim_L_1")
                vol_cim_neto = base_c * alt_c * lar_c
            else:
                cor_c   = st.number_input("Corona C [m]",0.0,value=0.60,step=0.1,key="kc_cim_C")
                base_c  = st.number_input("Base B [m]",0.0,value=1.0,step=0.1,key="kc_cim_B_2")
                alt_c   = st.number_input("Altura H [m]",0.0,value=1.0,step=0.1,key="kc_cim_H_2")
                lar_c   = st.number_input("Largo [m]",0.0,value=1.0,step=0.1,key="kc_cim_L_2")
                vol_cim_neto = (cor_c + base_c) / 2 * alt_c * lar_c
            cant_cim = st.number_input("Cantidad",min_value=1,value=1,key="kc_cim_qty_1")
            st.caption(f"Vol neto: {vol_cim_neto:.2f}m³")
        with ci2:
            st.markdown("** Piedra**")
            pct_piedra = st.selectbox("% Piedra:",[60,65,70,75,80,85,90],index=2,key="kc_cim_ppiedra")
            desp_p = st.select_slider("Desp. Piedra [%]",[1,2,3,4,5,6,7,8,10],value=7,key="kc_cim_dp")
            st.markdown("** Mortero de Pega**")
            dos_mort = st.selectbox("Dosificación:",["1:2","1:3","1:4","1:5","1:6"],index=3,key="kc_cim_dos")
            pct_mort = st.selectbox("% Mortero:",[20,25,30,35,40],index=2,key="kc_cim_pmort")
            desp_m = st.select_slider("Desp. Mortero [%]",[1,2,3,4,5,6,8,10],value=10,key="kc_cim_dm")
        with ci3:
            st.markdown("**Cantidad**")
            # Display only

        # Calc
        vol_t = vol_cim_neto * cant_cim
        vol_piedra = vol_t * (pct_piedra/100) * (1+desp_p/100)
        vol_mort_n = vol_t * (pct_mort/100)
        vol_mort   = vol_mort_n * (1+desp_m/100)
        mx = _MORT[dos_mort]
        bol_cim  = math.ceil(vol_mort * mx["bol"])
        arena_cim = vol_mort * mx["sand"]
        agua_cim  = vol_mort * mx["water"]
        costo_c = vol_piedra*float(p.get("piedra",65)) + bol_cim*float(p.get("cemento",0)) + arena_cim*float(p.get("arena",0))
        st.markdown("---"); st.markdown("####  Resultados")
        rci1,rci2 = st.columns(2)
        with rci1:
            st.metric(" Piedra",f"{vol_piedra:.2f}m³"); st.metric(" Cemento",f"{bol_cim} bultos")
            st.metric(" Arena",f"{arena_cim:.2f}m³"); st.metric(" Agua",f"{agua_cim:.0f} lt")
        with rci2:
            st.metric("Dosif.",dos_mort); st.metric("% Piedra",f"{pct_piedra}%"); st.metric("Vol Piedra",f"{vol_piedra:.2f}m³"); st.metric("Vol Mortero",f"{vol_mort:.2f}m³")

    # ─────── ZAPATAS AISLADAS ───────
    elif modo_cim in ["⬜ Zapata Central","⬜ Zapata Lateral","⬜ Zapata Esquinera"]:
        with ci1:
            st.markdown("** Losa / Concreto**")
            esp_z  = st.number_input("Espesor [m]",0.0,value=0.30,step=0.05,key="kc_cim_ez")
            lar_z  = st.number_input("Largo [m]",0.0,value=1.0,step=0.1,key="kc_cim_lz")
            anc_z  = st.number_input("Ancho [m]",0.0,value=1.0,step=0.1,key="kc_cim_az")
            fc_z   = st.selectbox("Resistencia [kg/cm²]:",[m["fc_kgcm2"] for m in MIX_DESIGNS],index=3,key="kc_cim_fcz")
            cant_cim = st.number_input("Cantidad",min_value=1,value=1,key="kc_cim_qty_2")
            mix_z  = next(m for m in MIX_DESIGNS if m["fc_kgcm2"]==fc_z)
            rec_z  = st.number_input("Recub. lat. [m]",0.0,value=0.05,step=0.01,key="kc_cim_recz")
            st.markdown("**Desperdicios**")
            desp_cz = st.select_slider("Concreto [%]",[1,2,3,4,5,6,8,10],value=4,key="kc_cim_dcz")
            desp_az = st.select_slider("Acero [%]",[1,2,3,4,5,6,8,10],value=5,key="kc_cim_daz")
        with ci2:
            st.markdown("** Acero en X**")
            long_x_z  = st.number_input("Longitud X [m]",0.0,value=lar_z-2*rec_z,step=0.05,key="kc_cim_lxz")
            sep_x_z   = st.number_input("Separación X [m]",0.0,value=0.30,step=0.05,key="kc_cim_sxz")
            var_x_z   = st.selectbox("Ø X:",[v["nombre"] for v in VARILLAS],index=3,key="kc_cim_vxz")
            vkd_x_z   = next(v for v in VARILLAS if v["nombre"]==var_x_z)
            trasl_xz  = st.radio("Traslapes X:",["NO","SI"],horizontal=True,key="kc_cim_txz")
            lt_xz = st.number_input("Long. traslape X [m]",0.0,value=0.35,step=0.05,key="kc_cim_ltxz") if trasl_xz=="SI" else 0.0
        with ci3:
            st.markdown("** Acero en Y**")
            long_y_z  = st.number_input("Longitud Y [m]",0.0,value=anc_z-2*rec_z,step=0.05,key="kc_cim_lyz")
            sep_y_z   = st.number_input("Separación Y [m]",0.0,value=0.30,step=0.05,key="kc_cim_syz")
            var_y_z   = st.selectbox("Ø Y:",[v["nombre"] for v in VARILLAS],index=3,key="kc_cim_vyz")
            vkd_y_z   = next(v for v in VARILLAS if v["nombre"]==var_y_z)
            trasl_yz  = st.radio("Traslapes Y:",["NO","SI"],horizontal=True,key="kc_cim_tyz")
            lt_yz = st.number_input("Long. traslape Y [m]",0.0,value=0.35,step=0.05,key="kc_cim_ltyz") if trasl_yz=="SI" else 0.0

        vol_z = lar_z*anc_z*esp_z*cant_cim*(1+desp_cz/100)
        bol_cim  = math.ceil(mix_z["cem_kg"]*vol_z/R["peso_bolsa"])
        arena_cim = mix_z["arena_m3"]*vol_z; grava_cim=mix_z["grava_m3"]*vol_z; agua_cim=mix_z["agua_lt"]*vol_z
        n_bars_x_z = math.floor(anc_z/max(sep_x_z,0.01))+1
        n_bars_y_z = math.floor(lar_z/max(sep_y_z,0.01))+1
        kg_x_z = n_bars_x_z*(long_x_z+lt_xz)*vkd_x_z["kg_m"]*cant_cim*(1+desp_az/100)
        kg_y_z = n_bars_y_z*(long_y_z+lt_yz)*vkd_y_z["kg_m"]*cant_cim*(1+desp_az/100)
        kg_ace_z = kg_x_z + kg_y_z
        costo_c = bol_cim*float(p.get("cemento",0))+arena_cim*float(p.get("arena",0))+grava_cim*float(p.get("grava",0))+kg_ace_z*float(p.get("acero_kg",0))
        st.markdown("---"); st.markdown("####  Resultados")
        rz1,rz2,rz3,rz4 = st.columns(4)
        with rz1:
            st.metric(" Cemento",f"{bol_cim} bultos"); st.metric(" Arena",f"{arena_cim:.2f}m³")
            st.metric(" Grava",f"{grava_cim:.2f}m³"); st.metric(" Agua",f"{agua_cim:.0f} lt")
        with rz2: st.metric(f" Acero X ({n_bars_x_z}brs)",f"{kg_x_z:.2f} kg")
        with rz3: st.metric(f" Acero Y ({n_bars_y_z}brs)",f"{kg_y_z:.2f} kg")
        with rz4: st.metric(" Acero total",f"{kg_ace_z:.2f} kg"); st.metric(" Costo",f"{moneda} {costo_c:,.0f}")

    # ─────── LOSA CIMENTACIÓN POR % ───────
    elif modo_cim == "% Losa Cim · Por %":
        with ci1:
            st.markdown("** Dimensiones**")
            anc_lc = st.number_input("Ancho [m]",0.0,value=1.0,step=0.1,key="kc_cim_alcpct")
            lar_lc = st.number_input("Largo [m]",0.0,value=1.0,step=0.1,key="kc_cim_llcpct")
            esp_lc = st.number_input("Espesor [m]",0.0,value=1.0,step=0.1,key="kc_cim_elcpct")
            cant_cim = st.number_input("Cantidad",min_value=1,value=1,key="kc_cim_qty_3")
            fc_lc  = st.selectbox("Resistencia [kg/cm²]:",[m["fc_kgcm2"] for m in MIX_DESIGNS],index=3,key="kc_cim_fclcpct")
            mix_lc = next(m for m in MIX_DESIGNS if m["fc_kgcm2"]==fc_lc)
            st.markdown("**Desperdicios**"); desp_clc=st.select_slider("Concreto [%]",[1,2,3,4,5,6,8,10],value=6,key="kc_cim_dclcp"); desp_alc=st.select_slider("Acero [%]",[1,2,3,4,5,6,8,10],value=6,key="kc_cim_dalcp")
        with ci2:
            st.markdown("** Acero**")
            prop_acero = st.selectbox("Proporción % acero:",[0.1,0.14,0.18,0.24,0.30,0.34,0.40,0.50,0.60,0.80,1.0],index=5,key="kc_cim_prop")
            var_lc = st.selectbox("Diámetro:",[v["nombre"] for v in VARILLAS],index=3,key="kc_cim_vlcpct")
        vol_lc = anc_lc*lar_lc*esp_lc*cant_cim*(1+desp_clc/100)
        bol_cim=math.ceil(mix_lc["cem_kg"]*vol_lc/R["peso_bolsa"]); arena_cim=mix_lc["arena_m3"]*vol_lc; grava_cim=mix_lc["grava_m3"]*vol_lc; agua_cim=mix_lc["agua_lt"]*vol_lc
        kg_ace_lc = anc_lc*lar_lc*cant_cim*(prop_acero/100)*7850*(1+desp_alc/100)
        costo_c = bol_cim*float(p.get("cemento",0))+arena_cim*float(p.get("arena",0))+grava_cim*float(p.get("grava",0))+kg_ace_lc*float(p.get("acero_kg",0))
        st.markdown("---"); st.markdown("####  Resultados")
        rp1,rp2 = st.columns(2)
        with rp1: st.metric(" Cemento",f"{bol_cim} bultos"); st.metric(" Arena",f"{arena_cim:.2f}m³"); st.metric(" Grava",f"{grava_cim:.2f}m³"); st.metric(" Agua",f"{agua_cim:.0f}lt")
        with rp2: st.metric(" Acero",f"{kg_ace_lc:.2f} kg"); st.metric(" Dosif",mix_lc["dos"]); st.metric(" Costo",f"{moneda} {costo_c:,.0f}")

    # ─────── LOSA CIMENTACIÓN DIMENSIÓN ───────
    elif modo_cim == "D Losa Cim · Dimensión":
        with ci1:
            st.markdown("** Losa / Concreto**")
            esp_ld  = st.number_input("Espesor [m]",0.0,value=0.20,step=0.05,key="kc_cim_eldm")
            lar_ld  = st.number_input("Largo [m]",0.0,value=6.0,step=0.5,key="kc_cim_lldm")
            anc_ld  = st.number_input("Ancho [m]",0.0,value=4.0,step=0.5,key="kc_cim_aldm")
            fc_ld   = st.selectbox("Resistencia [kg/cm²]:",[m["fc_kgcm2"] for m in MIX_DESIGNS],index=2,key="kc_cim_fcldm")
            cant_cim = st.number_input("Cantidad",min_value=1,value=1,key="kc_cim_qty_4")
            mix_ld  = next(m for m in MIX_DESIGNS if m["fc_kgcm2"]==fc_ld)
            rec_ld  = st.number_input("Recub. lat. [m]",0.0,value=0.05,step=0.01,key="kc_cim_recldm")
            st.markdown("**Desperdicios**"); desp_cld=st.select_slider("Concreto [%]",[1,2,3,4,5,6,8,10],value=3,key="kc_cim_dcldm"); desp_ald=st.select_slider("Acero [%]",[1,2,3,4,5,6,8,10],value=5,key="kc_cim_daldm")
        with ci2:
            st.markdown("** Acero X — capa (a)**")
            lxa=st.number_input("Long. Xa [m]",0.0,value=6.50,step=0.1,key="kc_cim_lxa"); sepa=st.number_input("Sep. Xa [m]",0.0,value=0.30,step=0.05,key="kc_cim_sepa")
            vaxa=st.selectbox("Ø Xa:",[v["nombre"] for v in VARILLAS],index=3,key="kc_cim_vaxa"); vkdxa=next(v for v in VARILLAS if v["nombre"]==vaxa)
            st.markdown("** Acero X — capa (b)**")
            lxb=st.number_input("Long. Xb [m]",0.0,value=6.50,step=0.1,key="kc_cim_lxb"); sepb=st.number_input("Sep. Xb [m]",0.0,value=0.35,step=0.05,key="kc_cim_sepb")
            vaxb=st.selectbox("Ø Xb:",[v["nombre"] for v in VARILLAS],index=3,key="kc_cim_vaxb"); vkdxb=next(v for v in VARILLAS if v["nombre"]==vaxb)
            st.markdown("**Traslapes X**"); trx=st.radio("Trasl. X:",["NO","SI"],horizontal=True,key="kc_cim_trxd"); ltx=st.number_input("Long.trasl.X[m]",0.0,value=0.35,step=0.05,key="kc_cim_ltxd") if trx=="SI" else 0.0
        with ci3:
            st.markdown("** Acero Y — capa (a)**")
            lya=st.number_input("Long. Ya [m]",0.0,value=2.20,step=0.1,key="kc_cim_lya"); sepa_y=st.number_input("Sep. Ya [m]",0.0,value=0.40,step=0.05,key="kc_cim_sepa_y")
            vaya=st.selectbox("Ø Ya:",[v["nombre"] for v in VARILLAS],index=3,key="kc_cim_vaya"); vkdya=next(v for v in VARILLAS if v["nombre"]==vaya)
            st.markdown("** Acero Y — capa (b)**")
            lyb=st.number_input("Long. Yb [m]",0.0,value=2.30,step=0.1,key="kc_cim_lyb"); sepb_y=st.number_input("Sep. Yb [m]",0.0,value=0.45,step=0.05,key="kc_cim_sepb_y")
            vayb=st.selectbox("Ø Yb:",[v["nombre"] for v in VARILLAS],index=3,key="kc_cim_vayb"); vkdyb=next(v for v in VARILLAS if v["nombre"]==vayb)
            st.markdown("**Traslapes Y**"); try_=st.radio("Trasl. Y:",["NO","SI"],horizontal=True,key="kc_cim_tryd"); lty=st.number_input("Long.trasl.Y[m]",0.0,value=0.35,step=0.05,key="kc_cim_ltyd") if try_=="SI" else 0.0

        vol_ld = esp_ld*lar_ld*anc_ld*cant_cim*(1+desp_cld/100)
        bol_cim=math.ceil(mix_ld["cem_kg"]*vol_ld/R["peso_bolsa"]); arena_cim=mix_ld["arena_m3"]*vol_ld; grava_cim=mix_ld["grava_m3"]*vol_ld; agua_cim=mix_ld["agua_lt"]*vol_ld
        nb_xa=math.floor(anc_ld/max(sepa,0.01))+1; nb_xb=math.floor(anc_ld/max(sepb,0.01))+1
        nb_ya=math.floor(lar_ld/max(sepa_y,0.01))+1; nb_yb=math.floor(lar_ld/max(sepb_y,0.01))+1
        kg_xa=nb_xa*(lxa+ltx)*vkdxa["kg_m"]*cant_cim*(1+desp_ald/100); kg_xb=nb_xb*(lxb+ltx)*vkdxb["kg_m"]*cant_cim*(1+desp_ald/100)
        kg_ya=nb_ya*(lya+lty)*vkdya["kg_m"]*cant_cim*(1+desp_ald/100); kg_yb=nb_yb*(lyb+lty)*vkdyb["kg_m"]*cant_cim*(1+desp_ald/100)
        kg_ace_ld=kg_xa+kg_xb+kg_ya+kg_yb
        costo_c=bol_cim*float(p.get("cemento",0))+arena_cim*float(p.get("arena",0))+grava_cim*float(p.get("grava",0))+kg_ace_ld*float(p.get("acero_kg",0))
        st.markdown("---"); st.markdown("####  Resultados")
        rd1,rd2,rd3,rd4 = st.columns(4)
        with rd1: st.metric(" Cemento",f"{bol_cim} bultos"); st.metric(" Arena",f"{arena_cim:.2f}m³"); st.metric(" Grava",f"{grava_cim:.2f}m³"); st.metric(" Agua",f"{agua_cim:.0f}lt")
        with rd2: st.metric(f" Xa ({nb_xa}brs)",f"{kg_xa:.2f}kg"); st.metric(f" Xb ({nb_xb}brs)",f"{kg_xb:.2f}kg")
        with rd3: st.metric(f" Ya ({nb_ya}brs)",f"{kg_ya:.2f}kg"); st.metric(f" Yb ({nb_yb}brs)",f"{kg_yb:.2f}kg")
        with rd4: st.metric(" Total Acero",f"{kg_ace_ld:.2f}kg"); st.metric(" Costo",f"{moneda} {costo_c:,.0f}")

    # ─────── ZAPATA CORRIDA ───────
    elif modo_cim == "→ Zapata Corrida":
        with ci1:
            st.markdown("** Dimensiones**")
            esp_zc2=st.number_input("Espesor [m]",0.0,value=0.30,step=0.05,key="kc_cim_eszc")
            lar_zc2=st.number_input("Largo [m]",0.0,value=10.0,step=0.5,key="kc_cim_larzc")
            anc_zc2=st.number_input("Ancho [m]",0.0,value=1.0,step=0.1,key="kc_cim_anczc")
            fc_zc2=st.selectbox("Resistencia [kg/cm²]:",[m["fc_kgcm2"] for m in MIX_DESIGNS],index=3,key="kc_cim_fczc")
            cant_cim=st.number_input("Cantidad",min_value=1,value=1,key="kc_cim_qty_5")
            mix_zc2=next(m for m in MIX_DESIGNS if m["fc_kgcm2"]==fc_zc2)
            rec_zc2=st.number_input("Recub. lat. [m]",0.0,value=0.05,step=0.01,key="kc_cim_reczc")
            desp_czc=st.select_slider("Desp.Concreto [%]",[1,2,3,4,5,6],value=1,key="kc_cim_dczc"); desp_azc=st.select_slider("Desp.Acero [%]",[1,2,3,4,5,6],value=1,key="kc_cim_dazc")
        with ci2:
            st.markdown("** Acero X (transversal)**")
            lx_zc=st.number_input("Longitud X [m]",0.0,value=anc_zc2-2*rec_zc2,step=0.05,key="kc_cim_lxzc")
            sx_zc=st.number_input("Separación X [m]",0.0,value=0.30,step=0.05,key="kc_cim_sxzc")
            vx_zc=st.selectbox("Ø X:",[v["nombre"] for v in VARILLAS],index=3,key="kc_cim_vxzc"); vkdxzc=next(v for v in VARILLAS if v["nombre"]==vx_zc)
        with ci3:
            st.markdown("** Acero Y (longitudinal)**")
            ly_zc=st.number_input("Longitud Y [m]",0.0,value=lar_zc2-2*rec_zc2,step=0.1,key="kc_cim_lyzc")
            sy_zc=st.number_input("Separación Y [m]",0.0,value=0.30,step=0.05,key="kc_cim_syzc")
            vy_zc=st.selectbox("Ø Y:",[v["nombre"] for v in VARILLAS],index=3,key="kc_cim_vyzc"); vkdyzc=next(v for v in VARILLAS if v["nombre"]==vy_zc)
            trzy=st.radio("Traslapes Y:",["NO","SI"],horizontal=True,key="kc_cim_trzcy"); ltzy=st.number_input("Long.trasl.[m]",0.0,value=0.35,step=0.05,key="kc_cim_ltzcy") if trzy=="SI" else 0.0
        vol_zc2=esp_zc2*lar_zc2*anc_zc2*cant_cim*(1+desp_czc/100)
        bol_cim=math.ceil(mix_zc2["cem_kg"]*vol_zc2/R["peso_bolsa"]); arena_cim=mix_zc2["arena_m3"]*vol_zc2; grava_cim=mix_zc2["grava_m3"]*vol_zc2; agua_cim=mix_zc2["agua_lt"]*vol_zc2
        nb_xzc=math.floor(anc_zc2/max(sx_zc,0.01))+1; nb_yzc=math.floor(lar_zc2/max(sy_zc,0.01))+1
        kg_xzc=nb_xzc*lx_zc*vkdxzc["kg_m"]*cant_cim*(1+desp_azc/100); kg_yzc=nb_yzc*(ly_zc+ltzy)*vkdyzc["kg_m"]*cant_cim*(1+desp_azc/100)
        kg_ace_zc=kg_xzc+kg_yzc; costo_c=bol_cim*float(p.get("cemento",0))+arena_cim*float(p.get("arena",0))+grava_cim*float(p.get("grava",0))+kg_ace_zc*float(p.get("acero_kg",0))
        st.markdown("---"); st.markdown("####  Resultados")
        rzc1,rzc2,rzc3,rzc4=st.columns(4)
        with rzc1: st.metric(" Cemento",f"{bol_cim} bultos"); st.metric(" Arena",f"{arena_cim:.2f}m³"); st.metric(" Grava",f"{grava_cim:.2f}m³"); st.metric(" Agua",f"{agua_cim:.0f}lt")
        with rzc2: st.metric(f" Acero X ({nb_xzc}brs)",f"{kg_xzc:.2f}kg")
        with rzc3: st.metric(f" Acero Y ({nb_yzc}brs)",f"{kg_yzc:.2f}kg")
        with rzc4: st.metric(" Total",f"{kg_ace_zc:.2f}kg"); st.metric(" Costo",f"{moneda} {costo_c:,.0f}")

    # ─────── ELECTROMALLA ───────
    elif modo_cim == "⚡ Electromalla":
        kg_ace_lc = 0
        with ci1:
            st.markdown("** Dimensiones**")
            anc_em=st.number_input("Ancho [m]",0.0,value=1.0,step=0.1,key="kc_cim_aem")
            lar_em=st.number_input("Largo [m]",0.0,value=12.0,step=0.5,key="kc_cim_lem")
            esp_em=st.number_input("Espesor [m]",0.0,value=0.30,step=0.05,key="kc_cim_eem")
            cant_cim=st.number_input("Cantidad",min_value=1,value=1,key="kc_cim_qty_6")
            fc_em=st.selectbox("Resistencia [kg/cm²]:",[m["fc_kgcm2"] for m in MIX_DESIGNS],index=2,key="kc_cim_fcem")
            mix_em=next(m for m in MIX_DESIGNS if m["fc_kgcm2"]==fc_em)
            desp_cem=st.select_slider("Desp.Concreto [%]",[1,2,3,4,5,6,8],value=4,key="kc_cim_dcem"); desp_aem=st.select_slider("Desp.Electromalla [%]",[1,2,3,4,5,6,8],value=5,key="kc_cim_daem")
        with ci2:
            st.markdown("**⚡ Electromalla**")
            tipo_em=st.selectbox("Tipo:",list(_ELECTROMALLA_AREA.keys()),key="kc_cim_tem")
        area_em=anc_em*lar_em*cant_cim
        vol_em=area_em*esp_em*(1+desp_cem/100)
        bol_cim=math.ceil(mix_em["cem_kg"]*vol_em/R["peso_bolsa"]); arena_cim=mix_em["arena_m3"]*vol_em; grava_cim=mix_em["grava_m3"]*vol_em; agua_cim=mix_em["agua_lt"]*vol_em
        n_piezas_em=math.ceil(area_em*(1+desp_aem/100)/_ELECTROMALLA_AREA[tipo_em])
        costo_c=bol_cim*float(p.get("cemento",0))+arena_cim*float(p.get("arena",0))+grava_cim*float(p.get("grava",0))
        st.markdown("---"); st.markdown("####  Resultados")
        rem1,rem2=st.columns(2)
        with rem1: st.metric(" Cemento",f"{bol_cim} bultos"); st.metric(" Arena",f"{arena_cim:.2f}m³"); st.metric(" Grava",f"{grava_cim:.2f}m³"); st.metric(" Agua",f"{agua_cim:.0f}lt")
        with rem2: st.metric("⚡ Electromalla",f"{n_piezas_em} piezas"); st.metric(" Dosif.",mix_em["dos"]); st.metric(" Volumen",f"{vol_em:.2f}m³"); st.metric(" Costo",f"{moneda} {costo_c:,.0f}")

    # ─────── SUELO CEMENTO ───────
    elif modo_cim == " Suelo Cemento":
        kg_ace_lc = 0; arena_cim = 0; grava_cim = 0
        with ci1:
            st.markdown("** Dimensiones**")
            vol_sc=st.number_input("Volumen [m³]",0.0,value=11.0,step=0.5,key="kc_cim_vsc")
            cant_cim=st.number_input("Cantidad",min_value=1,value=1,key="kc_cim_qty_7")
        with ci2:
            st.markdown("** Suelo Cemento**")
            dos_sc=st.selectbox("Dosificación (Suelo:Cemento):",["20:1","15:1","10:1","8:1","6:1","5:1"],index=0,key="kc_cim_dossc")
            proctor=st.number_input("Proctor [kg/m³]",0.0,value=1550.0,step=10.0,key="kc_cim_proctor")
        dos_ratio = float(dos_sc.split(":")[0])
        masa_suelo = vol_sc * cant_cim * proctor
        masa_cem   = masa_suelo / dos_ratio
        bol_cim    = math.ceil(masa_cem / 50)
        vol_suelo  = vol_sc * cant_cim
        agua_cim   = 0; arena_cim = 0; grava_cim = 0
        costo_c    = bol_cim*float(p.get("cemento",0))
        st.markdown("---"); st.markdown("####  Resultados")
        rsc1,rsc2=st.columns(2)
        with rsc1: st.metric(" Cemento",f"{bol_cim} bultos"); st.metric(" Suelo Selecto",f"{vol_suelo:.2f} m³")
        with rsc2: st.metric("Dosif.",dos_sc); st.metric("Volumen",f"{vol_suelo:.2f} m³"); st.metric(" Costo cem",f"{moneda} {costo_c:,.0f}")

    # ─────── BOTÓN AGREGAR ───────
    # Summary strip
    st.markdown(
        f'<div style="background:#0d2137;border-radius:8px;padding:8px 16px;">'
        f'<span style="color:#ffcc80;">{modo_cim}</span></div>', unsafe_allow_html=True)
    if st.button(" Agregar Cimiento al Resumen", key="kc_add_cim", type="primary"):
        lb = desc_cim or "Cimiento"
        rows=[{"elemento":f"{lb} — {R['cemento']}","unidad":"bultos","cant":bol_cim,"precio":float(p.get("cemento",0))}]
        if arena_cim: rows.append({"elemento":f"{lb} — Arena","unidad":"m³","cant":round(arena_cim,3),"precio":float(p.get("arena",0))})
        if grava_cim: rows.append({"elemento":f"{lb} — Grava","unidad":"m³","cant":round(grava_cim,3),"precio":float(p.get("grava",0))})
        st.session_state.kc_rows.extend(rows)
        st.success(f" {lb} agregado al resumen")


# ══════════ TAB 7 — MURO DE CONTENCIÓN (CM-V3.0 · 11 modos) ══════════
with tabs[6]:
    st.subheader(" Calculadora de Muro de Contención")
    st.caption(f"CM-V3.0 | Ciclópeo (6) + Estructural (4) | {norma_sel}")

    # Mortar mix constants per m³ mortar (calibrated from CM-V3.0)
    _MCIM = {
        "1:2": {"bol":13.0,"sand":0.60,"water":180},
        "1:3": {"bol":12.0,"sand":1.10,"water":260},
        "1:4": {"bol":8.0, "sand":1.20,"water":235},
        "1:5": {"bol":6.3, "sand":1.25,"water":252},
        "1:6": {"bol":5.0, "sand":1.30,"water":260},
    }

    MODOS_MURO = [
        " Ciclópeo Trapecio",
        " Ciclópeo Trapecio+Pie",
        "▮  Ciclópeo Rectangular",
        "▮ Dos Secciones",
        "▮ Dos Secciones+Pie",
        "m³ Volumen Directo",
        " Estructural L-básico",
        " Estructural L-completo",
        " Estructural 2 Capas",
        " Estructural 2 Capas+H.Cim",
    ]
    modo_m = st.radio("Modo:", MODOS_MURO, horizontal=True, key="kc_muro_modo")
    st.markdown("---")
    desc_m = st.text_input("Descripción", placeholder="Ej: Muro MC-1", key="kc_muro_desc")

    # ======= DIAGRAMA TÉCNICO PROFESIONAL DE MURO =======
    _is_ciclopeo_m = "Ciclópeo" in modo_m
    _is_estructural_m = "Estructural" in modo_m
    _is_voldir_m = "Volumen" in modo_m

    _w_svg_m, _h_svg_m = 260, 200
    _svg_corte_m, _svg_planta_m = "", ""

    _defs_m = """<defs>
      <pattern id="soil_m" width="8" height="8" patternUnits="userSpaceOnUse" patternTransform="rotate(45)">
        <line x1="0" y1="0" x2="0" y2="8" stroke="#5d4037" stroke-width="0.8" opacity="0.4"/>
        <line x1="2" y1="0" x2="2" y2="8" stroke="#5d4037" stroke-width="0.3" opacity="0.2"/>
      </pattern>
      <pattern id="stone_m" width="16" height="16" patternUnits="userSpaceOnUse">
        <rect width="16" height="16" fill="#1e3a5f" opacity="0.9"/>
        <path d="M2,2 Q4,0 6,2 T5,6 Q2,8 1,5 Z m8,8 Q12,8 14,10 T13,14 Q10,16 9,13 Z m-6,4 Q4,10 6,12" fill="#7f8c8d" opacity="0.5"/>
      </pattern>
      <pattern id="conc_m" width="10" height="10" patternUnits="userSpaceOnUse">
        <rect width="10" height="10" fill="#1e3a5f" opacity="0.8"/>
        <circle cx="2" cy="2" r="0.8" fill="#4a9eff" opacity="0.5"/>
        <circle cx="7" cy="6" r="1.2" fill="#4a9eff" opacity="0.5"/>
      </pattern>
    </defs>"""

    _soil_bg_m = '<rect x="0" y="80" width="130" height="120" fill="url(#soil_m)"/><line x1="0" y1="80" x2="130" y2="80" stroke="#795548" stroke-width="2"/><path d="M 15 80 L 10 70 L 20 70 Z" fill="#795548"/>'

    if _is_ciclopeo_m:
        if "Rectangular" in modo_m:
            _svg_corte_m += f'{_soil_bg_m}<rect x="130" y="30" width="60" height="150" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
            _svg_planta_m += '<rect x="60" y="20" width="70" height="160" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
        elif "Trapecio" in modo_m and "Pie" not in modo_m:
            _svg_corte_m += f'{_soil_bg_m}<polygon points="130,30 170,30 210,180 130,180" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
            _svg_planta_m += '<rect x="50" y="20" width="100" height="160" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
        elif "Trapecio+Pie" in modo_m:
            _svg_corte_m += f'{_soil_bg_m}<polygon points="130,30 170,30 210,150 130,150" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/><rect x="100" y="150" width="110" height="30" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
            _svg_planta_m += '<rect x="40" y="20" width="120" height="160" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
        elif "Dos Secciones" in modo_m and "Pie" not in modo_m:
            _svg_corte_m += f'{_soil_bg_m}<polygon points="130,30 170,30 190,140 130,140" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/><rect x="130" y="140" width="90" height="40" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
            _svg_planta_m += '<rect x="50" y="20" width="110" height="160" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
        elif "Dos Secciones+Pie" in modo_m:
            _svg_corte_m += f'{_soil_bg_m}<polygon points="130,30 170,30 190,130 130,130" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/><rect x="100" y="130" width="120" height="50" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
            _svg_planta_m += '<rect x="40" y="20" width="130" height="160" fill="url(#stone_m)" stroke="#bdc3c7" stroke-width="1.5"/>'
        
        if not _is_voldir_m:
            _svg_corte_m += '<line x1="240" y1="30" x2="240" y2="180" stroke="#ccc" stroke-width="1" stroke-dasharray="2,2"/><text x="235" y="105" fill="#ccc" font-size="10" transform="rotate(-90 235,105)" text-anchor="middle">Alt H</text>'

    elif _is_estructural_m:
        _svg_corte_m += f'{_soil_bg_m}'
        _svg_corte_m += '<rect x="90" y="150" width="140" height="30" fill="url(#conc_m)" stroke="#4a9eff" stroke-width="1.5"/>'
        _svg_corte_m += '<rect x="130" y="20" width="30" height="130" fill="url(#conc_m)" stroke="#4a9eff" stroke-width="1.5" stroke-dasharray="0,0,30,0"/>'
        _svg_corte_m += '<line x1="95" y1="174" x2="225" y2="174" stroke="#ff6b35" stroke-width="2"/>'
        for i in range(8): _svg_corte_m += f'<circle cx="{100 + i*17}" cy="170" r="2" fill="#00d4ff"/>'
        _svg_corte_m += '<line x1="135" y1="25" x2="135" y2="174" stroke="#ff6b35" stroke-width="2"/>'
        if "2 Cap" in modo_m: _svg_corte_m += '<line x1="155" y1="25" x2="155" y2="174" stroke="#ff6b35" stroke-width="2"/>'
        for i in range(8):
            _svg_corte_m += f'<circle cx="139" cy="{35 + i*15}" r="2" fill="#00d4ff"/>'
            if "2 Cap" in modo_m: _svg_corte_m += f'<circle cx="151" cy="{35 + i*15}" r="2" fill="#00d4ff"/>'
                
        _svg_planta_m += '<rect x="60" y="20" width="140" height="160" fill="url(#conc_m)" stroke="#4a9eff" stroke-width="1.5"/>'
        _svg_planta_m += '<rect x="100" y="20" width="30" height="160" fill="none" stroke="#bdc3c7" stroke-width="2" stroke-dasharray="4,2"/>'
        for i in range(8): _svg_planta_m += f'<line x1="65" y1="{30 + i*20}" x2="195" y2="{30 + i*20}" stroke="#00d4ff" stroke-width="1.5"/>'
        for i in range(5): _svg_planta_m += f'<line x1="{70 + i*30}" y1="20" x2="{70 + i*30}" y2="180" stroke="#ff6b35" stroke-width="1.5"/>'

    if not _is_voldir_m:
        _html_muro_diag = f'''
        <div style="background:#0a192f; padding:16px; border-radius:8px; border:1px solid #1e3a5f; margin:12px 0 24px 0; display:flex; justify-content:space-around; align-items:center; flex-wrap:wrap; box-shadow:0 4px 15px rgba(0,0,0,0.3);">
          <div style="text-align:center; min-width:270px;">
            <h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Corte Transversal</h4>
            <svg width="{_w_svg_m}" height="{_h_svg_m}" viewBox="0 0 {_w_svg_m} {_h_svg_m}" xmlns="http://www.w3.org/2000/svg">
              {_defs_m}{_svg_corte_m}
            </svg>
          </div>
          <div style="text-align:center; min-width:270px; border-left:1px dashed #1e3a5f;">
            <h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Vista en Planta</h4>
            <svg width="{_w_svg_m}" height="{_h_svg_m}" viewBox="0 0 {_w_svg_m} {_h_svg_m}" xmlns="http://www.w3.org/2000/svg">
              {_defs_m}{_svg_planta_m}
            </svg>
          </div>
        </div>
        '''
        import streamlit.components.v1 as _stc_m
        _stc_m.html(_html_muro_diag, height=270, scrolling=False)

    # ─────── CICLÓPEO SHAPES ───────
    if modo_m in [" Ciclópeo Trapecio"," Ciclópeo Trapecio+Pie",
                  "▮  Ciclópeo Rectangular","▮ Dos Secciones",
                  "▮ Dos Secciones+Pie","m³ Volumen Directo"]:
        mc1,mc2,mc3 = st.columns(3)
        with mc1:
            st.markdown("** Dimensiones**")
            if modo_m == "m³ Volumen Directo":
                vol_base_m = st.number_input("Volumen [m³]",0.0,value=5.0,step=0.5,key="kc_muro_voldir")
                cant_m = st.number_input("Cantidad",min_value=1,value=1,key="kc_muro_qty_1")
                vol_muro_neto = vol_base_m * cant_m
            else:
                corona_m = st.number_input("Corona C [m]",0.0,value=0.30,step=0.05,key="kc_muro_C")
                if modo_m != "▮  Ciclópeo Rectangular":
                    base_m = st.number_input("Base B [m]",0.0,value=1.0,step=0.1,key="kc_muro_B")
                else:
                    base_m = corona_m
                alt_m = st.number_input("Altura H [m]",0.0,value=1.50,step=0.1,key="kc_muro_H")
                if modo_m in [" Ciclópeo Trapecio+Pie","▮ Dos Secciones+Pie"]:
                    pie_m = st.number_input("Pie I [m]",0.0,value=0.30,step=0.05,key="kc_muro_I")
                else:
                    pie_m = 0.0
                if modo_m in ["▮ Dos Secciones","▮ Dos Secciones+Pie"]:
                    base2_m = st.number_input("Base 2 B [m]",0.0,value=2.0,step=0.1,key="kc_muro_B2")
                    alt2_m  = st.number_input("Altura 2 H [m]",0.0,value=0.30,step=0.05,key="kc_muro_H2")
                else:
                    base2_m = 0.0; alt2_m = 0.0
                largo_m = st.number_input("Largo [m]",0.0,value=5.0,step=0.5,key="kc_muro_L")
                cant_m  = st.number_input("Cantidad",min_value=1,value=1,key="kc_muro_qty_2")
                # Volume formula by shape
                if modo_m == "▮  Ciclópeo Rectangular":
                    vol_muro_neto = corona_m * alt_m * largo_m * cant_m
                elif modo_m in [" Ciclópeo Trapecio"," Ciclópeo Trapecio+Pie"]:
                    vol_muro_neto = ((corona_m+base_m)/2 * alt_m + corona_m*pie_m) * largo_m * cant_m
                else:  # Dos secciones
                    vol_muro_neto = ((corona_m+base_m)/2 * alt_m + base2_m*alt2_m) * largo_m * cant_m
                    if modo_m == "▮ Dos Secciones+Pie":
                        vol_muro_neto += corona_m * pie_m * largo_m * cant_m
                st.caption(f"Vol neto: {vol_muro_neto:.2f}m³")
        with mc2:
            st.markdown("** Piedra**")
            pct_p_m = st.selectbox("% Piedra:",[55,60,65,70,75,80,85,90],index=2,key="kc_muro_pp")
            desp_p_m = st.selectbox("Desp. Piedra [%]:",[1,2,3,4,5,6,7,8,10],index=3,key="kc_muro_dp")
            st.markdown("** Mortero de Pega**")
            dos_m = st.selectbox("Dosificación:",["1:2","1:3","1:4","1:5","1:6"],index=2,key="kc_muro_dos")
            pct_mort_m = st.selectbox("% Mortero:",[20,25,30,35,40,45,50],index=2,key="kc_muro_pmort")
            desp_mort_m = st.selectbox("Desp. Mortero [%]:",[1,2,3,4,5,6,7,8,10],index=3,key="kc_muro_dm")
        with mc3:
            st.caption("")  # spacing

        # Calc ciclópeo
        vol_p_neto = vol_muro_neto * (pct_p_m/100)
        vol_p_desp = vol_p_neto * (1+desp_p_m/100)
        vol_mort_n = vol_muro_neto * (pct_mort_m/100)
        vol_mort   = vol_mort_n * (1+desp_mort_m/100)
        mx_m = _MCIM[dos_m]
        bol_m  = math.ceil(vol_mort * mx_m["bol"])
        arena_m = vol_mort * mx_m["sand"]
        agua_m  = vol_mort * mx_m["water"]
        kg_ace_m = 0; grava_m = 0
        costo_m = vol_p_desp*float(p.get("piedra",65)) + bol_m*float(p.get("cemento",0)) + arena_m*float(p.get("arena",0))
        st.markdown("---"); st.markdown("####  Resultados")
        rm1,rm2 = st.columns(2)
        with rm1:
            st.metric(" Piedra",f"{vol_p_desp:.2f}m³"); st.metric(" Cemento",f"{bol_m} bultos")
            st.metric(" Arena",f"{arena_m:.2f}m³"); st.metric(" Agua",f"{agua_m:.0f} lt")
        with rm2:
            st.metric("Dosif.",dos_m); st.metric(f"% Piedra",f"{pct_p_m}%")
            st.metric("Vol Piedra",f"{vol_p_neto:.2f} m³"); st.metric("Vol Mortero",f"{vol_mort_n:.2f} m³")
            st.metric(" Costo est.",f"{moneda} {costo_m:,.0f}")

    # ─────── ESTRUCTURAL ───────
    else:
        is_2capas = "2 Cap" in modo_m
        has_hcim  = "H.Cim" in modo_m
        is_completo = "completo" in modo_m

        me1,me2,me3 = st.columns(3)
        with me1:
            st.markdown("** Concreto / Dimensiones**")
            alt1_e  = st.number_input("Altura 1 H [m]",0.0,value=3.0,step=0.1,key="kc_muro_h1e")
            alt2_e  = st.number_input("Altura 2 H [m]",0.0,value=0.40,step=0.05,key="kc_muro_h2e")
            base_e  = st.number_input("Base B [m]",0.0,value=2.0,step=0.1,key="kc_muro_be")
            corona_e= st.number_input("Corona C [m]",0.0,value=0.40,step=0.05,key="kc_muro_ce")
            largo_e = st.number_input("Largo [m]",0.0,value=5.0,step=0.5,key="kc_muro_le")
            cant_m  = st.number_input("Cantidad",min_value=1,value=1,key="kc_muro_qty_3")
            fc_e    = st.selectbox("Resistencia [kg/cm²]:",[m["fc_kgcm2"] for m in MIX_DESIGNS],index=3,key="kc_muro_fce")
            mix_e   = next(m for m in MIX_DESIGNS if m["fc_kgcm2"]==fc_e)
            rec_e   = st.number_input("Recubrimiento [m]",0.0,value=0.05,step=0.01,key="kc_muro_rece")
            desp_ce = st.select_slider("Desp.Concr.[%]",[1,2,3,4,5,6,7,8],value=5,key="kc_muro_dce")
            desp_ae = st.select_slider("Desp.Acero [%]",[1,2,3,4,5,6,7,8,9],value=5,key="kc_muro_dae")
        with me2:
            st.markdown("** Acero Vertical**")
            if not is_2capas:
                lv1=st.number_input("Long. Vert. Var1 [m]",0.0,value=3.5,step=0.1,key="kc_muro_lv1")
                sv1=st.number_input("Sep. Var1 [m]",0.0,value=0.30,step=0.05,key="kc_muro_sv1_1")
                vv1=st.selectbox("Ø Var1:",[v["nombre"] for v in VARILLAS],index=2,key="kc_muro_vv1_1"); vkdv1=next(v for v in VARILLAS if v["nombre"]==vv1)
                lv2=st.number_input("Long. Vert. Var2 [m]",0.0,value=3.0,step=0.1,key="kc_muro_lv2")
                sv2=st.number_input("Sep. Var2 [m]",0.0,value=0.40,step=0.05,key="kc_muro_sv2_1")
                vv2=st.selectbox("Ø Var2:",[v["nombre"] for v in VARILLAS],index=2,key="kc_muro_vv2_1"); vkdv2=next(v for v in VARILLAS if v["nombre"]==vv2)
                lv12=lv22=sv12=sv22=0.0; vkdv12=vkdv22=None
            else:
                lv1=st.number_input("Long. Var1.1 [m]",0.0,value=3.10,step=0.1,key="kc_muro_lv11"); lv12=st.number_input("Long. Var1.2 [m]",0.0,value=3.20,step=0.1,key="kc_muro_lv12")
                sv1=st.number_input("Sep. Var1 [m]",0.0,value=0.20,step=0.05,key="kc_muro_sv1_2")
                vv1=st.selectbox("Ø Var1:",[v["nombre"] for v in VARILLAS],index=2,key="kc_muro_vv1_2"); vkdv1=next(v for v in VARILLAS if v["nombre"]==vv1); vkdv12=vkdv1
                lv2=st.number_input("Long. Var2.1 [m]",0.0,value=3.30,step=0.1,key="kc_muro_lv21"); lv22=st.number_input("Long. Var2.2 [m]",0.0,value=3.40,step=0.1,key="kc_muro_lv22")
                sv2=st.number_input("Sep. Var2 [m]",0.0,value=0.30,step=0.05,key="kc_muro_sv2_2")
                vv2=st.selectbox("Ø Var2:",[v["nombre"] for v in VARILLAS],index=1,key="kc_muro_vv2_2"); vkdv2=next(v for v in VARILLAS if v["nombre"]==vv2); vkdv22=vkdv2
            if is_completo or has_hcim:
                st.markdown("** Acero H. Cimiento**")
                lhc1=st.number_input("Long. H.Cim. Var1 [m]",0.0,value=1.20,step=0.1,key="kc_muro_lhc"); shc1=st.number_input("Sep. H.Cim. [m]",0.0,value=0.30,step=0.05,key="kc_muro_shc")
                vhc1=st.selectbox("Ø H.Cim.:",[v["nombre"] for v in VARILLAS],index=2,key="kc_muro_vhc"); vkdhc=next(v for v in VARILLAS if v["nombre"]==vhc1)
            else:
                lhc1=0;vkdhc=None;shc1=0.30
        with me3:
            st.markdown("** Acero Longitudinal**")
            llm=st.number_input("Long. En Muro [m]",0.0,value=5.5,step=0.1,key="kc_muro_llm")
            slm=st.number_input("Sep. En Muro [m]",0.0,value=0.50,step=0.05,key="kc_muro_slm")
            vlm=st.selectbox("Ø Long. Muro:",[v["nombre"] for v in VARILLAS],index=3,key="kc_muro_vlm"); vkdlm=next(v for v in VARILLAS if v["nombre"]==vlm)
            llc=st.number_input("Long. En Cimiento [m]",0.0,value=5.5,step=0.1,key="kc_muro_llc")
            slc=st.number_input("Sep. En Cimiento [m]",0.0,value=0.60,step=0.05,key="kc_muro_slc")
            vlc=st.selectbox("Ø Long. Cim.:",[v["nombre"] for v in VARILLAS],index=3,key="kc_muro_vlc"); vkdlc=next(v for v in VARILLAS if v["nombre"]==vlc)
            trasl_le=st.radio("Traslapes Long.:",["NO","SI"],horizontal=True,key="kc_muro_trl"); lt_le=st.number_input("Long.Trasl.[m]",0.0,value=0.35,step=0.05,key="kc_muro_ltl") if trasl_le=="SI" else 0.0

        # Calc estructural
        vol_e = (corona_e * alt1_e + base_e * alt2_e) * largo_e * cant_m * (1+desp_ce/100)
        bol_m   = math.ceil(mix_e["cem_kg"]*vol_e/R["peso_bolsa"])
        arena_m = mix_e["arena_m3"]*vol_e; grava_m = mix_e["grava_m3"]*vol_e; agua_m = mix_e["agua_lt"]*vol_e
        # Vertical bars (run along LARGO, spaced along wall/base)
        n_v1 = math.floor(largo_e/max(sv1,0.01))
        n_v2 = math.floor(largo_e/max(sv2,0.01))
        kg_v1 = n_v1 * lv1 * vkdv1["kg_m"] * cant_m * (1+desp_ae/100)
        kg_v2 = n_v2 * lv2 * vkdv2["kg_m"] * cant_m * (1+desp_ae/100)
        if is_2capas and vkdv12 and vkdv22:
            kg_v1 += n_v1 * lv12 * vkdv12["kg_m"] * cant_m * (1+desp_ae/100)
            kg_v2 += n_v2 * lv22 * vkdv22["kg_m"] * cant_m * (1+desp_ae/100)
        # Longitudinal muro (horizontal bars along height)
        n_lm = math.floor(alt1_e/max(slm,0.01))
        n_lc = math.floor(base_e/max(slc,0.01))
        kg_lm = n_lm * (llm+lt_le) * vkdlm["kg_m"] * cant_m * (1+desp_ae/100)
        kg_lc = n_lc * (llc+lt_le) * vkdlc["kg_m"] * cant_m * (1+desp_ae/100)
        # H. Cimiento horizontal bars in slab thickness direction
        kg_hc = 0
        if (is_completo or has_hcim) and vkdhc and lhc1>0:
            n_hc = math.floor(largo_e/max(shc1,0.01))
            kg_hc = n_hc * lhc1 * vkdhc["kg_m"] * cant_m * (1+desp_ae/100)
        kg_ace_m = kg_v1 + kg_v2 + kg_lm + kg_lc + kg_hc
        costo_m = bol_m*float(p.get("cemento",0))+arena_m*float(p.get("arena",0))+grava_m*float(p.get("grava",0))+kg_ace_m*float(p.get("acero_kg",0))
        st.markdown("---"); st.markdown("####  Resultados")
        re1,re2,re3,re4 = st.columns(4)
        with re1: st.metric(" Cemento",f"{bol_m} bultos"); st.metric(" Arena",f"{arena_m:.2f} m³"); st.metric(" Grava",f"{grava_m:.2f} m³"); st.metric(" Agua",f"{agua_m:.0f} lt")
        with re2: st.metric(" Acero Vert_1",f"{kg_v1:.2f}kg"); st.metric(" Acero Vert_2",f"{kg_v2:.2f}kg")
        with re3: st.metric(" Long. Muro",f"{kg_lm:.2f}kg"); st.metric(" Long. Cim.",f"{kg_lc:.2f}kg"); st.metric(" H.Cim.",f"{kg_hc:.2f}kg")
        with re4: st.metric(" Acero Total",f"{kg_ace_m:.2f}kg"); st.metric(" Vol.Concreto",f"{vol_e:.2f}m³"); st.metric(" Costo",f"{moneda} {costo_m:,.0f}")

    # ─────── BOTÓN AGREGAR ───────
    st.markdown(f'<div style="background:#0d2137;border-radius:8px;padding:8px 16px;"><span style="color:#ffcc80;">{modo_m}</span></div>',unsafe_allow_html=True)
    if st.button(" Agregar Muro al Resumen", key="kc_add_muro", type="primary"):
        lb = desc_m or f"Muro {modo_m[:10]}"
        rows=[{"elemento":f"{lb} — {R['cemento']}","unidad":"bultos","cant":bol_m,"precio":float(p.get("cemento",0))}]
        if arena_m: rows.append({"elemento":f"{lb} — Arena","unidad":"m³","cant":round(arena_m,3),"precio":float(p.get("arena",0))})
        if grava_m: rows.append({"elemento":f"{lb} — Grava","unidad":"m³","cant":round(grava_m,3),"precio":float(p.get("grava",0))})
        if kg_ace_m: rows.append({"elemento":f"{lb} — Acero","unidad":"kg","cant":round(kg_ace_m,2),"precio":float(p.get("acero_kg",0))})
        st.session_state.kc_rows.extend(rows)
        st.success(f" {lb}: {bol_m} bultos cemento | {kg_ace_m:.1f} kg acero | Costo: {moneda} {costo_m:,.0f}")


# ══════════ TAB 14 — SALARIO MÍNIMO Y LIQUIDACIÓN ══════════
with tabs[14]:
    sal = SALARIOS_MIN.get(pais)
    if sal:
        st.subheader(f" Salario Mínimo y Liquidación Laboral — {pais}")
        st.caption("Cálculo estimado de nómina y prestaciones sociales.")
        
        # Selector de año
        _hist = HIST_SAL.get(pais, {})
        _anios = list(_hist.keys()) if _hist else [2026]
        anio_sel = st.selectbox(" Año de liquidación:", _anios, key="kc_anio_liq")
        
        # Obtener valores
        s_base = sal.get("salario_base", 0)
        aux_tr = sal.get("auxilio_transporte", 0)
        nota_anio = sal.get("nota", "")
        
        if anio_sel in _hist:
            s_base = _hist[anio_sel]["salario_base"]
            aux_tr = _hist[anio_sel].get("auxilio_transporte", 0)
            nota_anio = _hist[anio_sel].get("nota", "")
        
        devengado = s_base + aux_tr
        
        # Tarjeta info 
        st.markdown(
            f'<div style="background:#0f2a0f;border-radius:8px;padding:12px;margin-bottom:12px;border:1px solid #2e592e;">'
            f'<div style="display:flex;justify-content:space-between;margin-bottom:4px;">'
            f'<span style="color:#a8e6cf;font-weight:600;"> Salario base</span>'
            f'<span style="color:#c8e6c9;">{moneda} {s_base:,.0f}/mes</span></div>'
            f'<div style="display:flex;justify-content:space-between;margin-bottom:4px;">'
            f'<span style="color:#a8e6cf;font-weight:600;"> Auxilio transporte</span>'
            f'<span style="color:#c8e6c9;">{moneda} {aux_tr:,.0f}/mes</span></div>'
            f'<div style="display:flex;justify-content:space-between;border-top:1px solid #2e592e;padding-top:4px;">'
            f'<span style="color:#ffd54f;font-weight:700;"> Total devengado</span>'
            f'<span style="color:#fff;font-weight:700;">{moneda} {devengado:,.0f}/mes</span></div>'
            f'<div style="display:flex;justify-content:space-between;margin-top:8px;">'
            f'<span style="color:#a8e6cf;font-weight:500;"> Jornal diario</span>'
            f'<span style="color:#81d4fa;">{moneda} {(devengado/30):,.0f}/día</span></div>'
            f'<div style="display:flex;justify-content:space-between;">'
            f'<span style="color:#a8e6cf;font-weight:500;">⏱ Costo por hora</span>'
            f'<span style="color:#81d4fa;">{moneda} {(devengado/240):,.0f}/h</span></div>'
            f'</div>', unsafe_allow_html=True
        )
        st.caption(f"ℹ {nota_anio}")
        
        # Simulador liquidación rápida
        st.markdown("---")
        st.markdown("###  Simulación de Liquidación (Fin de Contrato)")
        col_f1, col_f2 = st.columns(2)
        import datetime
        with col_f1:
            fecha_ingreso = st.date_input("Fecha de Ingreso", datetime.date(anio_sel, 1, 1), key="kc_f_ing")
        with col_f2:
            fecha_salida = st.date_input("Fecha de Retiro", datetime.date(anio_sel, 12, 31), key="kc_f_sal")
            
        if fecha_salida >= fecha_ingreso:
            dias_trabajados = (fecha_salida - fecha_ingreso).days + 1
            st.info(f"⏳ **Tiempo laborado:** {dias_trabajados} días")
            
            # Cálculo genérico proporcional al salario
            base_liq = devengado
            cesantias = (base_liq * dias_trabajados) / 360
            intereses = (cesantias * 0.12 * dias_trabajados) / 360 # Ojo: 12% anual sobre cesantías
            prima = (base_liq * dias_trabajados) / 360
            vacaciones = (s_base * dias_trabajados) / 720 # sobre salario base sin aux transporte
            
            total_liq = cesantias + intereses + prima + vacaciones
            
            st.markdown(f"- **Cesantías:** `{moneda} {cesantias:,.0f}`")
            st.markdown(f"- **Intereses:** `{moneda} {intereses:,.0f}`")
            st.markdown(f"- **Prima:** `{moneda} {prima:,.0f}`")
            st.markdown(f"- **Vacaciones:** `{moneda} {vacaciones:,.0f}`")
            st.markdown(f"#### Total: `{moneda} {total_liq:,.0f}`")
            
            # Export to DOC
            html_doc = f"""<html xmlns:o='urn:schemas-microsoft-com:office:office' xmlns:w='urn:schemas-microsoft-com:office:word' xmlns='http://www.w3.org/TR/REC-html40'>
            <head><meta charset='utf-8'></head><body>
            <h2>Liquidación Laboral - Konte ({pais})</h2>
            <p><b>Fecha de Ingreso:</b> {fecha_ingreso.strftime('%d/%m/%Y')}</p>
            <p><b>Fecha de Retiro:</b> {fecha_salida.strftime('%d/%m/%Y')}</p>
            <p><b>Días Trabajados:</b> {dias_trabajados}</p>
            <p><b>Salario Base de Liquidación:</b> {moneda} {base_liq:,.0f}</p>
            <hr>
            <ul>
                <li><b>Cesantías:</b> {moneda} {cesantias:,.0f}</li>
                <li><b>Intereses a las Cesantías:</b> {moneda} {intereses:,.0f}</li>
                <li><b>Prima de Servicios:</b> {moneda} {prima:,.0f}</li>
                <li><b>Vacaciones:</b> {moneda} {vacaciones:,.0f}</li>
            </ul>
            <h3>Total Liquidación: {moneda} {total_liq:,.0f}</h3>
            <p><small>Generado con Konte Calculadora - Estimación referencial, verifique con la normativa local vigente.</small></p>
            </body></html>"""
            
            st.download_button(" Exportar a DOC", data=html_doc.encode('utf-8'), file_name="Liquidacion_Laboral.doc", mime="application/msword", use_container_width=True)
        else:
            st.warning("La fecha de retiro debe ser mayor al ingreso.")

        # Tabla comparativa últimos años si hay histórico
        if _hist:
            st.markdown("---")
            st.markdown(f"** Histórico salarios ({pais}):**")
            filas_h = []
            for yr in sorted(_hist.keys(), reverse=True):
                dh = _hist[yr]
                filas_h.append({
                    "Año": yr,
                    f"Salario ({moneda})": f"{dh['salario_base']:,.0f}",
                    "Aux.Transp.": f"{dh.get('auxilio_transporte',0):,.0f}",
                    "Total": f"{dh['salario_base']+dh.get('auxilio_transporte',0):,.0f}",
                })
            import pandas as pd
            st.dataframe(pd.DataFrame(filas_h).set_index("Año"), use_container_width=True)

# ══════════ TAB 16 — CONFIGURACIÓN (13 subtabs) ══════════
with tabs[15]:
    st.subheader("⚙ Configuración del Sistema")
    st.caption(f"Personalice dosificaciones, materiales y costos | {norma_sel} | {moneda}")

    # ── Init config session state ──
    import pandas as pd

    def _init_cfg(key, default_df):
        if key not in st.session_state:
            st.session_state[key] = default_df

    _init_cfg("kc_cfg_mix", pd.DataFrame(st.session_state.get("_MIX_DESIGNS", [
        {"Dos.": "1:2:2",   "fc(kg/cm²)": 280, "Cem.(kg)": 420, "Arena(m³)": 0.67, "Grava(m³)": 0.67, "Agua(lt)": 190},
        {"Dos.": "1:2:2.5", "fc(kg/cm²)": 240, "Cem.(kg)": 380, "Arena(m³)": 0.60, "Grava(m³)": 0.76, "Agua(lt)": 180},
        {"Dos.": "1:2:3",   "fc(kg/cm²)": 226, "Cem.(kg)": 350, "Arena(m³)": 0.55, "Grava(m³)": 0.84, "Agua(lt)": 170},
        {"Dos.": "1:2:3.5", "fc(kg/cm²)": 210, "Cem.(kg)": 320, "Arena(m³)": 0.52, "Grava(m³)": 0.90, "Agua(lt)": 170},
        {"Dos.": "1:2:4",   "fc(kg/cm²)": 200, "Cem.(kg)": 300, "Arena(m³)": 0.48, "Grava(m³)": 0.95, "Agua(lt)": 158},
        {"Dos.": "1:2.5:4", "fc(kg/cm²)": 189, "Cem.(kg)": 280, "Arena(m³)": 0.55, "Grava(m³)": 0.89, "Agua(lt)": 158},
        {"Dos.": "1:3:3",   "fc(kg/cm²)": 168, "Cem.(kg)": 300, "Arena(m³)": 0.72, "Grava(m³)": 0.72, "Agua(lt)": 158},
        {"Dos.": "1:3:4",   "fc(kg/cm²)": 159, "Cem.(kg)": 260, "Arena(m³)": 0.63, "Grava(m³)": 0.83, "Agua(lt)": 163},
        {"Dos.": "1:3:5",   "fc(kg/cm²)": 140, "Cem.(kg)": 230, "Arena(m³)": 0.55, "Grava(m³)": 0.92, "Agua(lt)": 148},
        {"Dos.": "1:3:6",   "fc(kg/cm²)": 119, "Cem.(kg)": 210, "Arena(m³)": 0.50, "Grava(m³)": 1.00, "Agua(lt)": 143},
        {"Dos.": "1:4:7",   "fc(kg/cm²)": 109, "Cem.(kg)": 175, "Arena(m³)": 0.55, "Grava(m³)": 0.98, "Agua(lt)": 133},
        {"Dos.": "1:4:8",   "fc(kg/cm²)": 99,  "Cem.(kg)": 160, "Arena(m³)": 0.55, "Grava(m³)": 1.03, "Agua(lt)": 125},
    ])))
    _init_cfg("kc_cfg_mort", pd.DataFrame([
        {"Dos.":"1:2","Cem.(kg/m³)":610,"Arena(m³/m³)":0.97,"Agua(lt/m³)":250},
        {"Dos.":"1:3","Cem.(kg/m³)":600,"Arena(m³/m³)":1.10,"Agua(lt/m³)":250},
        {"Dos.":"1:4","Cem.(kg/m³)":364,"Arena(m³/m³)":1.16,"Agua(lt/m³)":240},
        {"Dos.":"1:5","Cem.(kg/m³)":302,"Arena(m³/m³)":1.20,"Agua(lt/m³)":240},
        {"Dos.":"1:6","Cem.(kg/m³)":261,"Arena(m³/m³)":1.20,"Agua(lt/m³)":235},
    ]))
    _init_cfg("kc_cfg_acero", pd.DataFrame([
        {"Diámetro":"N2-1/4\"","m/var":6.0,"Kg/m":0.248},
        {"Diámetro":"N3-3/8\"","m/var":6.0,"Kg/m":0.558},
        {"Diámetro":"N4-1/2\"","m/var":6.0,"Kg/m":0.994},
        {"Diámetro":"N5-5/8\"","m/var":6.0,"Kg/m":1.542},
        {"Diámetro":"N5.5-11/16\"","m/var":6.0,"Kg/m":1.632},
        {"Diámetro":"N6-3/4\"","m/var":6.0,"Kg/m":2.222},
        {"Diámetro":"N7-7/8\"","m/var":6.0,"Kg/m":3.022},
        {"Diámetro":"N8-1\"","m/var":6.0,"Kg/m":3.950},
        {"Diámetro":"N10-1¼\"","m/var":6.0,"Kg/m":6.173},
        {"Diámetro":"N12-1½\"","m/var":6.0,"Kg/m":8.892},
        {"Diámetro":"N16-2\"","m/var":6.0,"Kg/m":15.803},
    ]))
    _init_cfg("kc_cfg_mamp_blq", pd.DataFrame([
        {"Nombre":"B-10x20x40","UN/m²":12.50,"Mortero m³/m²":0.00775},
        {"Nombre":"B-12x20x40","UN/m²":12.50,"Mortero m³/m²":0.009025},
        {"Nombre":"B-15x20x40","UN/m²":12.50,"Mortero m³/m²":0.011500},
        {"Nombre":"B-20x20x40","UN/m²":12.50,"Mortero m³/m²":0.015250},
        {"Nombre":"B-Per1","UN/m²":6.00,"Mortero m³/m²":0.005300},
        {"Nombre":"B-Per2","UN/m²":5.00,"Mortero m³/m²":0.005000},
    ]))
    _init_cfg("kc_cfg_mamp_lad", pd.DataFrame([
        {"Nombre":"L-4x10.5x22cm","UN/m²":15.0,"Mortero m³/m²":0.007750},
        {"Nombre":"L-6x10.5x22cm","UN/m²":15.0,"Mortero m³/m²":0.009025},
        {"Nombre":"L-10x10x20cm","UN/m²":15.0,"Mortero m³/m²":0.011500},
        {"Nombre":"L-20x15x40cm","UN/m²":15.0,"Mortero m³/m²":0.015250},
        {"Nombre":"L-A","UN/m²":5.0,"Mortero m³/m²":0.005000},
        {"Nombre":"L-B","UN/m²":5.0,"Mortero m³/m²":0.005000},
    ]))
    _init_cfg("kc_cfg_bovedilla", pd.DataFrame([
        {"Nombre":"bovedilla 1","Ancho(m)":0.30,"Largo(m)":0.30},
        {"Nombre":"bovedilla 2","Ancho(m)":0.30,"Largo(m)":0.30},
        {"Nombre":"bovedilla 3","Ancho(m)":0.30,"Largo(m)":0.30},
    ]))
    _init_cfg("kc_cfg_electromalla", pd.DataFrame([
        {"Nombre":"electromalla 1","Ancho(m)":2.50,"Largo(m)":1.00},
        {"Nombre":"electromalla 2","Ancho(m)":2.50,"Largo(m)":5.00},
        {"Nombre":"electromalla 3","Ancho(m)":2.50,"Largo(m)":10.00},
    ]))
    _init_cfg("kc_cfg_perfiles_pared", pd.DataFrame([
        {"Nombre":"Parante","Longitud(m)":2.44},
        {"Nombre":"Parante","Longitud(m)":2.44},
        {"Nombre":"Canal","Longitud(m)":2.44},
        {"Nombre":"Canal","Longitud(m)":2.44},
    ]))
    _init_cfg("kc_cfg_perfiles_cielo", pd.DataFrame([
        {"Nombre":"Omegas","Longitud(m)":2.44},
        {"Nombre":"Omegas","Longitud(m)":2.60},
        {"Nombre":"Viguetas","Longitud(m)":2.44},
        {"Nombre":"Viguetas","Longitud(m)":2.60},
        {"Nombre":"Ang.Per","Longitud(m)":2.44},
        {"Nombre":"Ang.Per","Longitud(m)":2.60},
    ]))
    _init_cfg("kc_cfg_panel_yeso", pd.DataFrame([
        {"Nombre":"Panel","Ancho(m)":1.22,"Largo(m)":2.44,"Área(m²)":1.22*2.44},
        {"Nombre":"Panel","Ancho(m)":1.23,"Largo(m)":2.45,"Área(m²)":1.23*2.45},
    ]))
    _init_cfg("kc_cfg_laminas", pd.DataFrame([
        {"Nombre":"ZincAlum",          "Ancho(m)":1.00,"Tras.L(m)":0.20,"Tras.T(m)":0.20},
        {"Nombre":"Metal Galvanizada", "Ancho(m)":1.00,"Tras.L(m)":0.20,"Tras.T(m)":0.20},
        {"Nombre":"Lamina 3",          "Ancho(m)":1.00,"Tras.L(m)":0.20,"Tras.T(m)":0.20},
        {"Nombre":"Lamina 4",          "Ancho(m)":1.10,"Tras.L(m)":0.10,"Tras.T(m)":0.15},
    ]))
    _init_cfg("kc_cfg_tejas", pd.DataFrame([
        {"Nombre":"teja 1","unidades/m²":5.0},
        {"Nombre":"teja 2","unidades/m²":6.0},
        {"Nombre":"teja 3","unidades/m²":7.0},
        {"Nombre":"teja 4","unidades/m²":8.0},
        {"Nombre":"teja 5","unidades/m²":9.0},
    ]))
    _init_cfg("kc_cfg_ceramica", pd.DataFrame([
        {"Nombre":"CE 1","Ancho(m)":0.25,"Largo(m)":0.25,"Área(m²)":0.25*0.25,"UN/caja":10},
        {"Nombre":"CE 2","Ancho(m)":0.40,"Largo(m)":0.40,"Área(m²)":0.40*0.40,"UN/caja": 8},
        {"Nombre":"CE 3","Ancho(m)":0.50,"Largo(m)":0.50,"Área(m²)":0.50*0.50,"UN/caja": 6},
    ]))
    _init_cfg("kc_cfg_porcelanato", pd.DataFrame([
        {"Nombre":"PN 1","Ancho(m)":0.30,"Largo(m)":0.30,"Área(m²)":0.30*0.30,"UN/caja":10},
        {"Nombre":"PN 2","Ancho(m)":0.40,"Largo(m)":0.40,"Área(m²)":0.40*0.40,"UN/caja": 8},
        {"Nombre":"PN 3","Ancho(m)":0.15,"Largo(m)":0.60,"Área(m²)":0.15*0.60,"UN/caja": 6},
    ]))
    _init_cfg("kc_cfg_adh_cer", pd.DataFrame([{"Material":"AD Cerámica","REN(m²/bol)":2.0,"Agua(lt/bol)":9.45}]))
    _init_cfg("kc_cfg_boq_cer", pd.DataFrame([{"Material":"BQ Cerámica","REN(m²/bol)":5.0,"Agua(lt/bol)":8.50}]))
    _init_cfg("kc_cfg_adh_pn",  pd.DataFrame([{"Material":"AD Porcelanato","REN(m²/bol)":2.0,"Agua(lt/bol)":9.45}]))
    _init_cfg("kc_cfg_boq_pn",  pd.DataFrame([{"Material":"BQ Porcelanato","REN(m²/bol)":5.0,"Agua(lt/bol)":8.50}]))

    # ── Aproximaciones init ──
    _init_cfg("kc_cfg_aprox", pd.DataFrame([
        {"Material":"Bolsas Cemento y Afinado","Aprox":"no"},
        {"Material":"m3/m2/m","Aprox":"no"},
        {"Material":"lt/barr/gal","Aprox":"no"},
        {"Material":"Blocks/Ladrillos","Aprox":"no"},
        {"Material":"Cajas Piso","Aprox":"no"},
        {"Material":"kg","Aprox":"no"},
        {"Material":"Bovedillas Losa","Aprox":"no"},
        {"Material":"Tejas Techo","Aprox":"no"},
        {"Material":"Láminas Techo","Aprox":"no"},
        {"Material":"Panel Yeso","Aprox":"no"},
        {"Material":"Moneda","Aprox":"no"},
        {"Material":"Masilla Yeso","Aprox":"no"},
        {"Material":"Electromalla","Aprox":"no"},
    ]))

    # Reference prices PER COUNTRY  (no USD from CM-V3.0)
    _ref_cos = PRICE_URLS.get(pais, {}).get("precios_ref", {"cemento":30,"arena":50,"grava":60,"acero_kg":3})
    _is_cop  = (moneda == "COP $")
    _init_cfg("kc_cfg_costos", pd.DataFrame([
        {"Material": f"Cemento bolsa (bolsa)",         "Costo": float(p.get("cemento", _ref_cos.get("cemento",30)))},
        {"Material": f"Arena (m³)",                    "Costo": float(p.get("arena",   _ref_cos.get("arena",  50)))},
        {"Material": f"Grava (m³)",                    "Costo": float(p.get("grava",   _ref_cos.get("grava",  60)))},
        {"Material": f"Agua (lt)",                     "Costo": 0.003 if not _is_cop else 1.5},
        {"Material": f"Agua (barr)",                   "Costo": 3.0   if not _is_cop else 2000},
        {"Material": f"Agua (gal)",                    "Costo": 0.010 if not _is_cop else 5.0},
        {"Material": f"Bloque 10x20x40 (unidad)",      "Costo": float(p.get("bloque",  2.0 if not _is_cop else 1200))},
        {"Material": f"Bloque 12x20x40 (unidad)",      "Costo": 2.2  if not _is_cop else 1400},
        {"Material": f"Bloque 15x20x40 (unidad)",      "Costo": 2.5  if not _is_cop else 1600},
        {"Material": f"Bloque 20x20x40 (unidad)",      "Costo": 3.0  if not _is_cop else 2000},
        {"Material": f"Ladrillo (unidad)",              "Costo": 0.4  if not _is_cop else 400},
        {"Material": f"Piedra (m³)",                   "Costo": 40.0 if not _is_cop else 65000},
        {"Material": f"Acero/Varilla (kg)",            "Costo": float(p.get("acero_kg",_ref_cos.get("acero_kg",3)))},
        {"Material": f"Pintura (galón)",               "Costo": float(p.get("pintura", 40.0 if not _is_cop else 55000))},
        {"Material": f"Afinado instantáneo (bolsa)",   "Costo": 9.0  if not _is_cop else 9400},
        {"Material": f"Parante 2.44m (unidad)",        "Costo": 2.5  if not _is_cop else 8500},
        {"Material": f"Canal 2.44m (unidad)",          "Costo": 2.5  if not _is_cop else 7500},
        {"Material": f"Yeso Panel 1.22x2.44m (unidad)","Costo": 8.0  if not _is_cop else 25000},
        {"Material": "Tornillo Estructura 6x1\" (un)","Costo": 0.02 if not _is_cop else 50},
        {"Material": "Tornillo Panel 7x7/16\" (un)",  "Costo": 0.03 if not _is_cop else 65},
        {"Material": f"Masilla Yeso (cubeta 6kg)",     "Costo": 12.0 if not _is_cop else 35000},
        {"Material": f"Cinta Papel (rollo 100m)",      "Costo": 2.0  if not _is_cop else 5500},
        {"Material": f"Lija (unidad)",                 "Costo": 0.1  if not _is_cop else 350},
        {"Material": f"Aislante T.Acústico (m²)",      "Costo": 2.0  if not _is_cop else 6500},
        {"Material": f"Omega 2.44m (unidad)",          "Costo": 2.5  if not _is_cop else 7500},
        {"Material": f"Vigueta 2.44m (unidad)",        "Costo": 2.6  if not _is_cop else 8500},
        {"Material": f"Ang.Perim 2.44m (unidad)",      "Costo": 2.3  if not _is_cop else 6000},
        {"Material": f"ZincAlum (m)",                  "Costo": 6.9  if not _is_cop else 18000},
        {"Material": f"Metal Galvanizada (m)",         "Costo": 5.0  if not _is_cop else 13000},
        {"Material": f"Teja Barro (unidad)",           "Costo": 0.25 if not _is_cop else 650},
        {"Material": f"CE 1 Cerámica (caja)",          "Costo": float(p.get("ceramica",12.0 if not _is_cop else 35000))},
        {"Material": f"CE 2 Cerámica (caja)",          "Costo": 14.0 if not _is_cop else 42000},
        {"Material": f"PN 1 Porcelanato (caja)",       "Costo": 16.0 if not _is_cop else 55000},
        {"Material": f"PN 2 Porcelanato (caja)",       "Costo": 18.0 if not _is_cop else 68000},
        {"Material": f"AD Cerámica (bolsa)",           "Costo": 8.0  if not _is_cop else 22000},
        {"Material": f"BQ Cerámica (bolsa)",           "Costo": 6.0  if not _is_cop else 18000},
        {"Material": f"AD Porcelanato (bolsa)",        "Costo": 10.0 if not _is_cop else 28000},
        {"Material": f"BQ Porcelanato (bolsa)",        "Costo": 7.0  if not _is_cop else 22000},
    ]))

    # ── 13 CONFIG SUB-TABS ──
    cfg_tabs = st.tabs([" Concreto"," Mortero"," Cemento/Afinado"," Unidades"," Aproximaciones",
                         f" Costos ({moneda})"," Mampostería"," Acero"," Losa"," Perfiles",
                         " Panel Yeso"," Techo"," Piso"])

    # ─ 0: Concreto ─
    with cfg_tabs[0]:
        st.markdown(f"**Dosificaciones del Concreto para 1m³** — Bolsa: {R['peso_bolsa']}kg")
        st.caption("Verifique las dosificaciones según su normativa local.")
        edited_mix = st.data_editor(st.session_state.kc_cfg_mix, num_rows="dynamic", use_container_width=True, key="de_mix")
        if st.button(" Guardar Concreto",key="cfg_sv_mix"):
            st.session_state.kc_cfg_mix = edited_mix; st.success(" Dosificaciones guardadas")

    # ─ 1: Mortero ─
    with cfg_tabs[1]:
        st.markdown("**Dosificaciones del Mortero para 1m³**")
        st.caption("Configure las mezclas de mortero para su región.")
        edited_mort = st.data_editor(st.session_state.kc_cfg_mort, num_rows="dynamic", use_container_width=True, key="de_mort")
        if st.button(" Guardar Mortero",key="cfg_sv_mort"):
            st.session_state.kc_cfg_mort = edited_mort; st.success(" Mortero guardado")

    # ─ 2: Cemento/Afinado ─
    with cfg_tabs[2]:
        ca1, ca2 = st.columns(2)
        with ca1:
            st.markdown("**Cemento**")
            cem_peso = st.number_input("Peso bolsa (Kg)", value=float(R["peso_bolsa"]), min_value=1.0, step=0.5, key="cfg_cem_pg")
            st.caption(f"Bolsas de {cem_peso}kg — {R['cemento']}")
        with ca2:
            st.markdown("**Afinado Instantáneo**")
            st.caption("Configure grosores y rendimiento según su proveedor.")
            afinado_df = pd.DataFrame([
                {"Tipo":"(nombre)","Grosor(mm)":2.00,"Grosor(m)":0.002,"REN(m²/bol)":7.00,"Agua(lt/bol)":9.45},
                {"Tipo":"-",       "Grosor(mm)":5.00,"Grosor(m)":0.005,"REN(m²/bol)":3.50,"Agua(lt/bol)": 0},
                {"Tipo":"-",       "Grosor(mm)":7.00,"Grosor(m)":0.007,"REN(m²/bol)":2.50,"Agua(lt/bol)": 0},
                {"Tipo":"-",       "Grosor(mm)":19.00,"Grosor(m)":0.019,"REN(m²/bol)":1.50,"Agua(lt/bol)":0},
            ])
            st.data_editor(afinado_df, use_container_width=True, key="de_afinado")

    # ─ 3: Unidades ─
    with cfg_tabs[3]:
        st.markdown("**Unidades de Agua**")
        agua_opt = st.selectbox("Unidades de Agua:", ["lt","barr","gal"], key="cfg_agua_ud")
        st.caption(f"Unidad seleccionada: **{agua_opt}** — Las unidades se aplican solo a los resultados, no al ingreso de datos.")
        st.info("ℹ Las conversiones: 1 barr ≈ 159 lt | 1 gal ≈ 3.785 lt")
        if st.button(" Guardar Unidades",key="cfg_sv_ud"):
            st.session_state.kc_cfg_agua_ud = agua_opt; st.success(f" Unidad de agua: {agua_opt}")

    # ─ 4: Aproximaciones ─
    with cfg_tabs[4]:
        st.markdown("**Aproximación de Unidades**")
        st.caption('"no" = ninguna aproximación | "entero" = aproximar al entero más próximo')
        aprox_opts = ["no","entero"]
        aprox_df = st.session_state.kc_cfg_aprox.copy()
        edited_aprox = st.data_editor(
            aprox_df,
            column_config={"Aprox": st.column_config.SelectboxColumn("Aprox", options=aprox_opts)},
            num_rows="fixed", use_container_width=True, key="de_aprox")
        if st.button(" Guardar Aproximaciones",key="cfg_sv_aprox"):
            st.session_state.kc_cfg_aprox = edited_aprox; st.success(" Aproximaciones guardadas")

    # ─ 5: Costos ─
    with cfg_tabs[5]:
        st.markdown(f"**Costos de Materiales — {moneda}**")
        st.caption(f" País: **{pais}** | Los costos están en **{moneda}**. Actualice según los precios actuales de su región.")
        cc1, cc2 = st.columns([3,1])
        with cc2:
            if st.button(" Consultar Precios en Vivo", key="cfg_live_prices", use_container_width=True):
                with st.spinner("Consultando precios..."):
                    live = get_live_prices(pais, _ref_cos); st.session_state.kc_precios.update(live)
                st.success(" Precios actualizados desde sitios web")
        with cc1:
            edited_cos = st.data_editor(st.session_state.kc_cfg_costos, num_rows="dynamic", use_container_width=True, key="de_costos",
                column_config={"Costo": st.column_config.NumberColumn(f"Costo ({moneda})", format="%.2f", min_value=0.0)})
            if st.button(" Guardar Costos",key="cfg_sv_cos"):
                st.session_state.kc_cfg_costos = edited_cos
                # Sync key prices back to p dict
                for row in edited_cos.itertuples():
                    nm = str(row.Material).lower()
                    if "cemento" in nm and "bultos" in nm: p["cemento"] = float(row.Costo)
                    elif "arena" in nm: p["arena"] = float(row.Costo)
                    elif "grava" in nm: p["grava"] = float(row.Costo)
                    elif "acero" in nm or "varilla" in nm: p["acero_kg"] = float(row.Costo)
                    elif "pintura" in nm: p["pintura"] = float(row.Costo)
                    elif "bloque" in nm and "10x20" in nm: p["bloque"] = float(row.Costo)
                st.session_state.kc_precios = p; st.success(f" Costos en {moneda} guardados y aplicados")

    # ─ 6: Mampostería ─
    with cfg_tabs[6]:
        m1, m2 = st.columns(2)
        with m1:
            st.markdown("**Mampostería Bloque**")
            st.caption("UN/m²: unidades por m². MORTERO: m³ por m².")
            edited_blq = st.data_editor(st.session_state.kc_cfg_mamp_blq, num_rows="dynamic", use_container_width=True, key="de_blq")
            if st.button(" Guardar Bloques",key="cfg_sv_blq"):
                st.session_state.kc_cfg_mamp_blq = edited_blq; st.success(" Bloques guardados")
        with m2:
            st.markdown("**Mampostería Ladrillo**")
            edited_lad = st.data_editor(st.session_state.kc_cfg_mamp_lad, num_rows="dynamic", use_container_width=True, key="de_lad")
            if st.button(" Guardar Ladrillos",key="cfg_sv_lad"):
                st.session_state.kc_cfg_mamp_lad = edited_lad; st.success(" Ladrillos guardados")

    # ─ 7: Acero ─
    with cfg_tabs[7]:
        st.markdown("**Configuración de Acero / Varillas**")
        st.caption("DIÁMETRO: nombre de la varilla | m/var: longitud de cada varilla | Kg/m: peso por metro.")
        edited_acer = st.data_editor(st.session_state.kc_cfg_acero, num_rows="dynamic", use_container_width=True, key="de_acer")
        if st.button(" Guardar Acero",key="cfg_sv_ace"):
            st.session_state.kc_cfg_acero = edited_acer; st.success(" Acero guardado")

    # ─ 8: Losa ─
    with cfg_tabs[8]:
        l1, l2 = st.columns(2)
        with l1:
            st.markdown("**Bovedillas de Losa**")
            st.caption("Ingrese ancho y largo según el esquema para calcular área.")
            edited_bov = st.data_editor(st.session_state.kc_cfg_bovedilla, num_rows="dynamic", use_container_width=True, key="de_bov")
            if st.button(" Guardar Bovedillas",key="cfg_sv_bov"):
                st.session_state.kc_cfg_bovedilla = edited_bov; st.success(" Bovedillas guardadas")
        with l2:
            st.markdown("**Electromalla**")
            edited_elm = st.data_editor(st.session_state.kc_cfg_electromalla, num_rows="dynamic", use_container_width=True, key="de_elm")
            if st.button(" Guardar Electromalla",key="cfg_sv_elm"):
                st.session_state.kc_cfg_electromalla = edited_elm; st.success(" Electromalla guardada")

    # ─ 9: Perfiles ─
    with cfg_tabs[9]:
        p1, p2 = st.columns(2)
        with p1:
            st.markdown("**Perfiles Pared: Panel Yeso**")
            edited_ppared = st.data_editor(st.session_state.kc_cfg_perfiles_pared, num_rows="dynamic", use_container_width=True, key="de_ppared")
            if st.button(" Guardar Perfiles Pared",key="cfg_sv_ppared"):
                st.session_state.kc_cfg_perfiles_pared = edited_ppared; st.success(" Guardado")
            st.markdown("---"); st.markdown("**Tornillos Pared (Tor/m²)**")
            st.number_input('Tornillo Estructura 6x1" [Tor/m²]', value=10.0, key="cfg_torn_est_pr")
            st.number_input('Tornillo Panel 7x7/16" [Tor/m²]',  value=5.0,  key="cfg_torn_pan_pr")
        with p2:
            st.markdown("**Perfiles Cielo Raso: Panel Yeso**")
            edited_pcielo = st.data_editor(st.session_state.kc_cfg_perfiles_cielo, num_rows="dynamic", use_container_width=True, key="de_pcielo")
            if st.button(" Guardar Perfiles Cielo",key="cfg_sv_pcielo"):
                st.session_state.kc_cfg_perfiles_cielo = edited_pcielo; st.success(" Guardado")
            st.markdown("---"); st.markdown("**Tornillos Cielo Raso (Tor/m²)**")
            st.number_input('Tornillo Estructura 6x1" [Tor/m²]', value=10.0, key="cfg_torn_est_cr")
            st.number_input('Tornillo Panel 7x7/16" [Tor/m²]',  value=5.0,  key="cfg_torn_pan_cr")

    # ─ 10: Panel Yeso ─
    with cfg_tabs[10]:
        py1, py2, py3 = st.columns(3)
        with py1:
            st.markdown("**Paneles Pared y Cielo Raso**")
            edited_py = st.data_editor(st.session_state.kc_cfg_panel_yeso, num_rows="dynamic", use_container_width=True, key="de_py")
            if st.button(" Guardar Paneles",key="cfg_sv_py"):
                st.session_state.kc_cfg_panel_yeso = edited_py; st.success("")
        with py2:
            st.markdown("**Masilla Yeso**")
            st.number_input("REN Masilla (Kg/m²)", value=1.1, key="cfg_mas_ren")
            st.number_input("CUBETA Masilla (Kg)", value=6.0, key="cfg_mas_cub")
            st.markdown("**Cinta Papel**")
            st.number_input("ROLLO (m)", value=100.0, key="cfg_cinta_rollo")
            st.number_input("REN Cinta (m/m²)", value=12.0, key="cfg_cinta_ren")
        with py3:
            st.markdown("**Chazo Puntilla (CP/m²)**")
            st.number_input("REN Chazo Puntilla", value=1.0, key="cfg_chazo_ren")
            st.markdown("**Lija Panel Yeso**")
            st.number_input("REN Lija (L/Pa)", value=0.10, key="cfg_lija_ren")
            st.markdown("**Aislante Termo-Acústico**")
            st.number_input("m² por unidad", value=10.0, key="cfg_aisl_m2")

    # ─ 11: Techo ─
    with cfg_tabs[11]:
        t1, t2 = st.columns(2)
        with t1:
            st.markdown("**Láminas — Dimensiones y Traslapes**")
            st.caption("ANCHO: ancho total de la lámina | TRAS.L: traslape longitudinal | TRAS.T: traslape transversal")
            edited_lam = st.data_editor(st.session_state.kc_cfg_laminas, num_rows="dynamic", use_container_width=True, key="de_lam")
            if st.button(" Guardar Láminas",key="cfg_sv_lam"):
                st.session_state.kc_cfg_laminas = edited_lam; st.success(" Láminas guardadas")
        with t2:
            st.markdown("**Teja — Cobertura (unidades/m²)**")
            edited_tja = st.data_editor(st.session_state.kc_cfg_tejas, num_rows="dynamic", use_container_width=True, key="de_tja")
            if st.button(" Guardar Tejas",key="cfg_sv_tja"):
                st.session_state.kc_cfg_tejas = edited_tja; st.success(" Tejas guardadas")

    # ─ 12: Piso ─
    with cfg_tabs[12]:
        ps1, ps2 = st.columns(2)
        with ps1:
            st.markdown("**Cerámica**")
            edited_cer = st.data_editor(st.session_state.kc_cfg_ceramica, num_rows="dynamic", use_container_width=True, key="de_cer")
            if st.button(" Guardar Cerámica",key="cfg_sv_cer"):
                st.session_state.kc_cfg_ceramica = edited_cer; st.success("")
            st.markdown("**Porcelanato**")
            edited_pn = st.data_editor(st.session_state.kc_cfg_porcelanato, num_rows="dynamic", use_container_width=True, key="de_pn")
            if st.button(" Guardar Porcelanato",key="cfg_sv_pn"):
                st.session_state.kc_cfg_porcelanato = edited_pn; st.success("")
        with ps2:
            st.markdown("**Adhesivo Cerámica**")
            st.data_editor(st.session_state.kc_cfg_adh_cer, use_container_width=True, key="de_adh_cer")
            st.markdown("**Boquilla Cerámica**")
            st.data_editor(st.session_state.kc_cfg_boq_cer, use_container_width=True, key="de_boq_cer")
            st.markdown("**Adhesivo Porcelanato**")
            st.data_editor(st.session_state.kc_cfg_adh_pn, use_container_width=True, key="de_adh_pn")
            st.markdown("**Boquilla Porcelanato**")
            st.data_editor(st.session_state.kc_cfg_boq_pn, use_container_width=True, key="de_boq_pn")
            if st.button(" Guardar Adhesivos / Boquillas",key="cfg_sv_adh"):
                st.success(" Constantes de instalación guardadas")


# ══════════ TAB RESUMEN — TECHO (CM-V3.0 · 7 modos) ══════════
with tabs[7]:
    st.subheader(" Calculadora de Techo")
    st.caption(f"CM-V3.0 | Tejas (1A/2A/4A/Área) + Láminas (1/2/3 Filas) | {norma_sel}")

    # Tile type lookup: typical coverage in Colombia (uds/m²)
    TIPOS_TEJA = {
        "Teja Barro Española": 5.0,
        "Teja Barro Media Canal": 6.0,
        "Teja Eternit Ondulada": 2.5,
        "Teja Ondulada Termo-acustica": 1.4,
        "Teja Canaleta Zinc": 1.3,
        "Personalizado": 0.0,
    }
    # Lamina effective coverage width (m) per type
    TIPOS_LAM = {
        "ZincAlum": {"eff": 0.72, "precio_m": 6.90},
        "Zinc calibrado": {"eff": 0.72, "precio_m": 5.50},
        "Aluminio": {"eff": 0.80, "precio_m": 9.00},
        "Galvalume": {"eff": 0.72, "precio_m": 7.50},
        "Policarbonato": {"eff": 0.90, "precio_m": 12.00},
        "Teja PVC": {"eff": 0.72, "precio_m": 8.50},
    }

    MODOS_TECHO = [" 1A — Mono inclinación"," 2A — Doble agua"," 4A — Cuatro aguas"," Área directa",
                   " Láminas 1 Fila"," Láminas 2 Filas"," Láminas 3 Filas"]
    modo_t = st.radio("Modo:", MODOS_TECHO, horizontal=True, key="kc_techo_modo")
    st.markdown("---")
    desc_t = st.text_input("Descripción", placeholder="Ej: Techo bloque 2", key="kc_techo_desc")

    # ─────── TEJAS (4 modos) ───────
    if "1A" in modo_t or "2A" in modo_t or "4A" in modo_t or "Área" in modo_t:
        tc1, tc2 = st.columns(2)
        with tc1:
            st.markdown("** Dimensiones**")
            if "Área" in modo_t:
                area_t = st.number_input("Área [m²]",0.0,value=8.0,step=0.5,key="kc_techo_area")
            else:
                ancho_t = st.number_input("Ancho [m]",0.0,value=6.0,step=0.5,key="kc_techo_anc")
                largo_t = st.number_input("Largo [m]",0.0,value=3.0,step=0.5,key="kc_techo_lar")
                alto_t  = st.number_input("Alto [m]",0.0,value=1.0,step=0.1,key="kc_techo_alt")
                if "1A" in modo_t:
                    area_t = ancho_t * math.sqrt(largo_t**2 + alto_t**2)
                elif "2A" in modo_t:
                    area_t = 2 * ancho_t * math.sqrt((largo_t/2)**2 + alto_t**2)
                else:  # 4A
                    slant_l = math.sqrt((largo_t/2)**2 + alto_t**2)
                    slant_a = math.sqrt((ancho_t/2)**2 + alto_t**2)
                    area_t = ancho_t * slant_l + largo_t * slant_a
                st.metric(" Área inclinada", f"{area_t:.2f} m²")
            cant_t = st.number_input("Cantidad (secciones)",min_value=1,value=1,key="kc_techo_qty_1")
            area_total_t = area_t * cant_t
        with tc2:
            st.markdown("** Tipo de Teja**")
            tipo_t = st.selectbox("Tipo Teja:", list(TIPOS_TEJA.keys()), key="kc_techo_tipo")
            if tipo_t == "Personalizado":
                density_t = st.number_input("Densidad [uds/m²]",0.0,value=5.0,step=0.1,key="kc_techo_dens")
            else:
                density_t = TIPOS_TEJA[tipo_t]
                st.caption(f"Cobertura: {density_t} uds/m²")
            desp_t = st.select_slider("Desperdicio [%]",[1,2,3,4,5,6,7,8,10],value=1,key="kc_techo_desp")
        n_tejas = math.ceil(area_total_t * density_t * (1 + desp_t/100))
        precio_teja = float(p.get("teja",1.0))
        costo_t = n_tejas * precio_teja
        st.markdown("---"); st.markdown("####  Resultados")
        rt1,rt2,rt3 = st.columns(3)
        with rt1: st.metric(" Tejas",f"{n_tejas} unidades"); st.metric(" Área",f"{area_total_t:.2f} m²")
        with rt2: st.metric("Tipo",tipo_t); st.metric("Densidad",f"{density_t} uds/m²")
        with rt3: st.metric(" Costo",f"{moneda} {costo_t:,.0f}")
        mat_res = [{"elemento":f"{desc_t or 'Techo'} — {tipo_t}","unidad":"unidad","cant":n_tejas,"precio":precio_teja}]

    # ─────── LÁMINAS (3 modos) ───────
    else:
        n_filas = 1 if "1 Fila" in modo_t else (2 if "2 Filas" in modo_t else 3)
        tl1, tl2 = st.columns(2)
        with tl1:
            st.markdown("** Techo**")
            ancho_l = st.number_input("Ancho [m]",0.0,value=5.0,step=0.5,key="kc_techo_lancw")
            largo_l = st.number_input("Largo [m]",0.0,value=6.0,step=0.5,key="kc_techo_llar")
            cant_t  = st.number_input("Cantidad (secciones)",min_value=1,value=1,key="kc_techo_qty_2")
            st.markdown("** Láminas**")
            tipo_lam = st.selectbox("Tipo:",list(TIPOS_LAM.keys()),key="kc_techo_tlam")
            lam_info = TIPOS_LAM[tipo_lam]
        with tl2:
            st.markdown("**Largo por fila [m]**")
            larg_filas = []
            for i in range(n_filas):
                default_fl = round(largo_l / n_filas, 2) if i < n_filas-1 else round(largo_l - round(largo_l/n_filas,2)*(n_filas-1),2)
                lf = st.number_input(f"Largo Lám.{i+1} [m]",0.0,value=float(default_fl),step=0.1,key=f"kc_techo_lf{i+1}")
                larg_filas.append(lf)

        # Calculation: n per row = ceil(Ancho / effective_width)
        n_per_row = math.ceil(ancho_l / lam_info["eff"])
        # Total metros with row-transition overlap of 0.20m per extra fila
        metros_laminas = n_per_row * (sum(larg_filas) + 0.20 * (n_filas - 1)) * cant_t
        n_laminas_total = n_per_row * n_filas * cant_t
        area_l = ancho_l * largo_l * cant_t
        precio_m_lam = lam_info["precio_m"]
        costo_t = metros_laminas * precio_m_lam
        st.markdown("---"); st.markdown("####  Resultados")
        rl1,rl2,rl3 = st.columns(3)
        with rl1: st.metric(" Láminas",f"{n_laminas_total} unidades"); st.metric(" Área",f"{area_l:.2f} m²")
        with rl2: st.metric("Tipo",tipo_lam); st.metric("Metros Lámina",f"{metros_laminas:.2f} m")
        with rl3: st.metric("$/m",f"{moneda} {precio_m_lam:.2f}"); st.metric(" Costo",f"{moneda} {costo_t:,.0f}")
        mat_res = [{"elemento":f"{desc_t or 'Techo'} — {tipo_lam}","unidad":"m","cant":round(metros_laminas,2),"precio":precio_m_lam}]

    # ─────── BOTÓN AGREGAR ───────
    st.markdown(f'<div style="background:#0d2137;border-radius:8px;padding:8px 16px;"><span style="color:#ffcc80;">{modo_t}</span></div>',unsafe_allow_html=True)
    if st.button(" Agregar Techo al Resumen", key="kc_add_techo", type="primary"):
        st.session_state.kc_rows.extend(mat_res)
        st.success(f" {desc_t or 'Techo'} agregado al resumen")


# ══════════ TAB 9 — PISOS (CM-V3.0 · 6 modos) ══════════
with tabs[8]:
    st.subheader(" Calculadora de Pisos")
    st.caption(f"CM-V3.0 | Cerámica · Porcelanato · Zócalo | {norma_sel} | {moneda}")

    # Tile type catalog: (width_m, height_m, uds_per_box)
    TIPOS_CER = {
        "CE 0.20×0.20": (0.20, 0.20, 24),
        "CE 0.25×0.25": (0.25, 0.25, 10),
        "CE 0.30×0.30": (0.30, 0.30,  9),
        "CE 0.30×0.45": (0.30, 0.45,  6),
        "CE 0.40×0.40": (0.40, 0.40,  8),
        "CE 0.45×0.45": (0.45, 0.45,  6),
        "CE 0.60×0.60": (0.60, 0.60,  4),
    }
    TIPOS_PN = {
        "PN 0.40×0.40": (0.40, 0.40, 8),
        "PN 0.45×0.45": (0.45, 0.45, 6),
        "PN 0.50×0.50": (0.50, 0.50, 5),
        "PN 0.60×0.60": (0.60, 0.60, 4),
        "PN 0.80×0.80": (0.80, 0.80, 2),
        "PN 0.60×1.00": (0.60, 1.00, 3),
        "PN 0.60×1.20": (0.60, 1.20, 2),
    }

    # Constantes calibradas desde CM-V3.0 (por m² de instalación)
    _ADH_BOL_M2  = 0.505   # bolsas de adhesivo/m²
    _ADH_LT_BOL  = 9.54    # litros agua por bolsa adhesivo
    _BOQ_BOL_M2  = 0.202   # bolsas boquilla/m²
    _BOQ_LT_BOL  = 8.68    # litros agua por bolsa boquilla

    MODOS_PISO = [
        "⬜ Cerámica — Dimensiones",
        "⬜ Cerámica — m²",
        " Porcelanato — Dimensiones",
        " Porcelanato — m²",
        "▮  Zócalo Cerámica",
        "▮  Zócalo Porcelanato",
    ]
    modo_pi = st.radio("Modo:", MODOS_PISO, horizontal=True, key="kc_piso_modo")
    st.markdown("---")
    desc_pi = st.text_input("Descripción", placeholder="Ej: Piso baño nivel 1", key="kc_piso_desc")
    cant_pi = st.number_input("Cantidad (ambientes)",min_value=1,value=1,key="kc_piso_qty")

    # ======= DIAGRAMA TÉCNICO PROFESIONAL DE PISO =======
    _is_zocalo_p = "Zócalo" in modo_pi
    _is_pn_p = "Porcelanato" in modo_pi
    _is_dim_p = "Dimensiones" in modo_pi

    _w_svg_p, _h_svg_p = 260, 200
    _svg_corte_p, _svg_planta_p = "", ""

    _color_tile = "#2c3e50" if _is_pn_p else "#bdc3c7"
    _color_grout = "#4a9eff" if _is_pn_p else "#7f8c8d"

    _defs_p = f"""<defs>
      <pattern id="tile_bg_p" width="20" height="20" patternUnits="userSpaceOnUse">
        <rect width="20" height="20" fill="{_color_tile}" stroke="{_color_grout}" stroke-width="1.5" opacity="0.8"/>
      </pattern>
      <pattern id="concrete_p" width="10" height="10" patternUnits="userSpaceOnUse">
        <rect width="10" height="10" fill="#1e3a5f" opacity="0.6"/>
        <circle cx="2" cy="2" r="1" fill="#4a9eff" opacity="0.4"/>
      </pattern>
      <pattern id="zocalo_bg_p" width="10" height="20" patternUnits="userSpaceOnUse">
        <rect width="10" height="20" fill="{_color_tile}" stroke="{_color_grout}" stroke-width="0.8" opacity="0.9"/>
      </pattern>
    </defs>"""

    if not _is_zocalo_p:
        _svg_corte_p += '<rect x="40" y="100" width="180" height="30" fill="url(#concrete_p)" stroke="#4a9eff" stroke-width="1"/><text x="130" y="120" fill="#90caf9" font-size="10" text-anchor="middle">Placa / Base</text>'
        _svg_corte_p += '<rect x="40" y="94" width="180" height="6" fill="#7f8c8d" stroke="none"/><text x="130" y="100" fill="#2c3e50" font-size="8" text-anchor="middle">Adhesivo</text>'
        _svg_corte_p += f'<rect x="40" y="86" width="180" height="8" fill="{_color_tile}" stroke="{_color_grout}" stroke-width="1" stroke-dasharray="29,1"/><text x="130" y="93" fill="#fff" font-size="8" text-anchor="middle">Baldosas</text>'
        
        if _is_dim_p:
            _svg_planta_p += f'<rect x="50" y="30" width="160" height="140" fill="url(#tile_bg_p)" stroke="#4a9eff" stroke-width="2"/>'
            _svg_planta_p += '<line x1="220" y1="30" x2="220" y2="170" stroke="#ccc" stroke-width="1" stroke-dasharray="2,2"/><text x="215" y="100" fill="#ccc" font-size="10" transform="rotate(-90 215,100)" text-anchor="middle">Largo</text>'
            _svg_planta_p += '<line x1="50" y1="20" x2="210" y2="20" stroke="#ccc" stroke-width="1" stroke-dasharray="2,2"/><text x="130" y="15" fill="#ccc" font-size="10" text-anchor="middle">Ancho</text>'
        else:
            _svg_planta_p += f'<rect x="50" y="30" width="160" height="140" fill="url(#tile_bg_p)" stroke="#4a9eff" stroke-width="2"/>'
            _svg_planta_p += '<text x="130" y="105" fill="#fff" font-size="14" text-anchor="middle">Área de Piso</text>'
    else:
        _svg_corte_p += '<rect x="130" y="40" width="30" height="130" fill="url(#concrete_p)" stroke="#4a9eff" stroke-width="1"/><text x="145" y="105" fill="#90caf9" font-size="10" text-anchor="middle" transform="rotate(-90 145,105)">Muro</text>'
        _svg_corte_p += '<rect x="40" y="170" width="130" height="20" fill="url(#concrete_p)" stroke="#4a9eff" stroke-width="1"/><text x="85" y="184" fill="#90caf9" font-size="10" text-anchor="middle">Piso</text>'
        _svg_corte_p += '<rect x="124" y="140" width="6" height="30" fill="#7f8c8d" stroke="none"/>'
        _svg_corte_p += f'<rect x="120" y="140" width="4" height="30" fill="{_color_tile}" stroke="{_color_grout}" stroke-width="0.5"/>'
        _svg_corte_p += '<line x1="110" y1="140" x2="110" y2="170" stroke="#ccc" stroke-width="1" stroke-dasharray="2,2"/><text x="105" y="155" fill="#ccc" font-size="10" text-anchor="middle" transform="rotate(-90 105,155)">Alto</text>'

        _svg_planta_p += '<rect x="40" y="80" width="180" height="30" fill="url(#concrete_p)" stroke="#4a9eff" stroke-width="1"/><text x="130" y="98" fill="#90caf9" font-size="10" text-anchor="middle">Muro (Planta)</text>'
        _svg_planta_p += f'<rect x="40" y="110" width="180" height="5" fill="url(#zocalo_bg_p)" stroke="{_color_grout}" stroke-width="1"/>'
        _svg_planta_p += '<line x1="40" y1="125" x2="220" y2="125" stroke="#ccc" stroke-width="1" stroke-dasharray="2,2"/><text x="130" y="140" fill="#ccc" font-size="10" text-anchor="middle">Longitud Total</text>'

    _html_piso_diag = f'''
    <div style="background:#0a192f; padding:16px; border-radius:8px; border:1px solid #1e3a5f; margin:12px 0 24px 0; display:flex; justify-content:space-around; align-items:center; flex-wrap:wrap; box-shadow:0 4px 15px rgba(0,0,0,0.3);">
      <div style="text-align:center; min-width:270px;">
        <h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Corte Transversal</h4>
        <svg width="{_w_svg_p}" height="{_h_svg_p}" viewBox="0 0 {_w_svg_p} {_h_svg_p}" xmlns="http://www.w3.org/2000/svg">
          {_defs_p}{_svg_corte_p}
        </svg>
      </div>
      <div style="text-align:center; min-width:270px; border-left:1px dashed #1e3a5f;">
        <h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Vista en Planta</h4>
        <svg width="{_w_svg_p}" height="{_h_svg_p}" viewBox="0 0 {_w_svg_p} {_h_svg_p}" xmlns="http://www.w3.org/2000/svg">
          {_defs_p}{_svg_planta_p}
        </svg>
      </div>
    </div>
    '''
    import streamlit.components.v1 as _stc_piso
    _stc_piso.html(_html_piso_diag, height=270, scrolling=False)

    # ─────── CERÁMICA / PORCELANATO PISO ───────
    if "Zócalo" not in modo_pi:
        is_pn  = "Porcelanato" in modo_pi
        is_dim = "Dimensiones" in modo_pi
        tipos  = TIPOS_PN if is_pn else TIPOS_CER
        mat_lbl = "Porcelanato" if is_pn else "Cerámica"

        pp1, pp2, pp3 = st.columns(3)
        with pp1:
            st.markdown("** Área de Instalación**")
            if is_dim:
                ancho_pi = st.number_input("Ancho [m]",0.0,value=5.0,step=0.5,key="kc_piso_anc")
                largo_pi = st.number_input("Largo [m]",0.0,value=5.0,step=0.5,key="kc_piso_lar")
                area_pi  = ancho_pi * largo_pi * cant_pi
            else:
                area_pi = st.number_input("Área [m²]",0.0,value=25.0,step=1.0,key="kc_piso_area") * cant_pi
            st.metric(" Área total",f"{area_pi:.2f} m²")
        with pp2:
            st.markdown(f"** Tipo {mat_lbl}**")
            tipo_pi = st.selectbox("Tipo:",list(tipos.keys()),key="kc_cfg_piso_tipo")
            w_pi,h_pi,uds_caja = tipos[tipo_pi]
            st.caption(f"Tamaño: {w_pi}×{h_pi}m | {uds_caja} uds/caja")
            precio_caja_pi = st.number_input(f"Precio caja [{moneda}]",0.0,
                value=max(35000.0 if "COP" in moneda else 12.0,0.01),step=100.0,key="kc_piso_prec")
            st.markdown("** Desperdicios**")
            desp_pi  = st.select_slider(f"{mat_lbl} [%]",[1,2,3,4,5,6,8,10],value=1,key="kc_piso_dp")
            desp_adh = st.select_slider("Adhesivo [%]",[1,2,3,4,5,6,8,10],value=1,key="kc_piso_da")
            desp_boq = st.select_slider("Boquilla [%]",[1,2,3,4,5,6,8,10],value=1,key="kc_piso_db")
        with pp3:
            st.caption("")

        area_con_desp = area_pi * (1 + desp_pi/100)
        dens      = 1.0 / (w_pi * h_pi)         # tiles/m²
        n_tiles   = math.ceil(area_con_desp * dens)
        n_cajas   = n_tiles / uds_caja
        adh_bol   = math.ceil(area_pi * _ADH_BOL_M2 * (1+desp_adh/100) * 100)/100
        adh_agua  = adh_bol * _ADH_LT_BOL
        boq_bol   = math.ceil(area_pi * _BOQ_BOL_M2 * (1+desp_boq/100) * 100)/100
        boq_agua  = boq_bol * _BOQ_LT_BOL
        costo_pi  = n_cajas * precio_caja_pi + adh_bol*float(p.get("adhesivo_piso",0)) + boq_bol*float(p.get("boquilla",0))

        st.markdown("---"); st.markdown("####  Resultados")
        ri1,ri2,ri3 = st.columns(3)
        with ri1:
            st.metric(f" {mat_lbl}",f"{n_tiles} unidades")
            st.metric(" Cajas",f"{n_cajas:.2f} cajas")
            st.metric(f" Área",f"{area_pi:.2f} m²")
            st.metric("Tipo",tipo_pi)
            st.metric("Uds/Caja",str(uds_caja))
        with ri2:
            st.metric(" Adhesivo",f"{adh_bol:.2f} bultos")
            st.metric(" Agua adhesivo",f"{adh_agua:.2f} lt")
        with ri3:
            st.metric(" Boquilla",f"{boq_bol:.2f} bultos")
            st.metric(" Agua boquilla",f"{boq_agua:.2f} lt")
            st.metric(" Costo est.",f"{moneda} {costo_pi:,.0f}")

        mat_res = [
            {"elemento":f"{desc_pi or 'Piso'} — {mat_lbl} ({tipo_pi})","unidad":"caja","cant":round(n_cajas,2),"precio":precio_caja_pi},
            {"elemento":f"{desc_pi or 'Piso'} — Adhesivo","unidad":"bol","cant":adh_bol,"precio":float(p.get("adhesivo_piso",0))},
            {"elemento":f"{desc_pi or 'Piso'} — Boquilla","unidad":"bol","cant":boq_bol,"precio":float(p.get("boquilla",0))},
        ]

    # ─────── ZÓCALO ───────
    else:
        is_pn_z = "Porcelanato" in modo_pi
        tipos_z = TIPOS_PN if is_pn_z else TIPOS_CER
        mat_lbl = "Porcelanato" if is_pn_z else "Cerámica"

        pz1,pz2,pz3 = st.columns(3)
        with pz1:
            st.markdown("** Zócalo**")
            long_z  = st.number_input("Longitud [m]",0.0,value=15.0,step=0.5,key="kc_piso_zlong")
            alto_z  = st.number_input("Altura Zócalo [m]",0.0,value=0.07,step=0.01,key="kc_piso_zalto")
            area_z  = long_z * alto_z * cant_pi
            st.metric(" Área zócalo",f"{area_z:.3f} m²")
        with pz2:
            st.markdown(f"** Tipo {mat_lbl}**")
            tipo_z  = st.selectbox("Tipo:",list(tipos_z.keys()),key="kc_piso_ztipo")
            wz,hz,uds_caja_z = tipos_z[tipo_z]
            tile_dim = max(wz,hz)  # larger dimension along length
            st.caption(f"Tamaño: {wz}×{hz}m | {uds_caja_z} uds/caja")
            precio_caja_z = st.number_input(f"Precio caja [{moneda}]",0.0,
                value=max(30000.0 if "COP" in moneda else 10.0,0.01),step=100.0,key="kc_piso_zprec")
            st.markdown("** Desperdicios**")
            desp_zi  = st.select_slider(f"{mat_lbl} [%]",[1,2,3,4,5,6,8],value=1,key="kc_piso_zdp")
            desp_adh_z = st.select_slider("Adhesivo [%]",[1,2,3,4,5,6,8],value=1,key="kc_piso_zda")
            desp_boq_z = st.select_slider("Boquilla [%]",[1,2,3,4,5,6,8],value=1,key="kc_piso_zdb")

        # Zócalo calculation: how many physical tiles needed (cut-tile counting)
        piezas_per_row_len = math.ceil(long_z * cant_pi / tile_dim)
        cuts_per_tile = max(1, int(tile_dim / max(alto_z, 0.01)))  # pieces cut from 1 tile
        n_phys_tiles  = math.ceil(piezas_per_row_len * (1+desp_zi/100) / cuts_per_tile)
        n_cajas_z     = n_phys_tiles / uds_caja_z
        n_zocalo_pcs  = piezas_per_row_len  # pieces that go on wall
        adh_bol_z  = math.ceil(area_z * _ADH_BOL_M2 * (1+desp_adh_z/100) * 100)/100
        adh_agua_z = adh_bol_z * _ADH_LT_BOL
        boq_bol_z  = math.ceil(area_z * _BOQ_BOL_M2 * (1+desp_boq_z/100) * 100)/100
        boq_agua_z = boq_bol_z * _BOQ_LT_BOL
        costo_z    = n_cajas_z*precio_caja_z + adh_bol_z*float(p.get("adhesivo_piso",0)) + boq_bol_z*float(p.get("boquilla",0))

        st.markdown("---"); st.markdown("####  Resultados")
        rz1,rz2,rz3 = st.columns(3)
        with rz1:
            st.metric(f" {mat_lbl} (físicas)",f"{n_phys_tiles} unidades")
            st.metric(" Cajas",f"{n_cajas_z:.2f} cajas")
            st.metric("Longitud",f"{long_z*cant_pi:.1f} m")
            st.metric("Tipo",tipo_z)
            st.metric(f"Zócalos (piezas)",f"{n_zocalo_pcs}")
            st.metric("Uds/Caja",str(uds_caja_z))
        with rz2:
            st.metric(" Adhesivo",f"{adh_bol_z:.2f} bultos")
            st.metric(" Agua adh.",f"{adh_agua_z:.2f} lt")
        with rz3:
            st.metric(" Boquilla",f"{boq_bol_z:.2f} bultos")
            st.metric(" Agua boq.",f"{boq_agua_z:.2f} lt")
            st.metric(" Costo est.",f"{moneda} {costo_z:,.0f}")

        mat_res = [
            {"elemento":f"{desc_pi or 'Zócalo'} — {mat_lbl}","unidad":"caja","cant":round(n_cajas_z,2),"precio":precio_caja_z},
            {"elemento":f"{desc_pi or 'Zócalo'} — Adhesivo","unidad":"bol","cant":adh_bol_z,"precio":float(p.get("adhesivo_piso",0))},
            {"elemento":f"{desc_pi or 'Zócalo'} — Boquilla","unidad":"bol","cant":boq_bol_z,"precio":float(p.get("boquilla",0))},
        ]

    # ─────── BOTÓN AGREGAR ───────
    st.markdown(f'<div style="background:#0d2137;border-radius:8px;padding:8px 16px;"><span style="color:#ffcc80;">{modo_pi} | {moneda}</span></div>',unsafe_allow_html=True)
    if st.button(" Agregar Piso al Resumen", key="kc_add_piso_1", type="primary"):
        st.session_state.kc_rows.extend(mat_res)
        st.success(f" {desc_pi or 'Piso'} agregado | {moneda}")


# ══════════ TAB 10 — CIELO RASO (CM-V3.0 · Panel Yeso) ══════════
with tabs[9]:
    st.subheader(" Calculadora de Cielo Raso")
    st.caption(f"CM-V3.0 | Panel Yeso / Drywall | {norma_sel} | {moneda}")

    # Element standard lengths (all in 2.44m = 8ft)
    ELM_LENS = {
        "2.44m (8')": 2.44,
        "3.00m":      3.00,
        "2.00m":      2.00,
        "4.00m":      4.00,
    }
    PANEL_TYPES = {
        "Panel 1.22×2.44m": (1.22, 2.44),
        "Panel 1.20×2.40m": (1.20, 2.40),
        "Panel 1.20×2.50m": (1.20, 2.50),
        "Panel 1.22×3.00m": (1.22, 3.00),
    }
    # Calibrated constants from CM-V3.0
    _TORNILLOS_EST_M2 = 10.0    # tornillos estructura / m²
    _TORNILLOS_PAN_M2 = 5.0     # tornillos panel / m²
    _MASILLA_M2       = 1/6.54  # cubetas / m²  (1 cubeta=~6.54m² coverage)

    cr1, cr2, cr3 = st.columns(3)
    with cr1:
        st.markdown("** Dimensiones**")
        ancho_cr = st.number_input("Ancho [m]",0.0,value=10.0,step=0.5,key="kc_cr_anc")
        largo_cr = st.number_input("Largo [m]",0.0,value=12.0,step=0.5,key="kc_cr_lar")
        cant_cr  = st.number_input("Cantidad (ambientes)",min_value=1,value=1,key="kc_cr_qty")
        desc_cr  = st.text_input("Descripción",placeholder="Ej: Cielo sala",key="kc_cr_desc")
        area_cr  = ancho_cr * largo_cr * cant_cr
        perim_cr = 2*(ancho_cr + largo_cr) * cant_cr
        st.metric(" Área",f"{area_cr:.2f} m²"); st.metric(" Perímetro",f"{perim_cr:.2f} m")

    with cr2:
        st.markdown("** Estructura**")
        sep_v    = st.number_input("Sep. Viguetas [m]",0.0,value=0.40,step=0.05,key="kc_cr_sv")
        len_v    = st.selectbox("Dimensión Viguetas:",list(ELM_LENS.keys()),key="kc_cr_lv"); vigueta_len=ELM_LENS[len_v]
        sep_o    = st.number_input("Sep. Omegas [m]",0.0,value=0.45,step=0.05,key="kc_cr_so")
        len_o    = st.selectbox("Dimensión Omegas:",list(ELM_LENS.keys()),key="kc_cr_lo"); omega_len=ELM_LENS[len_o]
        len_a    = st.selectbox("Dimensión Ángulo Perim.:",list(ELM_LENS.keys()),key="kc_cr_la"); angulo_len=ELM_LENS[len_a]
        st.markdown("** Panel**")
        panel_t  = st.selectbox("Tipo Panel:",list(PANEL_TYPES.keys()),key="kc_cr_pt")
        panel_W,panel_H = PANEL_TYPES[panel_t]
        st.caption(f"Panel: {panel_W}×{panel_H}m = {panel_W*panel_H:.3f} m²/unidad")

    with cr3:
        st.markdown("** Desperdicios**")
        desp_panel = st.select_slider("Panel [%]",[1,2,3,4,5,6,8],value=1,key="kc_cr_dp")
        desp_otros = st.select_slider("Otros [%]",[1,2,3,4,5,6,8],value=1,key="kc_cr_do")
        st.markdown("** Precios**")
        _pp = 25000.0 if "COP" in moneda else 8.0
        _pv = 8500.0  if "COP" in moneda else 2.80
        _po = 7500.0  if "COP" in moneda else 2.50
        _pa = 6000.0  if "COP" in moneda else 2.00
        _pm = 35000.0 if "COP" in moneda else 12.0
        _pt = 50.0    if "COP" in moneda else 0.02
        precio_panel = st.number_input(f"Panel [{moneda}/un]",0.0,value=_pp,step=100.0,key="kc_cr_ppanel")
        precio_vig   = st.number_input(f"Vigueta [{moneda}/un]",0.0,value=_pv,step=100.0,key="kc_cr_pvig")
        precio_omg   = st.number_input(f"Omega [{moneda}/un]",0.0,value=_po,step=100.0,key="kc_cr_pomg")
        precio_ang   = st.number_input(f"Ángulo [{moneda}/un]",0.0,value=_pa,step=100.0,key="kc_cr_pang")
        precio_mas   = st.number_input(f"Masilla [{moneda}/cb]",0.0,value=_pm,step=100.0,key="kc_cr_pmas")
        precio_torn  = st.number_input(f"Tornillo [{moneda}/un]",0.0,value=_pt,step=1.0,key="kc_cr_ptorn")

    # ======= DIAGRAMA TÉCNICO PROFESIONAL DE CIELO RASO =======
    _w_svg_cr, _h_svg_cr = 260, 200
    
    _defs_cr = """<defs>
      <pattern id="slab_cr" width="20" height="10" patternUnits="userSpaceOnUse">
        <rect width="20" height="10" fill="#1e3a5f" opacity="0.6"/>
        <path d="M0 10 L20 0" stroke="#4a9eff" stroke-width="0.5" opacity="0.4"/>
      </pattern>
      <pattern id="grid_pan" width="30" height="60" patternUnits="userSpaceOnUse">
        <rect width="30" height="60" fill="#2c3e50" opacity="0.3"/>
        <path d="M30 0 L30 60 L0 60" stroke="#fff" stroke-width="0.5" opacity="0.5" fill="none"/>
      </pattern>
    </defs>"""

    _svg_corte_cr = '<rect x="30" y="30" width="200" height="20" fill="url(#slab_cr)" stroke="#4a9eff" stroke-width="1.5"/><text x="130" y="44" font-size="10" fill="#ccc" text-anchor="middle">Placa / Losa Superior</text>'
    _svg_corte_cr += '<line x1="70" y1="50" x2="70" y2="100" stroke="#95a5a6" stroke-width="1.5"/><line x1="130" y1="50" x2="130" y2="100" stroke="#95a5a6" stroke-width="1.5"/><line x1="190" y1="50" x2="190" y2="100" stroke="#95a5a6" stroke-width="1.5"/><text x="175" y="80" font-size="8" fill="#95a5a6">Cuelgues</text>'
    _svg_corte_cr += '<rect x="65" y="100" width="10" height="12" fill="#bdc3c7" stroke="#7f8c8d" stroke-width="1"/><rect x="125" y="100" width="10" height="12" fill="#bdc3c7" stroke="#7f8c8d" stroke-width="1"/><rect x="185" y="100" width="10" height="12" fill="#bdc3c7" stroke="#7f8c8d" stroke-width="1"/><text x="130" y="93" font-size="8" fill="#bdc3c7" text-anchor="middle">Viguetas</text>'
    _svg_corte_cr += '<rect x="40" y="112" width="180" height="6" fill="#ecf0f1" stroke="#bdc3c7" stroke-width="0.5"/><text x="130" y="126" font-size="8" fill="#ecf0f1" text-anchor="middle">Omegas</text>'
    _svg_corte_cr += '<rect x="30" y="130" width="200" height="8" fill="#fff" stroke="#95a5a6" stroke-width="1"/><text x="130" y="137" font-size="8" fill="#2c3e50" text-anchor="middle">Panel Yeso / Drywall</text>'

    _svg_planta_cr = f'<rect x="40" y="40" width="180" height="120" fill="url(#grid_pan)" stroke="#4a9eff" stroke-width="2"/>'
    for i in range(1, 9): _svg_planta_cr += f'<line x1="{40+i*20}" y1="40" x2="{40+i*20}" y2="160" stroke="#bdc3c7" stroke-width="2" opacity="0.6"/>'
    for j in range(1, 8): _svg_planta_cr += f'<line x1="40" y1="{40+j*15}" x2="220" y2="{40+j*15}" stroke="#ecf0f1" stroke-width="1" opacity="0.8"/>'
    _svg_planta_cr += '<text x="130" y="105" font-size="12" fill="#fff" text-anchor="middle" font-weight="bold">Cuadrícula Estructura</text>'
    
    _html_cr_diag = f'''
    <div style="background:#0a192f; padding:16px; border-radius:8px; border:1px solid #1e3a5f; margin:12px 0 24px 0; display:flex; justify-content:space-around; align-items:center; flex-wrap:wrap; box-shadow:0 4px 15px rgba(0,0,0,0.3);">
      <div style="text-align:center; min-width:270px;"><h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Corte Transversal</h4><svg width="{_w_svg_cr}" height="{_h_svg_cr}" viewBox="0 0 {_w_svg_cr} {_h_svg_cr}" xmlns="http://www.w3.org/2000/svg">{_defs_cr}{_svg_corte_cr}</svg></div>
      <div style="text-align:center; min-width:270px; border-left:1px dashed #1e3a5f;"><h4 style="color:#90caf9; font-size:0.95rem; margin-bottom:10px;">Vista en Planta</h4><svg width="{_w_svg_cr}" height="{_h_svg_cr}" viewBox="0 0 {_w_svg_cr} {_h_svg_cr}" xmlns="http://www.w3.org/2000/svg">{_defs_cr}{_svg_planta_cr}</svg></div>
    </div>
    '''
    import streamlit.components.v1 as _stc_cr
    _stc_cr.html(_html_cr_diag, height=270, scrolling=False)

    # ── Calculations ──
    # Panels
    n_panels     = area_cr / (panel_W * panel_H)
    n_panels_desp = n_panels * (1+desp_panel/100)

    # Structural (viguetas run along LARGO direction, spaced along ANCHO)
    n_viguetas   = math.ceil((ancho_cr/sep_v) * largo_cr / vigueta_len * cant_cr)
    n_omegas     = math.ceil((largo_cr/sep_o) * ancho_cr / omega_len  * cant_cr)
    n_angulo     = math.ceil(perim_cr / angulo_len * (1+desp_otros/100))

    # Finishing
    n_masilla    = area_cr * _MASILLA_M2
    n_torn_est   = math.ceil(area_cr * _TORNILLOS_EST_M2 * (1+desp_otros/100))
    n_torn_pan   = math.ceil(area_cr * _TORNILLOS_PAN_M2 * (1+desp_otros/100))

    # Costs
    costo_cr = (
        n_panels_desp * precio_panel +
        n_viguetas    * precio_vig +
        n_omegas      * precio_omg +
        n_angulo      * precio_ang +
        n_masilla     * precio_mas +
        (n_torn_est + n_torn_pan) * precio_torn
    )

    st.markdown("---"); st.markdown("####  Resultados")
    rc1,rc2,rc3 = st.columns(3)
    with rc1:
        st.markdown("**Paneles y Estructura**")
        st.metric(" Paneles",f"{n_panels_desp:.2f} unidades")
        st.metric(" Viguetas",f"{n_viguetas} unidades")
        st.metric(" Omegas",f"{n_omegas} unidades")
        st.metric(" Ángulo Perimetral",f"{n_angulo} unidades")
    with rc2:
        st.markdown("**Acabados**")
        st.metric(" Masilla",f"{n_masilla:.2f} cubetas")
        st.metric(" Tornillos Estructura",f"{n_torn_est}")
        st.metric(" Tornillos Paneles",f"{n_torn_pan}")
    with rc3:
        st.markdown("**Dimensiones**")
        st.metric(" Área",f"{area_cr:.2f} m²")
        st.metric(" Perímetro",f"{perim_cr:.2f} m")
        st.metric(" Costo total",f"{moneda} {costo_cr:,.0f}")

    st.markdown(
        f'<div style="background:#0d2137;border-radius:8px;padding:8px 16px;">'
        f'<span style="color:#ffcc80;">Panel Yeso | {panel_t} | {area_cr:.1f}m² | {moneda} {costo_cr:,.0f}</span></div>',
        unsafe_allow_html=True)
    if st.button(" Agregar Cielo Raso al Resumen", key="kc_add_cr", type="primary"):
        lb = desc_cr or "Cielo Raso"
        st.session_state.kc_rows.extend([
            {"elemento":f"{lb} — Paneles ({panel_t})","unidad":"unidad","cant":round(n_panels_desp,2),"precio":precio_panel},
            {"elemento":f"{lb} — Viguetas","unidad":"unidad","cant":n_viguetas,"precio":precio_vig},
            {"elemento":f"{lb} — Omegas","unidad":"unidad","cant":n_omegas,"precio":precio_omg},
            {"elemento":f"{lb} — Ángulo Perimetral","unidad":"unidad","cant":n_angulo,"precio":precio_ang},
            {"elemento":f"{lb} — Masilla","unidad":"cubeta","cant":round(n_masilla,2),"precio":precio_mas},
            {"elemento":f"{lb} — Tornillos Estruct.","unidad":"unidad","cant":n_torn_est,"precio":precio_torn},
            {"elemento":f"{lb} — Tornillos Panel","unidad":"unidad","cant":n_torn_pan,"precio":precio_torn},
        ])
        save_state()
        st.success(f" {lb}: {n_panels_desp:.1f} paneles | {n_viguetas}+{n_omegas} estruct. | {moneda} {costo_cr:,.0f}")


# ══════════ TAB 3 — VARILLAS ══════════
with tabs[10]:
    st.subheader(f" Calculadora de {R['varilla']}"); st.caption(f"Base de datos CM-V3.0 | {norma_sel}")
    v1,v2,v3,v4=st.columns(4)
    with v1:
        var_nombres=[v["nombre"] for v in VARILLAS]
        var_sel=st.selectbox(f"Diámetro:",var_nombres,index=2,key="kc_var_sel")
        vkd=next(v for v in VARILLAS if v["nombre"]==var_sel)
    with v2:
        longitud_var=st.number_input("Longitud total a colocar [m]",min_value=0.1,value=50.0,step=5.0,key="kc_var_long")
        traslape_pct=st.slider("Factor traslape [%]",0,40,10,key="kc_var_traslape")/100
    with v3:
        bars_by_length=st.radio("Long. barra en obra:",["6 m","9 m","12 m"],index=0,key="kc_var_bar_len")
        bar_len=float(bars_by_length.split()[0])
    with v4:
        st.metric("⚖ Peso por metro",f"{vkd['kg_m']:.3f} kg/m")
        st.metric(" Peso por barra",f"{vkd['kg_m']*bar_len:.2f} kg/{int(bar_len)}m")
    
    # ======= DIAGRAMA TÉCNICO PROFESIONAL DE VARILLA =======
    _w_svg_v, _h_svg_v = 500, 160
    _var_rad = max(4, min(14, vkd.get("diam_mm", 12) / 2)) # Visual scaling
    _svg_v = f'<rect x="50" y="{80-_var_rad}" width="400" height="{_var_rad*2}" fill="#7f8c8d" stroke="#34495e" stroke-width="1"/>'
    for i in range(50, 450, 10):
        _svg_v += f'<path d="M {i} {80-_var_rad} Q {i+5} {80} {i} {80+_var_rad}" stroke="#95a5a6" stroke-width="2" fill="none" opacity="0.8"/>'
    _svg_v += f'<text x="250" y="{80-_var_rad-15}" font-size="14" fill="#bdc3c7" text-anchor="middle" font-weight="bold">{var_sel} ({vkd.get("diam_mm", "?")} mm)</text>'
    _svg_v += f'<line x1="50" y1="{80+_var_rad+20}" x2="450" y2="{80+_var_rad+20}" stroke="#4a9eff" stroke-width="1.5"/><line x1="50" y1="{80+_var_rad+15}" x2="50" y2="{80+_var_rad+25}" stroke="#4a9eff" stroke-width="1.5"/><line x1="450" y1="{80+_var_rad+15}" x2="450" y2="{80+_var_rad+25}" stroke="#4a9eff" stroke-width="1.5"/><text x="250" y="{80+_var_rad+35}" font-size="12" fill="#4a9eff" text-anchor="middle">Long. comercial: {int(bar_len)}m | Traslape: {traslape_pct*100}%</text>'
    
    _html_var = f'<div style="background:#0a192f; padding:16px; border-radius:8px; border:1px solid #1e3a5f; margin:12px 0; text-align:center;"><svg width="{_w_svg_v}" height="{_h_svg_v}" viewBox="0 0 {_w_svg_v} {_h_svg_v}" xmlns="http://www.w3.org/2000/svg">{_svg_v}</svg></div>'
    import streamlit.components.v1 as _stc_var
    _stc_var.html(_html_var, height=210, scrolling=False)
    long_total=longitud_var*(1+traslape_pct); kg_total=long_total*vkd["kg_m"]
    barras_total=math.ceil(long_total/bar_len); costo_acero=kg_total*p["acero_kg"]
    st.markdown("---"); va1,va2,va3,va4,va5=st.columns(5)
    va1.metric(" Longitud + traslape",f"{long_total:.2f} m"); va2.metric(f" Barras {int(bar_len)}m",f"{barras_total}")
    va3.metric("⚖ Peso total",f"{kg_total:.2f} kg"); va4.metric("⚖ Toneladas",f"{kg_total/1000:.3f} ton")
    va5.metric(" Costo est.",f"{moneda} {costo_acero:,.0f}")
    with st.expander(" Tabla completa de varillas"):
        df_var=pd.DataFrame(VARILLAS)
        df_var["Precio/barra 6m"]=df_var["kg_6m"].apply(lambda x: f"{moneda} {x*p['acero_kg']:,.0f}")
        st.dataframe(df_var.rename(columns={"nombre":"Varilla","diam_mm":"Ø mm","diam_pulg":"Ø pulg","kg_6m":"kg/6m","kg_m":"kg/m"}),use_container_width=True)
    if st.button(f" Agregar {R['varilla']} al Resumen",key="kc_add_var",type="primary"):
        st.session_state.kc_rows.append({"elemento":f"{R['varilla']} {var_sel}","unidad":"kg","cant":round(kg_total,2),"precio":p["acero_kg"]})
        st.success(f" {kg_total:.2f} kg de {var_sel} agregados")

# ══════════ TAB 4 — PINTURA ══════════
with tabs[11]:
    st.subheader(f" {_t('Calculadora de', 'Calculator for')} {R['pintura']}")
    p1,p2,p3=st.columns(3)
    with p1:
        lbl_area = f"{_t('Área', 'Area')} [{_u('m²','sq ft')}]"
        area_pin=st.number_input(lbl_area,min_value=1.0,value=50.0 if st.session_state["unidades"]=="Métrico" else 500.0,step=5.0,key="kc_area_pin")
        if "kc_manos_pin" not in st.session_state:
            st.session_state["kc_manos_pin"] = 2
        st.radio(_t("Número de manos:", "Number of coats:"), [1, 2, 3],
                 horizontal=True, key="kc_manos_pin")
        manos = st.session_state["kc_manos_pin"]
    with p2:
        lbl_rend = f"{_t('Rendimiento', 'Yield')} [{_u('m²/galón','sq ft/gal')}]"
        rend_gal=st.number_input(lbl_rend,min_value=4.0,max_value=500.0,value=float(R["rendimiento_pintura"]),step=0.5,key="kc_rend_pin")
        desp_pin=st.slider(_t("Desperdicio [%]", "Waste [%]"),0,15,5,key="kc_desp_pin")/100
    with p3: 
        tipos_pin = [_t("Interior vinílico", "Interior Latex"), _t("Exterior acrílico", "Exterior Acrylic"), _t("Epóxico", "Epoxy"), _t("Impermeabilizante", "Waterproofing")]
        tipo_pin=st.selectbox(_t("Tipo:", "Type:"),tipos_pin,key="kc_tipo_pin")
    
    # ── CÁLCULO (antes del iframe para evitar lag por rerender) ──
    galones_exactos = (area_pin * manos / rend_gal) * (1 + desp_pin)
    cunetes = math.floor(galones_exactos / 5)
    resto_5 = galones_exactos - (cunetes * 5)
    medios_cunetes = math.floor(resto_5 / 2.5)
    resto_2_5 = resto_5 - (medios_cunetes * 2.5)
    galones_sueltos = math.ceil(resto_2_5)
    galones_comerciales = (cunetes * 5) + (medios_cunetes * 2.5) + galones_sueltos
    costo_pin = galones_comerciales * p["pintura"]

    # ======= DIAGRAMA TÉCNICO PROFESIONAL DE PINTURA =======
    _w_svg_pt, _h_svg_pt = 500, 180
    _color_p = "#3498db" if "vinílico" in tipo_pin.lower() or "latex" in tipo_pin.lower() else ("#2ecc71" if "acrílico" in tipo_pin.lower() or "acrylic" in tipo_pin.lower() else ("#e74c3c" if "epóxi" in tipo_pin.lower() or "epoxy" in tipo_pin.lower() else "#f1c40f"))
    _svg_pt = f'<rect x="80" y="40" width="340" height="100" fill="#2c3e50" stroke="#1e3a5f" stroke-width="2"/>'
    _svg_pt += f'<rect x="80" y="40" width="220" height="100" fill="{_color_p}" opacity="0.8"/>'
    _svg_pt += f'<path d="M 300 40 Q 280 90 300 140" stroke="none" fill="{_color_p}" opacity="0.8"/>'
    _svg_pt += f'<rect x="250" y="60" width="16" height="60" rx="8" fill="#ecf0f1" stroke="#bdc3c7" stroke-width="1" transform="rotate(15 258 90)"/>'
    _svg_pt += f'<rect x="258" y="70" width="4" height="40" fill="#7f8c8d" transform="rotate(15 258 90)"/>'
    _svg_pt += f'<rect x="270" y="70" width="20" height="40" rx="4" fill="{_color_p}" transform="rotate(15 258 90)"/>'
    _svg_pt += f'<text x="250" y="30" font-size="14" fill="#bdc3c7" text-anchor="middle" font-weight="bold">{_t("Aplicación de", "Application of")} {tipo_pin} ({manos} {_t("manos", "coats")})</text>'
    _svg_pt += f'<text x="190" y="95" font-size="14" fill="#fff" text-anchor="middle" font-weight="bold">{_t("Área", "Area")}: {area_pin} {_u("m²","sq ft")}</text>'
    _html_pt = f'<div style="background:#0a192f; padding:16px; border-radius:8px; border:1px solid #1e3a5f; margin:12px 0; text-align:center;"><svg width="{_w_svg_pt}" height="{_h_svg_pt}" viewBox="0 0 {_w_svg_pt} {_h_svg_pt}" xmlns="http://www.w3.org/2000/svg">{_svg_pt}</svg></div>'
    import streamlit.components.v1 as _stc_pt
    _stc_pt.html(_html_pt, height=220, scrolling=False)

    st.markdown("---"); pp1,pp2,pp3,pp4=st.columns(4)
    pp1.metric(f" {R['pintura']}",f"{galones_comerciales:g} {_t('galones', 'gallons')}",f"{manos} {_t('manos', 'coats')}")
    
    # Texto para el empaque
    text_empaque = f"{cunetes} {_t('Cuñ. (5g)', 'Buck. (5g)')}" if cunetes > 0 else ""
    if medios_cunetes > 0: text_empaque += f" + {medios_cunetes} {_t('Med.Cuñ. (2.5g)', 'Half.Buck')}"
    if galones_sueltos > 0: text_empaque += f" + {galones_sueltos} {_t('gal.', 'gal.')}"
    if not text_empaque: text_empaque = f"0"
        
    pp2.metric(f" {_t('Presentación', 'Packaging')}", text_empaque.strip(" + "))
    
    pp3.metric(f" {_t('Área', 'Area')}",f"{area_pin:.1f} {_u('m²','sq ft')}")
    pp4.metric(f" {_t('Costo est.', 'Est. Cost')}",f"{moneda} {costo_pin:,.0f}")
    
    if st.button(f" {_t('Agregar Pintura al Resumen', 'Add Paint to Summary')}",key="kc_add_pin",type="primary"):
        st.session_state.kc_rows.append({"elemento":f"{R['pintura']} {tipo_pin}","unidad":_t("galón", "gallon"),"cant":galones_comerciales,"precio":p["pintura"]})
        st.success(f" {galones_comerciales:g} {_t('galones de', 'gallons of')} {tipo_pin} {_t('agregados', 'added')}")


# ══════════ TAB 6 — CUBIERTA ══════════
with tabs[12]:
    st.subheader(f" Calculadora de {R['cubierta']}")
    CUBIERTA_TIPOS={"Lámina ZincAlum (0.80m útil)":{"ancho_util":0.80,"tipo":"lamina"},"Lámina Galvanizada (0.80m útil)":{"ancho_util":0.80,"tipo":"lamina"},
                    "Teja española (5 und/m²)":{"und_m2":5,"tipo":"teja"},"Teja plana (6 und/m²)":{"und_m2":6,"tipo":"teja"},"Teja ondulada (7 und/m²)":{"und_m2":7,"tipo":"teja"}}
    cu1,cu2=st.columns(2)
    with cu1: cub_tipo=st.selectbox("Tipo:",list(CUBIERTA_TIPOS.keys()),key="kc_cub_tipo"); ckd=CUBIERTA_TIPOS[cub_tipo]
    with cu2: area_cub=st.number_input("Área [m²]",min_value=1.0,value=30.0,step=5.0,key="kc_area_cub"); desp_cub=st.slider("Desperdicio [%]",0,20,10,key="kc_desp_cub")/100
    area_t_cub=area_cub*(1+desp_cub)
    
    # ======= DIAGRAMA TÉCNICO PROFESIONAL DE CUBIERTA =======
    _w_svg_c, _h_svg_c = 500, 160
    _svg_c = ""
    if "lamina" in ckd["tipo"]:
        _svg_c += '<rect x="100" y="60" width="300" height="60" fill="#2c3e50" stroke="#1abc9c" stroke-width="2"/>'
        for i in range(120, 400, 20):
            _svg_c += f'<line x1="{i}" y1="60" x2="{i}" y2="120" stroke="#1abc9c" stroke-width="1" opacity="0.5"/>'
        _svg_c += f'<text x="250" y="50" font-size="14" fill="#bdc3c7" text-anchor="middle" font-weight="bold">Láminas ({cub_tipo})</text>'
    else:
        _svg_c += '<rect x="100" y="60" width="300" height="60" fill="#c0392b" stroke="#e67e22" stroke-width="2"/>'
        for i in range(120, 400, 15):
            _svg_c += f'<path d="M {i} 60 Q {i+5} 90 {i} 120" stroke="#e67e22" stroke-width="1" fill="none" opacity="0.7"/>'
        _svg_c += f'<text x="250" y="50" font-size="14" fill="#bdc3c7" text-anchor="middle" font-weight="bold">Tejado ({cub_tipo})</text>'
    _svg_c += f'<text x="250" y="100" font-size="14" fill="#fff" text-anchor="middle" font-weight="bold">Área Total: {area_t_cub:.1f} m²</text>'
    
    _html_c = f'<div style="background:#0a192f; padding:16px; border-radius:8px; border:1px solid #1e3a5f; margin:12px 0; text-align:center;"><svg width="{_w_svg_c}" height="{_h_svg_c}" viewBox="0 0 {_w_svg_c} {_h_svg_c}" xmlns="http://www.w3.org/2000/svg">{_svg_c}</svg></div>'
    import streamlit.components.v1 as _stc_c
    _stc_c.html(_html_c, height=210, scrolling=False)
    if ckd["tipo"]=="lamina":
        laminas=math.ceil(area_t_cub/ckd["ancho_util"]); st.metric(" Láminas",f"{laminas} und")
        if st.button(" Agregar Cubierta al Resumen",key="kc_add_cub",type="primary"):
            st.session_state.kc_rows.append({"elemento":f"Cubierta — {cub_tipo}","unidad":"und","cant":laminas,"precio":p.get("lamina",p["ceramica"]*5)}); st.success(f" {laminas} láminas agregadas")
    else:
        und_cub=math.ceil(area_t_cub*ckd["und_m2"]); st.metric(" Tejas",f"{und_cub:,} und")
        if st.button(" Agregar Cubierta al Resumen",key="kc_add_cub2",type="primary"):
            st.session_state.kc_rows.append({"elemento":f"Cubierta — {cub_tipo}","unidad":"und","cant":und_cub,"precio":p.get("teja",p["ceramica"]*0.3)}); st.success(f" {und_cub} tejas agregadas")

# ══════════ TAB 7 — IMPORTAR EXCEL ══════════
with tabs[13]:
    st.subheader(" Importar desde Excel")
    template_rows=[
        {"Elemento":"Columna C1","Tipo":"Concreto","Volumen_m3":0.5,"Area_m2":"","Longitud_m":"","fc_MPa":21,"Dosificacion":"1:2:4"},
        {"Elemento":"Viga V1","Tipo":"Concreto","Volumen_m3":0.8,"Area_m2":"","Longitud_m":"","fc_MPa":21,"Dosificacion":"1:2:4"},
        {"Elemento":"Muro eje A","Tipo":"Mamposteria","Volumen_m3":"","Area_m2":15,"Longitud_m":"","fc_MPa":"","Dosificacion":"Bloque 15x20x40 cm"},
        {"Elemento":"Varilla N4","Tipo":"Acero","Volumen_m3":"","Area_m2":"","Longitud_m":120,"fc_MPa":"","Dosificacion":"N4 - 1/2\""},
    ]
    tmpl_buf=BytesIO(); pd.DataFrame(template_rows).to_excel(tmpl_buf,index=False,engine="xlsxwriter"); tmpl_buf.seek(0)
    st.download_button(" Descargar Plantilla Excel",tmpl_buf,file_name="Konte_Plantilla.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    up_file=st.file_uploader(" Subir Excel:",type=["xlsx","xls"],key="kc_upload")
    if up_file:
        try:
            df_up=pd.read_excel(up_file); st.dataframe(df_up,use_container_width=True); new_rows=[]
            for _,row in df_up.iterrows():
                tipo=str(row.get("Tipo","")).strip().lower(); elem=str(row.get("Elemento","sin nombre"))
                if "concreto" in tipo or "hormigon" in tipo:
                    vol=float(row.get("Volumen_m3",1) or 1); dos_key=str(row.get("Dosificacion","1:2:4")).strip()
                    mix_i=next((m for m in MIX_DESIGNS if m["dos"]==dos_key),MIX_DESIGNS[4])
                    bolsas_i=math.ceil(mix_i["cem_kg"]*vol/R["peso_bolsa"])
                    new_rows.extend([{"elemento":f"{elem} — {R['cemento']}","unidad":"bultos","cant":bolsas_i,"precio":p["cemento"]},
                                      {"elemento":f"{elem} — Arena","unidad":"m³","cant":round(mix_i["arena_m3"]*vol,3),"precio":p["arena"]},
                                      {"elemento":f"{elem} — Grava","unidad":"m³","cant":round(mix_i["grava_m3"]*vol,3),"precio":p["grava"]}])
                elif "acero" in tipo or "varilla" in tipo or "fierro" in tipo:
                    long_i=float(row.get("Longitud_m",0) or 0); dos_i=str(row.get("Dosificacion","N4 - 1/2\""))
                    var_i=next((v for v in VARILLAS if dos_i in v["nombre"]),VARILLAS[2])
                    new_rows.append({"elemento":f"{elem} — {dos_i}","unidad":"kg","cant":round(long_i*var_i["kg_m"],2),"precio":p["acero_kg"]})
                elif "mamposter" in tipo or "muro" in tipo:
                    area_i=float(row.get("Area_m2",0) or 0)
                    new_rows.append({"elemento":elem,"unidad":"m²","cant":area_i,"precio":p["bloque"]*12.5})
            if new_rows: st.session_state.kc_rows.extend(new_rows); st.success(f" {len(new_rows)} materiales importados"); st.dataframe(pd.DataFrame(new_rows),use_container_width=True)
        except Exception as e: st.error(f" Error: {e}")

# ══════════ TAB RESUMEN — RESUMEN Y EXPORTAR ══════════

# ══════════ TAB 16 — RENDIMIENTOS MO ══════════
with tabs[16]:
    st.subheader("♂ Factor de Rendimiento (Mano de Obra)")
    st.caption("Consulte tiempos requeridos (días) según volumen de obra, basados en 'Rendimientos.xlsx'.")
    search_rend = st.text_input(" Buscar actividad:")
    
    df_rend = pd.DataFrame(RENDIMIENTOS_MO)
    if search_rend:
        df_rend = df_rend[df_rend['Actividad'].str.contains(search_rend, case=False, na=False)]
    
    st.dataframe(df_rend, use_container_width=True, hide_index=True)
    
    st.markdown("---")
    st.markdown("#### ⏱ Calculadora Rápida de Tiempos")
    calc_col1, calc_col2 = st.columns([3, 1])
    with calc_col1:
        actividad_sel = st.selectbox("Seleccione la Actividad:", df_rend['Actividad'].tolist())
    with calc_col2:
        qty_val = st.number_input("Cantidad de Obra", min_value=0.0, value=1.0, step=1.0)
    
    if actividad_sel:
        rend = df_rend[df_rend['Actividad'] == actividad_sel].iloc[0]
        factor = float(rend['Cantidad'])
        ud = rend['Unidad']
        # Formulas from Excel: If factor > 0, days = qty_val / factor
        dias_req = qty_val / factor if factor > 0 else 0
        st.success(f"**Resultado:** Para {qty_val:,.2f} unidades, se requieren **{dias_req:,.2f} días** (Rendimiento: {factor} {ud})")

# ══════════ TAB 17 — PRESUPUESTO ══════════
with tabs[17]:
    st.subheader(" Generador de Presupuestos APU")
    st.caption("Presupuesto preconfigurado con precios unitarios, mano de obra, IVA y AIU según la región seleccionada.")

    # ─ Parametros regionales ─────────────────────────────────────────────────
    APU_CONFIG = {
        "Colombia": {
            "smmlv_mes": 1423500, "aux_trans": 162000, "iva": 0.19, "aiu": 0.20,
            "moneda": "COP $", "factor_ps": 1.60,
        },
        "Ecuador":  {"smmlv_mes": 460, "aux_trans": 0, "iva": 0.15, "aiu": 0.18, "moneda": "USD $", "factor_ps": 1.35},
        "Perú":     {"smmlv_mes": 1025, "aux_trans": 102.5, "iva": 0.18, "aiu": 0.20, "moneda": "PEN S/", "factor_ps": 1.40},
        "México":   {"smmlv_mes": 7310, "aux_trans": 0, "iva": 0.16, "aiu": 0.15, "moneda": "MXN $", "factor_ps": 1.30},
        "Bolivia":  {"smmlv_mes": 2362, "aux_trans": 0, "iva": 0.13, "aiu": 0.18, "moneda": "BOB Bs.", "factor_ps": 1.25},
        "Argentina":{"smmlv_mes": 780000,"aux_trans": 0, "iva": 0.21,"aiu": 0.25, "moneda":"ARS $", "factor_ps": 1.45},
        "Venezuela":{"smmlv_mes": 130, "aux_trans": 40, "iva": 0.16, "aiu": 0.20, "moneda": "USD $", "factor_ps": 1.35},
        "USA":      {"smmlv_mes": 7.25*8*5*4, "aux_trans": 0, "iva": 0.0, "aiu": 0.20, "moneda": "USD $", "factor_ps": 1.25},
    }

    cfg_apu = APU_CONFIG.get(pais, APU_CONFIG["Colombia"])
    _smmlv_total = cfg_apu["smmlv_mes"] + cfg_apu.get("aux_trans", 0)
    jd = (_smmlv_total / 30) * cfg_apu["factor_ps"]
    _iva = cfg_apu["iva"]
    _aiu = cfg_apu["aiu"]

    # ─ Terminologia regional ──────────────────────────────────────────────────
    term = {
        "Colombia":  {"exc":"Excavación manual","cim":"Cimiento corrido","col":"Columna","viga":"Viga","losa":"Placa maciza","muro":"Muro de mampostería","piso":"Piso en cerámica","fachada":"Pecto / Estuco","pint":"Pintura de caucho","inst_hid":"Instalaciones hidráulicas","inst_el":"Instalaciones eléctricas","cubierta":"Cubierta en teja","adm":"Administración, Imprevistos y Utilidad (AIU)"},
        "Ecuador":   {"exc":"Excavación a mano","cim":"Cimiento corrido H°S","col":"Columna de hormigón","viga":"Cadena / Viga","losa":"Losa plana","muro":"Muro de bloque","piso":"Piso en cerámica","fachada":"Revoque / Enlucido","pint":"Pintura látex","inst_hid":"Instalaciones sanitarias","inst_el":"Instalaciones eléctricas","cubierta":"Cubierta en zinc","adm":"Administración e Imprevistos (AIU)"},
        "Perú":      {"exc":"Excavación a pulso","cim":"Cimiento corrido","col":"Columna de concreto","viga":"Viga principal","losa":"Losa aligerada","muro":"Muro de ladrillo King Kong","piso":"Piso en cerámico","fachada":"Tarrajeo / Revoque","pint":"Pintura látex","inst_hid":"Instalaciones sanitarias","inst_el":"Instalaciones eléctricas","cubierta":"Cobertura en calamina","adm":"Gastos generales y utilidad"},
        "México":    {"exc":"Excavación manual","cim":"Cimentación corrida","col":"Columna","viga":"Trabe / Viga","losa":"Losa de concreto","muro":"Muro de block","piso":"Piso en azulejo/cerámica","fachada":"Aplanado / Firme","pint":"Pintura vinilicá","inst_hid":"Instalación hidráulico-sanitaria","inst_el":"Instalaciones eléctricas","cubierta":"Lámina galvanizada","adm":"Gastos generales y utilidad"},
        "Bolivia":   {"exc":"Excavación a mano","cim":"Cimiento de hormigón","col":"Columna","viga":"Viga","losa":"Losa de hormigón","muro":"Muro de ladrillo","piso":"Piso en cerámica","fachada":"Revoque / Enlucido","pint":"Pintura látex","inst_hid":"Plomería","inst_el":"Electricidad","cubierta":"Cubierta","adm":"Gastos admin. y utilidad"},
        "Argentina": {"exc":"Excavación manual","cim":"Cimiento / Platea","col":"Columna","viga":"Viga / Encadenado","losa":"Losa de hormigón","muro":"Muro de ladrillo cerámico","piso":"Piso en cerámica","fachada":"Revoque / Jaharro","pint":"Pintura látex","inst_hid":"Plomería","inst_el":"Electricidad","cubierta":"Chapa metálica","adm":"Indirectos y utilidad (AIU)"},
        "Venezuela": {"exc":"Excavación manual","cim":"Cimiento corrido","col":"Columna","viga":"Viga","losa":"Losa","muro":"Muro de ladrillo","piso":"Piso en cerámica","fachada":"Friso / Revoque","pint":"Pintura de caucho","inst_hid":"Plomería","inst_el":"Electricidad","cubierta":"Lámina acanalada","adm":"Overheads y utilidad"},
        "USA":       {"exc":"Manual Excavation","cim":"Strip Footing","col":"Concrete Column","viga":"Concrete Beam","losa":"Concrete Slab","muro":"CMU Block Wall","piso":"Ceramic Tile Floor","fachada":"Stucco / Plaster","pint":"Latex Paint","inst_hid":"Plumbing rough-in","inst_el":"Electrical rough-in","cubierta":"Metal Roof","adm":"Overheads and Profit (O&P)"},
    }
    T = term.get(pais, term["Colombia"])

    # ─ Precios unitarios base ──────────────────────────
    _cem  = float(p.get("cemento", 34000 if _iva else 30))
    _are  = float(p.get("arena", 80000 if _iva else 50))
    _gra  = float(p.get("grava", 90000 if _iva else 60))
    _ace  = float(p.get("acero_kg", 4800 if _iva else 3))
    _blo  = float(p.get("bloque", 1600 if _iva else 2.5))
    _pin  = float(p.get("pintura", 55000 if _iva else 40))
    _cer  = float(p.get("ceramica", 38000 if _iva else 12))
    _tb   = float(p.get("teja_barro", 1800 if _iva else 0.3))

    def pu(mat, mo_dias, aiu_pct=_aiu, iva_pct=_iva):
        base = mat + (mo_dias * jd)
        return round(base * (1 + aiu_pct) * (1 + iva_pct), 0)

    # ─ Tabla APU preconfigurada ───────────────────────────────────────────────
    APU_ROWS = [
        {"N°":"1", "DESCRIPCIÓN":"CAPITULO 1 — PRELIMINARES", "UDS":"", "CANTIDAD":0, "PU":0},
        {"N°":"1.1","DESCRIPCIÓN":"Limpieza y demarcación del terreno","UDS":"m²","CANTIDAD":1.0, "PU": pu(0, 0.02)},
        {"N°":"1.2","DESCRIPCIÓN":T["exc"] + " en terreno normal h<1.5m","UDS":"m³","CANTIDAD":1.0, "PU": pu(0, 1.2)},
        {"N°":"1.3","DESCRIPCIÓN":"Retiro de material sobrante","UDS":"m³","CANTIDAD":1.0, "PU": pu(jd*0.5, 0.3)},
        {"N°":"2", "DESCRIPCIÓN":"CAPITULO 2 — CIMENTACIÓN", "UDS":"", "CANTIDAD":0, "PU":0},
        {"N°":"2.1","DESCRIPCIÓN":T["cim"] + " en concreto 21 MPa","UDS":"m³","CANTIDAD":1.0, "PU": pu(_cem*7+_are*0.48+_gra*0.95+_ace*80, 1.5)},
        {"N°":"2.2","DESCRIPCIÓN":"Solado/plantilla de limpieza e=5cm","UDS":"m²","CANTIDAD":1.0, "PU": pu(_cem*0.8+_are*0.05, 0.08)},
        {"N°":"2.3","DESCRIPCIÓN":"Suministro e instalación acero de refuerzo","UDS":"kg", "CANTIDAD":1.0, "PU": pu(_ace, 0.02)},
        {"N°":"3", "DESCRIPCIÓN":"CAPITULO 3 — ESTRUCTURA", "UDS":"", "CANTIDAD":0, "PU":0},
        {"N°":"3.1","DESCRIPCIÓN":T["col"] + " sencilla 0.25x0.25m — concreto 21 MPa","UDS":"ml", "CANTIDAD":1.0, "PU": pu(_cem*1.6+_are*0.12+_gra*0.24+_ace*15, 0.25)},
        {"N°":"3.2","DESCRIPCIÓN":T["viga"] + " rectangular — concreto 21 MPa","UDS":"ml", "CANTIDAD":1.0, "PU": pu(_cem*1.2+_are*0.09+_gra*0.18+_ace*12, 0.20)},
        {"N°":"3.3","DESCRIPCIÓN":T["losa"] + " e=12cm — concreto 21 MPa","UDS":"m²","CANTIDAD":1.0, "PU": pu(_cem*2.5+_are*0.09+_gra*0.18+_ace*9, 0.35)},
        {"N°":"4", "DESCRIPCIÓN":"CAPITULO 4 — MAMPOSTERÍA", "UDS":"", "CANTIDAD":0, "PU":0},
        {"N°":"4.1","DESCRIPCIÓN":T["muro"] + " bloque 15cm — mortero 1:4","UDS":"m²","CANTIDAD":1.0, "PU": pu(_blo*12.5+_cem*0.9+_are*0.055, 0.20)},
        {"N°":"4.2","DESCRIPCIÓN":"Ladrillo frente tipo tolete (58 und/m²)","UDS":"m²","CANTIDAD":1.0, "PU": pu(_blo*2.8*58+_cem*0.7+_are*0.04, 0.22)},
        {"N°":"4.3","DESCRIPCIÓN":T["fachada"] + " interior e=15mm — mortero 1:3","UDS":"m²","CANTIDAD":1.0, "PU": pu(_cem*0.55+_are*0.04, 0.12)},
        {"N°":"5", "DESCRIPCIÓN":"CAPITULO 5 — CUBIERTAS", "UDS":"", "CANTIDAD":0, "PU":0},
        {"N°":"5.1","DESCRIPCIÓN":T["cubierta"] + " — teja barro media canal (6 und/m²)","UDS":"m²","CANTIDAD":1.0, "PU": pu(_tb*6, 0.12)},
        {"N°":"5.2","DESCRIPCIÓN":"Cubierta zinc ondulado cal. 26 + estructura","UDS":"m²","CANTIDAD":1.0, "PU": pu(18000 if _iva else 6.5, 0.10)},
        {"N°":"6", "DESCRIPCIÓN":"CAPITULO 6 — PISOS Y ENCHAPES", "UDS":"", "CANTIDAD":0, "PU":0},
        {"N°":"6.1","DESCRIPCIÓN":T["piso"] + " 30x30 cm (Tipo CE1)","UDS":"m²","CANTIDAD":1.0, "PU": pu(_cer*1.1+_cem*0.4+_are*0.03, 0.12)},
        {"N°":"6.2","DESCRIPCIÓN":"Porcelanato rectificado 60x60 (PN1)","UDS":"m²","CANTIDAD":1.0, "PU": pu((55000 if _iva else 16)*1.1+_cem*0.5, 0.14)},
        {"N°":"6.3","DESCRIPCIÓN":"Mortero de nivelación e=3cm (1:4)","UDS":"m²","CANTIDAD":1.0, "PU": pu(_cem*0.55+_are*0.035, 0.08)},
        {"N°":"7", "DESCRIPCIÓN":"CAPITULO 7 — PINTURA Y ACABADOS", "UDS":"", "CANTIDAD":0, "PU":0},
        {"N°":"7.1","DESCRIPCIÓN":T["pint"] + " interior — 2 manos","UDS":"m²","CANTIDAD":1.0, "PU": pu(_pin/10, 0.04)},
        {"N°":"7.2","DESCRIPCIÓN":"Pintura exterior impresión + 2 manos","UDS":"m²","CANTIDAD":1.0, "PU": pu(_pin/8, 0.05)},
        {"N°":"7.3","DESCRIPCIÓN":"Estuco o afin (pared)—listo para pintar","UDS":"m²","CANTIDAD":1.0, "PU": pu(9400 if _iva else 3.0, 0.08)},
        {"N°":"8", "DESCRIPCIÓN":"CAPITULO 8 — INSTALACIONES", "UDS":"", "CANTIDAD":0, "PU":0},
        {"N°":"8.1","DESCRIPCIÓN":T["inst_hid"] + " punto hidráulico","UDS":"pto","CANTIDAD":1.0, "PU": pu(80000 if _iva else 25, 0.3)},
        {"N°":"8.2","DESCRIPCIÓN":T["inst_hid"] + " punto sanitario","UDS":"pto","CANTIDAD":1.0, "PU": pu(60000 if _iva else 20, 0.25)},
        {"N°":"8.3","DESCRIPCIÓN":T["inst_el"] + " punto toma corriente GFCI","UDS":"pto","CANTIDAD":1.0, "PU": pu(35000 if _iva else 12, 0.15)},
        {"N°":"8.4","DESCRIPCIÓN":T["inst_el"] + " punto iluminación + interruptor","UDS":"pto","CANTIDAD":1.0, "PU": pu(28000 if _iva else 10, 0.12)},
        {"N°":"9", "DESCRIPCIÓN":f'CAPITULO 9 — {T["adm"].upper()} ({_aiu*100:.0f}%)', "UDS":"%","CANTIDAD":_aiu*100, "PU":0},
    ]

    # ─ Init session ────────────────────────────────────────────
    if "kc_presupuesto" not in st.session_state or st.session_state.get("kc_pres_pais","") != pais:
        st.session_state.kc_presupuesto = pd.DataFrame(APU_ROWS)
        st.session_state.kc_pres_pais = pais

    c_p1, c_p2 = st.columns([2, 2])
    with c_p1:
        cliente_p = st.text_input("Cliente:", value="Nombre del Cliente", key="pres_cliente")
    with c_p2:
        proyecto_p = st.text_input("Proyecto:", value="Proyecto Nueva Obra", key="pres_proy")

    st.markdown(f"####  Parámetros Base — {pais}")
    i1, i2, i3, i4, i5 = st.columns(5)
    i1.metric("Salario Mínimo (Mes)", f"{moneda} {cfg_apu['smmlv_mes']:,.0f}")
    i2.metric("Auxilios/Adicionales", f"{moneda} {cfg_apu.get('aux_trans', 0):,.0f}")
    i3.metric("Jornal c/prestaciones", f"{moneda} {jd:,.0f}")
    i4.metric("IVA", f"{_iva*100:.0f}%")
    i5.metric("AIU", f"{_aiu*100:.0f}%")
    st.caption(f" Los Precios Unitarios (PU) calculados abajo incluyen Materiales + Mano de Obra + AIU + IVA según normativas de {pais}.")

    st.markdown("Edite cantidades. **TOTAL = CANTIDAD × PU**. Exporte a Excel para el formato presupuestos.")

    edited_pres = st.data_editor(
        st.session_state.kc_presupuesto,
        num_rows="dynamic",
        use_container_width=True,
        key="de_presupuesto",
        column_config={
            "N°":         st.column_config.TextColumn("N°", width="small"),
            "DESCRIPCIÓN":st.column_config.TextColumn("DESCRIPCIÓN", width="large"),
            "UDS":        st.column_config.TextColumn("UDS", width="small"),
            "CANTIDAD":   st.column_config.NumberColumn("CANTIDAD", format="%.2f"),
            "PU":         st.column_config.NumberColumn(f"P.U. ({moneda})", format="%.0f"),
        }
    )

    try:
        _d = edited_pres.copy()
        _tot = _d[_d["PU"] > 0]["CANTIDAD"] * _d[_d["PU"] > 0]["PU"]
        st.metric(f" TOTAL PRESUPUESTO PRELIMINAR ESTIMADO", f"{moneda} {_tot.sum():,.0f}")
    except: pass

    bp1, bp2 = st.columns(2)
    with bp1:
        if st.button(" Guardar cambios", key="btn_sv_pres", use_container_width=True):
            st.session_state.kc_presupuesto = edited_pres
            st.success(" Actualizado")
    with bp2:
        if st.button(" Restaurar APU original", key="btn_reset_pres", use_container_width=True, type="secondary"):
            del st.session_state["kc_presupuesto"]
            del st.session_state["kc_pres_pais"]
            st.rerun()

    st.markdown("---")
    excel_buf = build_excel_presupuesto(edited_pres, cliente_p, proyecto_p)
    st.download_button(
        "⬇ Descargar Presupuesto.xlsx",
        data=excel_buf,
        file_name=f"Presupuesto_{pais}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", use_container_width=True
    )

# ══════════ TAB 18 — RESUMEN Y EXPORTAR ══════════
with tabs[18]:
    st.subheader(" Resumen de Materiales y Exportación")
    if not st.session_state.kc_rows:
        st.info("ℹ Agrega materiales desde cualquier tab de cálculo y aparecerán aquí.")
    else:
        df_res=pd.DataFrame(st.session_state.kc_rows); df_res["subtotal"]=df_res["cant"]*df_res["precio"]
        total_global=df_res["subtotal"].sum()
        col_tbl,col_chart=st.columns([3,2])
        with col_tbl:
            st.markdown("####  Tabla de Materiales")
            df_d=df_res.copy()
            df_d["cant"]=df_d["cant"].apply(lambda x:f"{x:,.2f}"); df_d["precio"]=df_d["precio"].apply(lambda x:f"{moneda} {x:,.2f}")
            df_d["subtotal"]=df_d["subtotal"].apply(lambda x:f"{moneda} {x:,.2f}"); df_d.columns=["Elemento","Unidad","Cantidad","P. Unit.","Subtotal"]
            st.dataframe(df_d,use_container_width=True,height=350); st.metric(" TOTAL ESTIMADO",f"{moneda} {total_global:,.2f}")
        with col_chart:
            st.markdown("####  Distribución de Costos")
            df_grp=df_res.groupby("elemento")["subtotal"].sum().nlargest(10)
            fig_pie=go.Figure(go.Pie(labels=df_grp.index.tolist(),values=df_grp.values.tolist(),hole=0.45,textinfo="label+percent",
                marker=dict(colors=["#1565c0","#1976d2","#2196f3","#42a5f5","#90caf9","#e53935","#ef5350","#ff7043","#ffa726","#4caf50"])))
            fig_pie.update_layout(showlegend=False,height=350,plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",font_color="white")
            st.plotly_chart(fig_pie,use_container_width=True)
        st.markdown("---"); ex1,ex2,ex3=st.columns(3)
        with ex1:
            excel_buf=build_excel_resumen(st.session_state.kc_rows,p,R)
            st.download_button(" Exportar a Excel",excel_buf,file_name=f"Konte_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
        with ex2:
            csv_str=df_res.to_csv(index=False).encode("utf-8")
            st.download_button(" Exportar a CSV",csv_str,file_name=f"Konte_{datetime.datetime.now().strftime('%Y%m%d')}.csv",mime="text/csv",use_container_width=True)
        with ex3:
            if st.button(" Limpiar Resumen",use_container_width=True,type="secondary"):
                st.session_state.kc_rows=[]
                save_state()
                st.rerun()